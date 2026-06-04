#!/usr/bin/env python3
"""Generate SEL-700G Generator Protection Relay Test Results Template (.xlsx)."""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import CellIsRule

# ── Colours ──────────────────────────────────────────────────────────────────
CB  = "1F4E79"   # deep blue  – headers
CLB = "D9E2F3"   # light blue – labels
CY  = "FFF2CC"   # yellow     – tolerance / caution
CCR = "FFF8E7"   # cream      – expected / reference
CPF = "C6EFCE"; CPFF = "006100"   # PASS
CFL = "FFC7CE"; CFFL = "9C0006"   # FAIL
CGR = "D9D9D9"   # grey       – auto-calculated
CAL = "F2F2F2"   # alt row
CUI = "0000FF"   # blue text  – user input
CWH = "FFFFFF"

def _fill(c): return PatternFill("solid", fgColor=c)
def _font(bold=False, color="000000", size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def h1(ws, row, col1, col2, text):
    ws.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    c = ws.cell(row=row, column=col1, value=text)
    c.fill = _fill(CB); c.font = _font(bold=True, color="FFFFFF", size=11)
    c.alignment = _align("left", "center"); c.border = _thin()

def label(ws, row, col, text, bg=CLB, bold=False, cols=1, wrap=False):
    if cols > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+cols-1)
    c = ws.cell(row=row, column=col, value=text)
    c.fill = _fill(bg); c.font = _font(bold=bold)
    c.alignment = _align("left", "center", wrap=wrap); c.border = _thin()

def inp(ws, row, col, value=None, cols=1):
    if cols > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+cols-1)
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(color=CUI); c.alignment = _align("left", "center")
    c.border = _thin(); c.fill = _fill(CWH)

def tol(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(color=CUI); c.alignment = _align("center", "center")
    c.border = _thin(); c.fill = _fill(CY)

def chdr(ws, row, col, text, bg=CB, fg="FFFFFF"):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = _fill(bg); c.font = _font(bold=True, color=fg, size=10)
    c.alignment = _align("center", "center", wrap=True); c.border = _thin()

def auto(ws, row, col, formula_or_value, fmt=None, bold=False):
    c = ws.cell(row=row, column=col, value=formula_or_value)
    c.fill = _fill(CGR); c.font = _font(bold=bold)
    c.alignment = _align("center", "center"); c.border = _thin()
    if fmt: c.number_format = fmt

def note_row(ws, row, col1, col2, text, height=32):
    ws.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    c = ws.cell(row=row, column=col1, value=text)
    c.fill = _fill(CY); c.font = _font(italic=True, size=10)
    c.alignment = _align("left", "center", wrap=True); c.border = _thin()
    ws.row_dimensions[row].height = height

def title_row(ws, row, col1, col2, text):
    ws.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    c = ws.cell(row=row, column=col1, value=text)
    c.fill = _fill(CB); c.font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 26

def blank(ws, row, col, bg=CWH):
    c = ws.cell(row=row, column=col)
    c.fill = _fill(bg); c.border = _thin()


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 – COVER
# ═══════════════════════════════════════════════════════════════════════════════
def build_cover(wb):
    ws = wb.active
    ws.title = "1. Cover"
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [30, 22, 14, 14, 14, 14]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 6, "SEL-700G GENERATOR PROTECTION RELAY — TEST RESULTS TEMPLATE")

    # ── Section 1: Relay ID ──────────────────────────────────────────────────
    r += 2                                          # r = 3
    h1(ws, r, 1, 6, "1.  RELAY IDENTIFICATION")
    rows_id = ["Relay Type / Model", "Relay ID (RID)", "Terminal ID (TID)",
               "Serial Number", "Firmware ID (FID)", "Checksum ID (CID)",
               "Protected Equipment (generator tag)", "Site / Substation",
               "Test Date", "Test Engineer"]
    defaults = ["SEL-700G"] + [""] * 9
    for lbl, dflt in zip(rows_id, defaults):
        r += 1
        label(ws, r, 1, lbl)
        inp(ws, r, 2, dflt, cols=5)

    # ── Section 2: Generator / CT / VT data ─────────────────────────────────
    r += 2                                          # r = 15
    h1(ws, r, 1, 6, "2.  GENERATOR AND CT/VT PARAMETERS")
    r += 1                                          # r = 16  column headers
    for col, hdr in enumerate(["Parameter", "Setting Label", "Value", "Units", "", ""], 1):
        chdr(ws, r, col, hdr)

    # Each tuple: (label, setting, default_value, unit)
    gen_data = [
        ("Generator MVA Rating",                     "MVA / INOM",           "",     "MVA / A nominal"),
        ("Generator Terminal Voltage (L-L)",          "VNOM_X",               "",     "kV (primary)"),
        ("Phase Rotation",                            "PHROT",                "ABC",  "ABC / ACB"),
        ("X-Side (Stator) CT Connection",             "CTCONY / DELTA_X",     "WYE",  "WYE / DELTA"),
        ("X-Side CT Ratio — Primary",                 "CTRX (primary)",       "",     "A primary"),
        ("X-Side CT Ratio — Secondary",               "CTRX (secondary)",     "5",    "A secondary"),
        ("Y-Side (87 Input) CT Ratio — Primary",      "CTRY (primary)",       "",     "A primary"),
        ("Y-Side (87 Input) CT Ratio — Secondary",    "CTRY (secondary)",     "5",    "A secondary"),
        ("Neutral CT Ratio — Primary",                "CTRN (primary)",       "",     "A primary"),
        ("Neutral CT Ratio — Secondary",              "CTRN (secondary)",     "5",    "A secondary"),
        ("Main VT Ratio — Primary",                   "PTRX (primary)",       "",     "V primary"),
        ("Main VT Ratio — Secondary",                 "PTRX (secondary)",     "",     "V secondary"),
        ("VT Connection Type (X-side)",               "DELTA_X",              "WYE",  "WYE / DELTA"),
        ("Differential TAP (X-side)",                 "TAPX",                 "",     "A secondary"),
        ("Differential TAP (Y-side)",                 "TAPY",                 "",     "A secondary"),
    ]
    data_start_row = r + 1   # = 17
    for lbl, slbl, dflt, unit in gen_data:
        r += 1
        label(ws, r, 1, lbl)
        label(ws, r, 2, slbl, bg=CCR)
        inp(ws, r, 3, dflt)
        label(ws, r, 4, unit, bg=CWH)
        blank(ws, r, 5); blank(ws, r, 6)

    # Row numbers for ratios (column C):
    # index 4 = CTRX primary → data_start_row + 4
    ctrx_pri = data_start_row + 4   # 21
    ctrx_sec = data_start_row + 5   # 22
    ctry_pri = data_start_row + 6   # 23
    ctry_sec = data_start_row + 7   # 24
    ctrn_pri = data_start_row + 8   # 25
    ctrn_sec = data_start_row + 9   # 26
    ptrx_pri = data_start_row + 10  # 27
    ptrx_sec = data_start_row + 11  # 28

    # ── Section 3: Tolerances ────────────────────────────────────────────────
    r += 2                                          # r = 33
    h1(ws, r, 1, 6, "3.  TEST TOLERANCES  (yellow cells are editable)")
    tol_rows = {}
    for name, dflt, unit in [
        ("Voltage Magnitude Tolerance",  1.0, "%"),
        ("Current Magnitude Tolerance",  1.0, "%"),
        ("Angle Tolerance",              2.0, "degrees"),
        ("Pickup Tolerance",             5.0, "%"),
        ("Timing Tolerance",             5.0, "%"),
    ]:
        r += 1
        label(ws, r, 1, name)
        tol(ws, r, 2, dflt)
        label(ws, r, 3, unit, bg=CWH)
        blank(ws, r, 4); blank(ws, r, 5); blank(ws, r, 6)
        tol_rows[name] = r
    # r = 34..38

    # ── Section 4: Overall result ────────────────────────────────────────────
    r += 2                                          # r = 40
    h1(ws, r, 1, 6, "4.  OVERALL TEST RESULT")
    r += 1                                          # r = 41  auto
    label(ws, r, 1, "Auto-calculated cross-check")
    auto_formula = (
        "=IF(COUNTIF('4. Analogue Inputs'!M:M,\"FAIL\")"
        "+COUNTIF('7. Element Results'!G:G,\"FAIL\")>0,\"FAIL\","
        "IF(COUNTIF('4. Analogue Inputs'!M:M,\"PASS\")"
        "+COUNTIF('7. Element Results'!G:G,\"PASS\")>0,\"PASS\",\"INCOMPLETE\"))"
    )
    c = ws.cell(row=r, column=2, value=auto_formula)
    c.font = _font(bold=True, size=12); c.alignment = _align("center", "center")
    c.border = _thin(); c.fill = _fill(CWH)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    label(ws, r, 5, "(auto only — see signed result)", bg=CWH, cols=2)

    r += 1                                          # r = 42  signed
    label(ws, r, 1, "SIGNED OVERALL RESULT", bold=True)
    signed_row = r
    c = ws.cell(row=r, column=2, value="PASS")
    c.font = _font(bold=True, size=12, color=CUI); c.alignment = _align("center", "center")
    c.border = _thin(); c.fill = _fill(CY)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    label(ws, r, 5, "PASS / PASS WITH NOTES / FAIL", bg=CWH, cols=2)

    dv = DataValidation(type="list", formula1='"PASS,PASS WITH NOTES,FAIL"')
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=signed_row, column=2))

    return ws, tol_rows, ctrx_pri, ctrx_sec, ctry_pri, ctry_sec, \
           ctrn_pri, ctrn_sec, ptrx_pri, ptrx_sec, signed_row


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 – EQUIPMENT & REFERENCES
# ═══════════════════════════════════════════════════════════════════════════════
def build_equip(wb):
    ws = wb.create_sheet("2. Equip & Refs")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [30, 40, 22, 16]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 4, "EQUIPMENT & REFERENCES")

    r += 2
    h1(ws, r, 1, 4, "TEST EQUIPMENT USED")
    r += 1
    for col, hdr in enumerate(["Equipment", "Manufacturer / Model", "Asset / Serial No.", "Cal. Due"], 1):
        chdr(ws, r, col, hdr)
    for _ in range(6):
        r += 1
        for col in range(1, 5): inp(ws, r, col)

    r += 2
    h1(ws, r, 1, 4, "DOCUMENTS & DRAWINGS")
    r += 1
    for col, hdr in enumerate(["Document", "Title / Description", "Document No.", "Rev"], 1):
        chdr(ws, r, col, hdr)
    for doc in ["Relay Settings File (.sel / AcSELerator QuickSet)",
                "AC Elementary / Schematic",
                "DC Control Schematic",
                "Generator Protection Single-Line Diagram",
                "Protection Coordination Study",
                "Maintenance Work Order"]:
        r += 1
        label(ws, r, 1, doc)
        for col in range(2, 5): inp(ws, r, col)
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 – PRE-TEST
# ═══════════════════════════════════════════════════════════════════════════════
def build_pretest(wb):
    ws = wb.create_sheet("3. Pre-Test")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [40, 32, 16, 22]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 4, "PRE-TEST CHECKS")

    def check_section(ws, r, title, rows):
        r += 2
        h1(ws, r, 1, 4, title)
        r += 1
        for col, hdr in enumerate(["Check", "Expected / Action", "Result", "Notes"], 1):
            chdr(ws, r, col, hdr)
        for chk, exp in rows:
            r += 1
            label(ws, r, 1, chk)
            label(ws, r, 2, exp, bg=CCR)
            inp(ws, r, 3); inp(ws, r, 4)
        return r

    r = check_section(ws, r, "3.1  COMMUNICATIONS & IDENTIFICATION", [
        ("Connect serial cable or Ethernet to relay",
         "Terminal prompt '=' or '-' appears"),
        ("Issue QUIT <Enter> — verify response header",
         "RID and TID match asset register"),
        ("Issue ACC <Enter> + Level 1 password",
         "Access Level 1 granted"),
        ("ENABLED LED illuminated (green, front panel)",
         "LED ON = relay healthy and in service"),
        ("Firmware ID (FID) — record string",
         "Matches approved firmware version"),
        ("Checksum ID (CID) — record string",
         "Matches as-left CID from last test"),
    ])

    # 3.2 STA self-test
    r += 2
    h1(ws, r, 1, 4,
       "3.2  RELAY SELF-TEST STATUS  (issue STATUS or STA <Enter> at serial port or front-panel STATUS)")
    r += 1
    note_row(ws, r, 1, 4,
             "All items should show OK. 'W' prefix = warning (issue STA C to clear). "
             "Latched FAIL = protection disabled — investigate immediately. "
             "HALARM = latched hardware alarm bit. SALARM = pulsed software alarm bit.", 40)
    r += 1
    for col, hdr in enumerate(["Self-Test Item (Table 11.9)",
                                "Normal Range / Description",
                                "Status / Measured Value",
                                "Action if WARN or FAIL"], 1):
        chdr(ws, r, col, hdr)

    # Full Table 11.9 self-test list
    sta_items = [
        ("FPGA — Mainboard (power up)",
         "FPGA accepts program + correct version"),
        ("FPGA — Mainboard (run time)",
         "No data acquisition interrupt loss; CRC OK"),
        ("GPSB — Backplane communications",
         "GPSB not busy on entry to processing interval"),
        ("HMI — Front-panel (power up)",
         "2-line: ID registers match; Touchscreen: diagnostics pass"),
        ("RAM — External (power up)",
         "Read/write test on system RAM"),
        ("RAM — External (run time)",
         "Read/write test on system RAM"),
        ("RAM — Internal (power up)",
         "Read/write test on CPU RAM"),
        ("RAM — Internal (run time)",
         "Read/write test on CPU RAM"),
        ("NON_VOL — Data Flash (power up)",
         "Checksum on critical data"),
        ("NON_VOL — Data Flash (run time)",
         "Checksum on critical data"),
        ("CR_RAM — Critical RAM / settings (power up)",
         "Checksum on active settings copy"),
        ("CR_RAM — Critical RAM (run time)",
         "Instruction matches FLASH image"),
        ("I/O Board — CARD_C / CARD_D / CARD_E",
         "ID register matches part number"),
        ("Slot Z Board (power up)",
         "ID register matches part number"),
        ("Slot Z A/D Offset WARN",
         "Normal: −50 to +50 mV — issue STA C to clear warn"),
        ("Slot E Board (power up)",
         "ID register matches part number"),
        ("Slot E A/D Offset WARN",
         "Normal: −50 to +50 mV — issue STA C to clear warn"),
        ("+0.9 V supply",   "Normal: 0.855 – 0.945 V"),
        ("+1.2 V supply",   "Normal: 1.152 – 1.248 V"),
        ("+1.5 V supply",   "Normal: 1.35 – 1.65 V"),
        ("+1.8 V supply",   "Normal: 1.71 – 1.89 V"),
        ("+2.5 V supply",   "Normal: 2.32 – 2.68 V"),
        ("+3.3 V supply",   "Normal: 3.07 – 3.53 V"),
        ("+3.75 V supply",  "Normal: 3.48 – 4.02 V"),
        ("+5.0 V supply",   "Normal: 4.65 – 5.35 V"),
        ("−1.25 V supply",  "Normal: −1.16 to −1.34 V"),
        ("−5.0 V supply",   "Normal: −4.65 to −5.35 V"),
        ("Clock Battery",   "Normal: 2.3 – 3.5 V — issue STA C to clear warn"),
        ("Clock Chip",      "Clock chip comms + timekeeping — issue STA C to clear warn"),
        ("Clock Chip RAM",  "Clock chip static RAM — issue STA C to clear warn"),
        ("RTD module (External/Internal)", "Comm OK; no open/shorted inputs (if fitted)"),
        ("SEL-2664 Field Ground Module",   "OK/FAIL — fibre comms + module status (if fitted)"),
        ("CID file (configured IED description)", "CID file accessible"),
        ("HALARM relay word bit",  "Should be 0 — not continuously asserted"),
        ("SALARM relay word bit",  "Should be 0 — not pulsed"),
        ("Relay Enabled",          "Output: 'Relay Enabled'"),
    ]
    for item, desc in sta_items:
        r += 1
        label(ws, r, 1, item)
        label(ws, r, 2, desc, bg=CCR)
        inp(ws, r, 3); inp(ws, r, 4)

    r = check_section(ws, r, "3.3  CLOCK / TIME SYNCHRONISATION", [
        ("Issue TIM <Enter> — relay time",         "Within 5 s of reference clock"),
        ("Issue DAT <Enter> — relay date",         "Correct calendar date"),
        ("Time source shown (Internal / IRIG-B)",  "IRIG-B if time-code input connected"),
        ("Synchronise clock if out of tolerance",  "TIME hh:mm:ss   DATE mm/dd/yy"),
    ])

    r = check_section(ws, r, "3.4  EVENT RECORDS (pre-test baseline)", [
        ("Issue HIS <Enter> — review event history",     "No unexplained trips/alarms since last test"),
        ("Issue SER <Enter> — review SER log",           "No unexpected HALARM / SALARM events"),
        ("Issue SUM <Enter> — event summary",            "Review; record most recent event number"),
        ("Issue EVE <Enter> — most recent event report", "Review for unexpected generator trips"),
    ])

    r = check_section(ws, r, "3.5  SETTINGS VERIFICATION", [
        ("Issue SHOW <Enter> (or SHO G/P/L/R) — download settings", "Match approved settings file"),
        ("CTRX (X-side CT ratio) matches Cover sheet",               "As per relay data"),
        ("CTRY (Y-side/87 CT ratio) matches Cover sheet",            "As per relay data"),
        ("CTRN (neutral CT ratio) matches Cover sheet",              "As per relay data"),
        ("PTRX (main VT ratio) matches Cover sheet",                 "As per relay data"),
        ("TAPX / TAPY (differential TAP values)",                    "As per protection study"),
        ("VNOM_X, INOM, PHROT match generator nameplate",            "As per generator data"),
        ("E87, E87N, EREF, E64G, E64F enabling settings",           "Elements enabled per scheme"),
        ("Active setting group correct (Group 1 or as applicable)",  "As per operating procedure"),
        ("Settings unchanged since last test",                       "No unauthorised changes"),
    ])

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 – ANALOGUE INPUTS
# ═══════════════════════════════════════════════════════════════════════════════
def build_analogue(wb):
    ws = wb.create_sheet("4. Analogue Inputs")
    ws.sheet_view.showGridLines = False
    col_widths = [22, 14, 13, 16, 16, 13, 10, 10, 8, 8, 8, 8, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    r = 1
    title_row(ws, r, 1, 13, "ANALOGUE INPUT VERIFICATION — PRIMARY VALUE CHECK")
    r += 1
    note_row(ws, r, 1, 13,
             "Verify the relay reports the correct PRIMARY value (Injected Secondary × CT or VT ratio). "
             "Use METER <Enter> or front-panel METER > Fundamental. "
             "All relay readings are in PRIMARY units (CTRX × secondary, etc.).", 36)

    HDRS = ["Channel", "Injected Sec.\n(A or V)", "Inj. Angle\n(°)",
            "Expected Primary\n(auto)", "Relay Primary\nReading",
            "Relay Angle\nReading", "Mag\nErr %", "Angle\nErr °",
            "Mag\nTol %", "Ang\nTol °", "", "", "P/F"]

    def analogue_block(ws, r, title, ratio_name, tol_mag_name, channels):
        r += 2
        h1(ws, r, 1, 13, title)
        r += 1
        label(ws, r, 1,
              f"Ratio: {ratio_name}  |  Tolerance: {tol_mag_name}  (defined names on Cover sheet)",
              bg=CLB, cols=13)
        r += 1
        for col, hdr in enumerate(HDRS, 1):
            chdr(ws, r, col, hdr)
        for ch in channels:
            r += 1
            label(ws, r, 1, ch, bg=CLB)
            inp(ws, r, 2); inp(ws, r, 3)
            c = ws.cell(row=r, column=4,
                        value=f'=IFERROR(IF(B{r}="","",B{r}*{ratio_name}),"—")')
            c.fill = _fill(CGR); c.font = _font(bold=True)
            c.border = _thin(); c.alignment = _align("center", "center")
            inp(ws, r, 5); inp(ws, r, 6)
            c = ws.cell(row=r, column=7,
                        value=f'=IFERROR(IF(OR(D{r}="",E{r}=""),"",(E{r}-D{r})/D{r}*100),"—")')
            c.fill = _fill(CGR); c.border = _thin()
            c.number_format = "0.00"; c.alignment = _align("center", "center")
            c = ws.cell(row=r, column=8,
                        value=f'=IFERROR(IF(OR(C{r}="",F{r}=""),"",F{r}-C{r}),"—")')
            c.fill = _fill(CGR); c.border = _thin()
            c.number_format = "0.0"; c.alignment = _align("center", "center")
            c = ws.cell(row=r, column=9, value=f"={tol_mag_name}")
            c.fill = _fill(CY); c.border = _thin(); c.alignment = _align("center", "center")
            c = ws.cell(row=r, column=10, value="=TolAngle_deg")
            c.fill = _fill(CY); c.border = _thin(); c.alignment = _align("center", "center")
            blank(ws, r, 11); blank(ws, r, 12)
            pf = (f'=IFERROR(IF(AND(B{r}="",E{r}=""),"—",'
                  f'IF(OR(AND(ISNUMBER(G{r}),ABS(G{r})>{tol_mag_name}),'
                  f'AND(ISNUMBER(H{r}),ABS(H{r})>TolAngle_deg)),"FAIL","PASS")),"—")')
            c = ws.cell(row=r, column=13, value=pf)
            c.font = _font(bold=True); c.border = _thin(); c.alignment = _align("center", "center")
        return r

    r = analogue_block(ws, r,
        "4.1  X-SIDE PHASE CURRENTS  (IAX, IBX, ICX — stator CTs in Slot Z)  [CTRX]",
        "Ratio_CTRX", "TolI_pct",
        ["IAX — A-Phase X-side", "IBX — B-Phase X-side", "ICX — C-Phase X-side"])

    r = analogue_block(ws, r,
        "4.2  Y-SIDE PHASE CURRENTS  (IAY, IBY, ICY — 87 input CTs in Slot E)  [CTRY]",
        "Ratio_CTRY", "TolI_pct",
        ["IAY — A-Phase Y-side (87 CT)", "IBY — B-Phase Y-side (87 CT)", "ICY — C-Phase Y-side (87 CT)"])

    r = analogue_block(ws, r,
        "4.3  NEUTRAL CURRENT  (IN — neutral CT in Slot Z)  [CTRN]  (if neutral CT fitted)",
        "Ratio_CTRN", "TolI_pct",
        ["IN — Generator neutral current"])

    r = analogue_block(ws, r,
        "4.4  VOLTAGE INPUTS  (VAX, VBX, VCX — X-side VTs)  [PTRX]  (if VT card fitted)",
        "Ratio_PTRX", "TolV_pct",
        ["VAX — Phase A voltage (X-side)", "VBX — Phase B voltage (X-side)", "VCX — Phase C voltage (X-side)"])

    # 4.5 Differential meter check
    r += 2
    h1(ws, r, 1, 13,
       "4.5  DIFFERENTIAL METER CHECK  (front-panel METER > Differential  or  EVE DIF <Enter>)")
    r += 1
    note_row(ws, r, 1, 13,
             "Under balanced load: IOP (operate) ≈ 0.00 TAP, IRT (restraint) ≈ load current in TAP units. "
             "If IOP/IRT > 0.10, investigate CT polarity, TAPX/TAPY balance, and CT connections. "
             "Check also 2nd-harmonic content (I1F2) — should be < 10% during normal load.", 44)
    r += 1
    for col, hdr in enumerate(["Diff Element", "IOP\n(Operate, TAP)", "IRT\n(Restraint, TAP)",
                                "MM = IOP/IRT", "Accept if\nMM ≤ 0.10",
                                "2nd Harm\nI1F2 (TAP)", "5th Harm\nI1F5 (TAP)",
                                "", "", "", "", "", "Result"], 1):
        chdr(ws, r, col, hdr)
    for elem in ["87-1 (DIFF1)", "87-2 (DIFF2) if used", "87-3 (DIFF3) if used"]:
        r += 1
        label(ws, r, 1, elem, bg=CLB)
        inp(ws, r, 2); inp(ws, r, 3)
        c = ws.cell(row=r, column=4,
                    value=f'=IFERROR(IF(OR(B{r}="",C{r}=""),"",B{r}/C{r}),"—")')
        c.fill = _fill(CGR); c.border = _thin()
        c.number_format = "0.000"; c.alignment = _align("center", "center")
        label(ws, r, 5, "≤ 0.10", bg=CCR)
        inp(ws, r, 6); inp(ws, r, 7)
        blank(ws, r, 8); blank(ws, r, 9); blank(ws, r, 10); blank(ws, r, 11); blank(ws, r, 12)
        c = ws.cell(row=r, column=13,
                    value=f'=IFERROR(IF(D{r}="","—",IF(AND(ISNUMBER(D{r}),D{r}<=0.10),"PASS","FAIL")),"—")')
        c.font = _font(bold=True); c.border = _thin(); c.alignment = _align("center", "center")

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 – BINARY I/O
# ═══════════════════════════════════════════════════════════════════════════════
def build_binary(wb):
    ws = wb.create_sheet("5. Binary IO")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", [10, 30, 22, 24, 12]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 5, "BINARY I/O VERIFICATION")

    r += 2
    h1(ws, r, 1, 5, "5.1  DIGITAL INPUTS — Pickup Verification")
    r += 1
    note_row(ws, r, 1, 5,
             "Apply rated control voltage to each input. Use front-panel MAIN > Targets > Row 49 "
             "or serial TARGET command. Verify IN101/IN102 state changes 0→1 when voltage applied.", 30)
    r += 1
    for col, hdr in enumerate(["Input No.", "Function Label", "Control Voltage Applied", "Relay Word Bit", "Result"], 1):
        chdr(ws, r, col, hdr)
    inputs = [
        ("IN101", "Digital Input 1 (Slot A) — edit to match schematics", "IN101"),
        ("IN102", "Digital Input 2 (Slot A) — edit to match schematics", "IN102"),
        ("IN301", "Optional DI — Slot C Input 1 (if card fitted)",         "IN301"),
        ("IN302", "Optional DI — Slot C Input 2 (if card fitted)",         "IN302"),
        ("IN303", "Optional DI — Slot C Input 3 (if card fitted)",         "IN303"),
        ("IN304", "Optional DI — Slot C Input 4 (if card fitted)",         "IN304"),
    ]
    for inp_id, func, rwb in inputs:
        r += 1
        label(ws, r, 1, inp_id, bg=CLB, bold=True)
        inp(ws, r, 2, func)
        inp(ws, r, 3)
        label(ws, r, 4, rwb, bg=CCR)
        inp(ws, r, 5)

    r += 2
    h1(ws, r, 1, 5, "5.2  DIGITAL OUTPUTS — Continuity Test")
    r += 1
    note_row(ws, r, 1, 5,
             "Force each output via serial (OUT101:=1 etc.) or front-panel CONTROL menu. "
             "Verify continuity through wiring to trip coil / alarm annunciator / close circuit. "
             "Restore all SELOGIC equations after test.", 30)
    r += 1
    for col, hdr in enumerate(["Output No.", "Function Label", "Force Command", "Expected Result", "P/F"], 1):
        chdr(ws, r, col, hdr)
    outputs = [
        ("OUT101", "TRIP / Main trip (edit label)",        "OUT101:=1", "Contact closes — continuity to trip coil"),
        ("OUT102", "Alarm / Auxiliary (edit label)",       "OUT102:=1", "Contact closes — continuity verified"),
        ("OUT103", "Close / Auxiliary (edit label)",       "OUT103:=1", "Contact closes — continuity verified"),
        ("OUT301", "Optional DO — Slot C 1 (if fitted)",   "OUT301:=1", "Contact closes — continuity verified"),
        ("OUT302", "Optional DO — Slot C 2 (if fitted)",   "OUT302:=1", "Contact closes — continuity verified"),
    ]
    for out_id, func, cmd, exp in outputs:
        r += 1
        label(ws, r, 1, out_id, bg=CLB, bold=True)
        inp(ws, r, 2, func)
        label(ws, r, 3, cmd, bg=CCR)
        label(ws, r, 4, exp, bg=CCR)
        inp(ws, r, 5)

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6 – ELEMENT TRACKER
# ═══════════════════════════════════════════════════════════════════════════════
def build_tracker(wb):
    ws = wb.create_sheet("6. Element Tracker")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [6, 18, 42, 14, 16, 12]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 6, "PROTECTION ELEMENT TRACKER")
    r += 2
    for col, hdr in enumerate(["#", "Element ID", "Description", "Enabled\n(Y / N / NA)", "Test Result Row", "Result"], 1):
        chdr(ws, r, col, hdr)

    elements = [
        # Differential
        ("87-1",    "Phase Differential restrained — element 1  (87AP × TAPX)"),
        ("87-1 U",  "Phase Differential unrestrained — element 1  (U87P × TAPX)"),
        ("87-2",    "Phase Differential — element 2  (if E87 enables 2nd element)"),
        ("87-3",    "Phase Differential — element 3  (if E87 enables 3rd element)"),
        ("87N",     "Ground Differential  (87NAP, 87NBP)"),
        ("REF",     "Restricted Earth Fault  (50REF1P, REF1TC) — if EREF := Y"),
        # Stator/field ground
        ("64G-1",   "100% Stator Ground — 27TH element 1  (64G1P, 64G1D) — if E64G := Y"),
        ("64G-2",   "100% Stator Ground — 27TH element 2  (64G2P, 64G2D) — if E64G := Y"),
        ("64F",     "Field Ground  (via SEL-2664 module) — if E64F := Y"),
        # Loss-of-field
        ("40Z1",    "Loss of Field — zone 1  (40Z1P, 40Z1D)"),
        ("40Z2",    "Loss of Field — zone 2  (40Z2P, 40Z2D)"),
        # Thermal
        ("49T",     "Thermal Overload  (49TP, 49TAP, 49TTC)"),
        ("49RTD",   "RTD Thermal Protection  (RTDT, RTDFLT) — if E49RTD fitted"),
        # Current unbalance / negative sequence
        ("46",      "Current Unbalance / Negative-Sequence  (46Q1P, 46Q2P, 46Q1D)"),
        # Volts/Hz
        ("24D1",    "Volts/Hz Definite Time 1  (24D1P, 24D1D)"),
        ("24D2",    "Volts/Hz Definite Time 2  (24D2P, 24D2D)"),
        ("24C",     "Volts/Hz Inverse Time  (24IP, 24IC)"),
        # Out of step / vector shift
        ("78",      "Out of Step  (78P — loss-of-synchronism)"),
        ("78VS",    "Vector Shift  (78VSP — delta-angle loss-of-synchronism)"),
        # Inadvertent energisation / backup
        ("INAD",    "Inadvertent Energisation  (EINAD — 50/27 logic)"),
        ("21C",     "Compensator Distance (backup)  (21CX)"),
        ("51C",     "Voltage-Controlled Time OC  (51CP, 51CV)"),
        ("51V",     "Voltage-Restrained Time OC  (51VP, 51VTC)"),
        # Phase OC
        ("50PX1",   "X-side Phase Instantaneous OC — Stage 1  (50PX1P)"),
        ("50PX2",   "X-side Phase Instantaneous OC — Stage 2  (50PX2P)"),
        ("51PX",    "X-side Phase Time OC  (51PXP, 51PXTC)"),
        # Ground OC
        ("50GX1",   "X-side Ground Instantaneous OC — Stage 1  (50GX1P)"),
        ("51GX",    "X-side Ground Time OC  (51GXP, 51GXTC)"),
        # Negative-sequence OC
        ("50QX1",   "X-side Neg-Seq Instantaneous OC — Stage 1  (50QX1P)"),
        ("51QX",    "X-side Neg-Seq Time OC  (51QXP, 51QXTC)"),
        # Neutral OC
        ("50N",     "Neutral Instantaneous OC  (50NP, uses 67NnP relay word bits)"),
        ("51N",     "Neutral Time OC  (51NP, 51NTC)"),
        ("67N",     "Directional Neutral OC  (67NnP, if EDIRX := Y)"),
        # Voltage
        ("27X1",    "Undervoltage — element 1  (27X1P, 27X1T) — requires VT card"),
        ("27X2",    "Undervoltage — element 2  (27X2P, 27X2T) — requires VT card"),
        ("27I",     "Inverse-Time Undervoltage  (E27I, 27V1X) — requires VT card"),
        ("59X1",    "Overvoltage P — element 1  (59X1P, 59X1T) — requires VT card"),
        ("59X2",    "Overvoltage Q — element 2  (59X2P, 59X2T) — requires VT card"),
        ("59X3",    "Overvoltage G — element 3 (zero-seq) — requires VT card"),
        ("59I",     "Inverse-Time Overvoltage  (E59IX) — requires VT card"),
        # Power / freq
        ("32X",     "Directional Power (reverse/low-forward)  (32XP, 32XPWR)"),
        ("81X1",    "Over/Underfrequency — element 1  (81X1P, 81X1T) — requires VT card"),
        ("81X2",    "Over/Underfrequency — element 2  (81X2P, 81X2T) — requires VT card"),
        ("81RX",    "Rate-of-Change of Frequency  (81RX, E81RX) — requires VT card"),
        # Breaker
        ("BFX",     "Breaker Failure — X-side  (BFIX, BFTX, BFDX)"),
        ("60LOPX",  "Loss of Potential  (LOPBLKX, LOPPUX) — requires VT card"),
        ("25X",     "Synchronism Check — X-side  (25C, 25AX1, 25AX2) — if E25X := Y"),
    ]

    for i, (eid, desc) in enumerate(elements, 1):
        r += 1
        bg = CAL if i % 2 == 0 else CWH
        c = ws.cell(row=r, column=1, value=i)
        c.font = _font(bold=True); c.border = _thin(); c.fill = _fill(bg)
        c.alignment = _align("center", "center")
        label(ws, r, 2, eid, bg=CLB, bold=True)
        label(ws, r, 3, desc, bg=bg)
        inp(ws, r, 4); inp(ws, r, 5); inp(ws, r, 6)

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 7 – ELEMENT RESULTS
# ═══════════════════════════════════════════════════════════════════════════════
def build_results(wb):
    ws = wb.create_sheet("7. Element Results")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFG", [12, 38, 14, 14, 12, 14, 10]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 7, "PROTECTION ELEMENT TEST RESULTS")
    r += 1
    note_row(ws, r, 1, 7,
             "Setting column = relay pickup setting (secondary A, V, TAP pu, or %). "
             "Measured Pickup = value at which relay output asserted. "
             "Pickup Err % = auto-calculated. Result = PASS if |Err| ≤ TolPickup_pct.", 32)
    r += 1
    for col, hdr in enumerate(["Element", "Description",
                                "Pickup Setting\n(sec A, V, or TAP pu)",
                                "Measured\nPickup",
                                "Pickup\nErr %",
                                "Measured\nTime (s)", "Result"], 1):
        chdr(ws, r, col, hdr)

    test_elements = [
        ("87-1 R",   "Differential restrained pickup — 87AP × TAPX  (TAP pu)"),
        ("87-1 U",   "Differential unrestrained pickup — U87P × TAPX  (TAP pu)"),
        ("87N",      "Ground differential pickup — 87NAP  (TAP pu)"),
        ("REF",      "Restricted Earth Fault pickup — 50REF1P  (A secondary, if enabled)"),
        ("64G-1",    "Stator ground 27TH — element 1 pickup  (V secondary)"),
        ("64G-2",    "Stator ground 27TH — element 2 pickup  (V secondary)"),
        ("40Z1",     "Loss of Field — zone 1 pickup  (% Z secondary)"),
        ("40Z2",     "Loss of Field — zone 2 pickup  (% Z secondary)"),
        ("49T",      "Thermal Overload pickup — 49TP  (A secondary)"),
        ("46",       "Current Unbalance — 46Q1P  (A negative-sequence secondary)"),
        ("24D1",     "Volts/Hz — 24D1P  (%, i.e. V/Hz relative to rated)"),
        ("24C",      "Volts/Hz inverse — 24IP  (%)"),
        ("78",       "Out of Step — 78P  (% Z secondary)"),
        ("78VS",     "Vector Shift — 78VSP  (degrees)"),
        ("INAD",     "Inadvertent Energisation — 50PXP  (A secondary)"),
        ("21C",      "Compensator Distance — 21CX  (% Z secondary)"),
        ("51C",      "Voltage-Controlled TOC pickup — 51CP  (A secondary)"),
        ("51V",      "Voltage-Restrained TOC pickup — 51VP  (A secondary)"),
        ("50PX1",    "X-side Phase Inst OC — Stage 1  (A secondary)"),
        ("50PX2",    "X-side Phase Inst OC — Stage 2  (A secondary)"),
        ("51PX",     "X-side Phase TOC pickup  (A secondary)"),
        ("50GX1",    "X-side Ground Inst OC — Stage 1  (A secondary)"),
        ("51GX",     "X-side Ground TOC pickup  (A secondary)"),
        ("50QX1",    "X-side Neg-Seq Inst OC — Stage 1  (A secondary)"),
        ("51QX",     "X-side Neg-Seq TOC pickup  (A secondary)"),
        ("50N",      "Neutral Instantaneous OC — 50NP  (A secondary)"),
        ("51N",      "Neutral TOC pickup — 51NP  (A secondary)"),
        ("27X1",     "Undervoltage — element 1  (V secondary)"),
        ("27X2",     "Undervoltage — element 2  (V secondary)"),
        ("59X1",     "Overvoltage — element 1  (V secondary)"),
        ("32X",      "Directional Power — 32XPWR  (kVA secondary)"),
        ("81X1",     "Frequency — element 1  (Hz)"),
        ("81X2",     "Frequency — element 2  (Hz)"),
        ("81RX",     "Rate-of-Change of Frequency  (Hz/s)"),
        ("BFX",      "Breaker Failure — BFDX timer  (s)"),
        ("25X",      "Synchronism Check — 25C closing window  (degrees)"),
    ]

    for i, (eid, desc) in enumerate(test_elements, 1):
        r += 1
        bg = CAL if i % 2 == 0 else CWH
        label(ws, r, 1, eid, bg=CLB, bold=True)
        label(ws, r, 2, desc, bg=bg)
        inp(ws, r, 3); inp(ws, r, 4)
        c = ws.cell(row=r, column=5,
                    value=f'=IFERROR(IF(OR(C{r}="",D{r}=""),"",(D{r}-C{r})/C{r}*100),"—")')
        c.fill = _fill(CGR); c.border = _thin()
        c.number_format = "0.00"; c.alignment = _align("center", "center")
        inp(ws, r, 6)
        result = (f'=IFERROR(IF(AND(C{r}="",D{r}=""),"—",'
                  f'IF(AND(ISNUMBER(E{r}),ABS(E{r})<=TolPickup_pct),"PASS","FAIL")),"—")')
        c = ws.cell(row=r, column=7, value=result)
        c.font = _font(bold=True); c.border = _thin(); c.alignment = _align("center", "center")

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 8 – POST-TEST
# ═══════════════════════════════════════════════════════════════════════════════
def build_posttest(wb):
    ws = wb.create_sheet("8. Post-Test")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [40, 30, 14, 20]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 4, "POST-TEST CHECKS")

    def check_section(ws, r, title, rows):
        r += 2
        h1(ws, r, 1, 4, title)
        r += 1
        for col, hdr in enumerate(["Check", "Expected / Action", "Result", "Notes"], 1):
            chdr(ws, r, col, hdr)
        for chk, exp in rows:
            r += 1
            label(ws, r, 1, chk)
            label(ws, r, 2, exp, bg=CCR)
            inp(ws, r, 3); inp(ws, r, 4)
        return r

    r = check_section(ws, r, "8.1  SETTINGS RESTORE & VERIFICATION", [
        ("Issue SHOW — verify settings unchanged",             "Match approved settings file"),
        ("All test-forced SELOGIC equations restored",         "OUT101:=, SV variables correct"),
        ("Active setting group restored to normal",            "Group 1 or as applicable"),
        ("ENABLED LED illuminated (green)",                    "Relay in service"),
    ])

    r = check_section(ws, r, "8.2  FINAL SELF-TEST  (STATUS / STA Command)", [
        ("Issue STA <Enter> — all items OK",           "No WARN or FAIL items"),
        ("HALARM = 0 (no latched hardware alarm)",     "Relay word bit confirmed 0"),
        ("SALARM = 0 (no software alarm pulsing)",     "Normal operating condition"),
        ("Relay Enabled confirmed",                    "'Relay Enabled' in STA output"),
    ])

    # 8.3 Clear data buffers
    r += 2
    h1(ws, r, 1, 4, "8.3  CLEAR TEST DATA BUFFERS")
    r += 1
    note_row(ws, r, 1, 4,
             "Issue the serial commands below to clear data generated during testing. "
             "This prevents test data from contaminating operational event records.", 30)
    r += 1
    for col, hdr in enumerate(["Serial Command", "Action Performed", "Completed", "Notes"], 1):
        chdr(ws, r, col, hdr)
    for cmd, action in [
        ("SUM R <Enter>", "Reset event report and Summary Command buffers"),
        ("SER R <Enter>", "Reset Sequential Events Record buffer"),
        ("LDP C <Enter>", "Clear Load Profile Data"),
        ("HIS C <Enter>", "Clear event history (if required by procedure)"),
        ("GSH R <Enter>", "Reset Generator Autosynchronism Report (if synchronism check equipped)"),
    ]:
        r += 1
        label(ws, r, 1, cmd, bg=CLB, bold=True)
        label(ws, r, 2, action, bg=CCR)
        inp(ws, r, 3); inp(ws, r, 4)

    r = check_section(ws, r, "8.4  PLANT RESTORATION", [
        ("Remove test connections from CT / VT terminals",         "Field CTs and VTs reconnected"),
        ("Restore control voltage wiring",                         "All control circuits normal"),
        ("Remove CT secondary shorting links",                     "CT circuits not open-circuited"),
        ("Verify relay measuring load current — METER command",    "Compare to substation metering"),
        ("METER DIF check — IOP/IRT ≤ 0.10 under load",           "Differential balance confirmed"),
        ("SEL-2664 Field Ground Module Rf reading recorded",       "Normal: 20 000 kΩ (healthy field)"),
        ("Cancel all outstanding LOTO permits",                    "All permits cleared"),
        ("Notify operations — relay returned to service",          "Operations informed"),
    ])

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 9 – FILES & SIGN-OFF
# ═══════════════════════════════════════════════════════════════════════════════
def build_signoff(wb, signed_row):
    ws = wb.create_sheet("9. Files & Sign-Off")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [28, 38, 20, 16]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 4, "FILES INDEX & SIGN-OFF")

    r += 2
    h1(ws, r, 1, 4, "OVERALL TEST RESULT (from Cover sheet)")
    r += 1
    label(ws, r, 1, "Signed Overall Result", bold=True)
    c = ws.cell(row=r, column=2, value=f"='1. Cover'!B{signed_row}")
    c.font = _font(bold=True, size=13); c.alignment = _align("center", "center")
    c.border = _thin(); c.fill = _fill(CY)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)

    r += 2
    h1(ws, r, 1, 4, "TEST ARTEFACTS INDEX")
    r += 1
    for col, hdr in enumerate(["Artefact", "File Name / SharePoint Location", "Date Saved", "Notes"], 1):
        chdr(ws, r, col, hdr)
    for art in ["Relay settings file (.sel / AcSELerator QuickSet project)",
                "Pre-test event records (HIS, SER, EVE, SUM)",
                "Post-test event records",
                "Completed results template (this file)",
                "Oscillography / disturbance records  (EVE D)",
                "Generator Autosynchronism Reports  (CGSR / GSH)",
                "Photographs / site notes"]:
        r += 1
        label(ws, r, 1, art)
        inp(ws, r, 2); inp(ws, r, 3); inp(ws, r, 4)

    r += 2
    h1(ws, r, 1, 4, "DEFECTS / FINDINGS LOG")
    r += 1
    for col, hdr in enumerate(["#", "Defect Description", "Action Taken / Work Order", "Closed"], 1):
        chdr(ws, r, col, hdr)
    for i in range(1, 7):
        r += 1
        c = ws.cell(row=r, column=1, value=i)
        c.border = _thin(); c.alignment = _align("center", "center")
        for col in range(2, 5): inp(ws, r, col)

    r += 2
    h1(ws, r, 1, 4, "SIGN-OFF")
    r += 1
    for col, hdr in enumerate(["Role", "Name (print)", "Signature", "Date"], 1):
        chdr(ws, r, col, hdr)
    for role in ["Test Engineer", "Witness (Operations)", "Approving Engineer"]:
        r += 1
        ws.row_dimensions[r].height = 28
        label(ws, r, 1, role, bg=CLB, bold=True)
        inp(ws, r, 2); inp(ws, r, 3); inp(ws, r, 4)

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# DEFINED NAMES
# ═══════════════════════════════════════════════════════════════════════════════
def add_names(wb, ctrx_pri, ctrx_sec, ctry_pri, ctry_sec,
              ctrn_pri, ctrn_sec, ptrx_pri, ptrx_sec, tol_rows):
    sn = "'1. Cover'"
    defs = {
        "Ratio_CTRX":     f"{sn}!$C${ctrx_pri}/{sn}!$C${ctrx_sec}",
        "Ratio_CTRY":     f"{sn}!$C${ctry_pri}/{sn}!$C${ctry_sec}",
        "Ratio_CTRN":     f"{sn}!$C${ctrn_pri}/{sn}!$C${ctrn_sec}",
        "Ratio_PTRX":     f"{sn}!$C${ptrx_pri}/{sn}!$C${ptrx_sec}",
        "TolV_pct":       f"{sn}!$B${tol_rows['Voltage Magnitude Tolerance']}",
        "TolI_pct":       f"{sn}!$B${tol_rows['Current Magnitude Tolerance']}",
        "TolAngle_deg":   f"{sn}!$B${tol_rows['Angle Tolerance']}",
        "TolPickup_pct":  f"{sn}!$B${tol_rows['Pickup Tolerance']}",
        "TolTime_pct":    f"{sn}!$B${tol_rows['Timing Tolerance']}",
    }
    for name, attr in defs.items():
        dn = DefinedName(name=name, attr_text=attr)
        wb.defined_names[name] = dn


# ═══════════════════════════════════════════════════════════════════════════════
# CONDITIONAL FORMATTING
# ═══════════════════════════════════════════════════════════════════════════════
def add_cf(wb):
    pf = PatternFill("solid", fgColor=CPF)
    ff = PatternFill("solid", fgColor=CFL)
    pfont = Font(bold=True, color=CPFF, name="Calibri")
    ffont = Font(bold=True, color=CFFL, name="Calibri")
    for sheet_name, rng in [("4. Analogue Inputs", "M1:M300"),
                             ("7. Element Results",  "G1:G200"),
                             ("6. Element Tracker",  "F1:F200")]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.conditional_formatting.add(rng, CellIsRule("equal", ['"PASS"'], fill=pf, font=pfont))
            ws.conditional_formatting.add(rng, CellIsRule("equal", ['"FAIL"'], fill=ff, font=ffont))


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()

    (cover_ws, tol_rows,
     ctrx_pri, ctrx_sec, ctry_pri, ctry_sec,
     ctrn_pri, ctrn_sec, ptrx_pri, ptrx_sec,
     signed_row) = build_cover(wb)

    build_equip(wb)
    build_pretest(wb)
    build_analogue(wb)
    build_binary(wb)
    build_tracker(wb)
    build_results(wb)
    build_posttest(wb)
    build_signoff(wb, signed_row)

    add_names(wb, ctrx_pri, ctrx_sec, ctry_pri, ctry_sec,
              ctrn_pri, ctrn_sec, ptrx_pri, ptrx_sec, tol_rows)
    add_cf(wb)

    for ws in wb.worksheets:
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = "portrait"
        ws.page_margins.left = 0.4
        ws.page_margins.right = 0.4
        ws.page_margins.top = 0.6
        ws.page_margins.bottom = 0.6

    out = "/home/michael/claudeProject /protection-relay-checksheet/SEL-700G_Test_Results_Template.xlsx"
    wb.save(out)
    print(f"Saved: {out}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Signed result row: {signed_row}")
    print(f"CTRX rows: C{ctrx_pri} / C{ctrx_sec}")
    print(f"CTRY rows: C{ctry_pri} / C{ctry_sec}")
    print(f"Defined names: {list(wb.defined_names.keys())}")


if __name__ == "__main__":
    main()
