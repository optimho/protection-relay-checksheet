#!/usr/bin/env python3
"""Generate ABB REG670 Generator Protection IED Test Results Template (.xlsx).
   Formatting matches generate_RET670_template.py exactly.
   Replaces the earlier REG670 xlsx that was built with a different code style.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import CellIsRule

# ── Colours (identical to RET670 / SEL-787 family) ───────────────────────────
CB  = "1F4E79"; CLB = "D9E2F3"; CY  = "FFF2CC"; CCR = "FFF8E7"
CPF = "C6EFCE"; CPFF = "006100"
CFL = "FFC7CE"; CFFL = "9C0006"
CGR = "D9D9D9"; CAL = "F2F2F2"; CUI = "0000FF"; CWH = "FFFFFF"

def _fill(c): return PatternFill("solid", fgColor=c)
def _font(bold=False, color="000000", size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def h1(ws, row, col1, col2, text):
    ws.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    c = ws.cell(row=row, column=col1, value=text)
    c.fill = _fill(CB); c.font = _font(bold=True, color="FFFFFF", size=11)
    c.alignment = _align("left", "center"); c.border = _thin()

def label(ws, row, col, text, bg=CLB, bold=False, cols=1, wrap=False):
    if cols > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+cols-1)
    c = ws.cell(row=row, column=col, value=text)
    c.fill = _fill(bg); c.font = _font(bold=bold)
    c.alignment = _align("left", "center", wrap=wrap); c.border = _thin()

def inp(ws, row, col, value=None, cols=1):
    if cols > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+cols-1)
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(color=CUI); c.alignment = _align("left", "center")
    c.border = _thin(); c.fill = _fill(CWH)

def tol(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(color=CUI); c.alignment = _align("center", "center")
    c.border = _thin(); c.fill = _fill(CY)

def chdr(ws, row, col, text, bg=CB, fg="FFFFFF"):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = _fill(bg); c.font = _font(bold=True, color=fg, size=10)
    c.alignment = _align("center", "center", wrap=True); c.border = _thin()

def note_row(ws, row, col1, col2, text, height=32):
    ws.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    c = ws.cell(row=row, column=col1, value=text)
    c.fill = _fill(CY); c.font = _font(italic=True, size=10)
    c.alignment = _align("left", "center", wrap=True); c.border = _thin()
    ws.row_dimensions[row].height = height

def title_row(ws, row, col1, col2, text):
    ws.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    c = ws.cell(row=row, column=col1, value=text)
    c.fill = _fill(CB); c.font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 26

def blank(ws, row, col, bg=CWH):
    c = ws.cell(row=row, column=col)
    c.fill = _fill(bg); c.border = _thin()


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 – COVER
# ═══════════════════════════════════════════════════════════════════════════════
def build_cover(wb):
    ws = wb.active
    ws.title = "1. Cover"
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [32, 22, 14, 14, 14, 14]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 6, "ABB REG670 GENERATOR PROTECTION IED — TEST RESULTS TEMPLATE")

    # ── Section 1: IED Identification ────────────────────────────────────────
    r += 2                                              # r = 3
    h1(ws, r, 1, 6, "1.  IED IDENTIFICATION")
    id_rows = ["Relay Type / Model", "IED Name (Station / Bay)",
               "Description (Protected Equipment)", "Serial Number",
               "Ordering Number", "Firmware Version",
               "Production Date", "PCM600 Project File Name (.pcm)",
               "Site / Substation", "Test Date / Test Engineer"]
    defaults = ["REG670 v2.0 ANSI"] + [""] * 9
    for lbl_text, dflt in zip(id_rows, defaults):
        r += 1
        label(ws, r, 1, lbl_text)
        inp(ws, r, 2, dflt, cols=5)

    # ── Section 2: Generator / CT / VT parameters ────────────────────────────
    r += 2                                              # r = 15
    h1(ws, r, 1, 6, "2.  GENERATOR AND CT/VT PARAMETERS  (from PCM600 / TRM settings)")
    r += 1                                              # r = 16  column headers
    for col, hdr in enumerate(["Parameter", "PCM600 Setting", "Value", "Units", "", ""], 1):
        chdr(ws, r, col, hdr)

    gen_data = [
        # (label,                              setting,                  default,    unit)
        ("Generator MVA Rating",               "Sr / MVA",               "",         "MVA"),
        ("Generator Terminal Voltage (L-L)",   "GBASVAL Ur",             "",         "kV primary (L-L)"),
        ("Stator CT Ratio — Primary",          "CTprim (stator / X)",    "",         "A primary"),
        ("Stator CT Ratio — Secondary",        "CTsec (stator / X)",     "1",        "A secondary"),
        ("Y-Side (87G) CT Ratio — Primary",    "CTprim (Y-side / 87G)",  "",         "A primary"),
        ("Y-Side (87G) CT Ratio — Secondary",  "CTsec (Y-side / 87G)",   "1",        "A secondary"),
        ("Neutral CT Ratio — Primary",         "CTprim (neutral IN)",     "",         "A primary"),
        ("Neutral CT Ratio — Secondary",       "CTsec (neutral IN)",      "1",        "A secondary"),
        ("Main VT Primary Voltage",            "VTprim main (kV)",        "",         "kV primary (L-L)"),
        ("Main VT Secondary Voltage",          "VTsec main (V)",          "",         "V secondary (L-L)"),
        ("Neutral VT Primary Voltage",         "VTprim neutral (kV)",     "",         "kV primary (L-G)"),
        ("Neutral VT Secondary Voltage",       "VTsec neutral (V)",       "",         "V secondary (L-G)"),
        ("Stator CT Connection Direction",     "CT_WyePoint",             "ToObject", "ToObject / FromObject"),
        ("Active Setting Group",               "ACTVGRP",                 "1",        "1–6"),
    ]
    data_start_row = r + 1   # = 17
    for lbl_text, sname, dflt, unit in gen_data:
        r += 1
        label(ws, r, 1, lbl_text)
        label(ws, r, 2, sname, bg=CCR)
        inp(ws, r, 3, dflt)
        label(ws, r, 4, unit, bg=CWH)
        blank(ws, r, 5); blank(ws, r, 6)
    # r = 30 after 14 data rows

    # Row numbers for defined names (column C = col 3):
    ct_sta_pri = data_start_row + 2   # row 19  — stator CT primary
    ct_sta_sec = data_start_row + 3   # row 20
    ct_Y_pri   = data_start_row + 4   # row 21  — Y-side CT primary
    ct_Y_sec   = data_start_row + 5   # row 22
    ct_neut_pri = data_start_row + 6  # row 23  — neutral CT primary
    ct_neut_sec = data_start_row + 7  # row 24
    vt_main_kV  = data_start_row + 8  # row 25  — main VT primary (kV)
    vt_main_V   = data_start_row + 9  # row 26  — main VT secondary (V)
    vt_neut_kV  = data_start_row + 10 # row 27  — neutral VT primary (kV)
    vt_neut_V   = data_start_row + 11 # row 28  — neutral VT secondary (V)

    # ── Section 3: Tolerances ─────────────────────────────────────────────────
    r += 2                                              # r = 32
    h1(ws, r, 1, 6, "3.  TEST TOLERANCES  (yellow cells are editable)")
    tol_rows = {}
    for name, dflt, unit in [
        ("Voltage Magnitude Tolerance",  1.0, "%"),
        ("Current Magnitude Tolerance",  1.0, "%"),
        ("Angle Tolerance",              2.0, "degrees"),
        ("Pickup Tolerance",             5.0, "%"),
        ("Timing Tolerance",             5.0, "%"),
    ]:
        r += 1
        label(ws, r, 1, name)
        tol(ws, r, 2, dflt)
        label(ws, r, 3, unit, bg=CWH)
        blank(ws, r, 4); blank(ws, r, 5); blank(ws, r, 6)
        tol_rows[name] = r
    # r = 33..37

    # ── Section 4: Overall result ─────────────────────────────────────────────
    r += 2                                              # r = 39
    h1(ws, r, 1, 6, "4.  OVERALL TEST RESULT")
    r += 1                                              # r = 40  auto
    label(ws, r, 1, "Auto-calculated cross-check")
    auto_formula = (
        "=IF(COUNTIF('4. Analogue Inputs'!M:M,\"FAIL\")"
        "+COUNTIF('7. Element Results'!G:G,\"FAIL\")>0,\"FAIL\","
        "IF(COUNTIF('4. Analogue Inputs'!M:M,\"PASS\")"
        "+COUNTIF('7. Element Results'!G:G,\"PASS\")>0,\"PASS\",\"INCOMPLETE\"))"
    )
    c = ws.cell(row=r, column=2, value=auto_formula)
    c.font = _font(bold=True, size=12); c.alignment = _align("center", "center")
    c.border = _thin(); c.fill = _fill(CWH)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    label(ws, r, 5, "(auto only — see signed result)", bg=CWH, cols=2)

    r += 1                                              # r = 41  signed
    label(ws, r, 1, "SIGNED OVERALL RESULT", bold=True)
    signed_row = r
    c = ws.cell(row=r, column=2, value="PASS")
    c.font = _font(bold=True, size=12, color=CUI); c.alignment = _align("center", "center")
    c.border = _thin(); c.fill = _fill(CY)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    label(ws, r, 5, "PASS / PASS WITH NOTES / FAIL", bg=CWH, cols=2)

    dv = DataValidation(type="list", formula1='"PASS,PASS WITH NOTES,FAIL"')
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=signed_row, column=2))

    return (ws, tol_rows,
            ct_sta_pri, ct_sta_sec, ct_Y_pri, ct_Y_sec,
            ct_neut_pri, ct_neut_sec, vt_main_kV, vt_main_V,
            vt_neut_kV, vt_neut_V, signed_row)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 – EQUIPMENT & REFERENCES
# ═══════════════════════════════════════════════════════════════════════════════
def build_equip(wb):
    ws = wb.create_sheet("2. Equip & Refs")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [30, 40, 22, 16]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 4, "EQUIPMENT & REFERENCES")
    r += 2
    h1(ws, r, 1, 4, "TEST EQUIPMENT USED")
    r += 1
    for col, hdr in enumerate(["Equipment", "Manufacturer / Model", "Asset / Serial No.", "Cal. Due"], 1):
        chdr(ws, r, col, hdr)
    for _ in range(6):
        r += 1
        for col in range(1, 5): inp(ws, r, col)

    r += 2
    h1(ws, r, 1, 4, "DOCUMENTS & DRAWINGS")
    r += 1
    for col, hdr in enumerate(["Document", "Title / Description", "Document No.", "Rev"], 1):
        chdr(ws, r, col, hdr)
    for doc in ["PCM600 Project File / Signal Matrix Tool (SMT)",
                "AC Elementary / Schematic",
                "DC Control Schematic",
                "Generator Protection Single-Line Diagram",
                "Protection Coordination Study",
                "Maintenance Work Order"]:
        r += 1
        label(ws, r, 1, doc)
        for col in range(2, 5): inp(ws, r, col)
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 – PRE-TEST
# ═══════════════════════════════════════════════════════════════════════════════
def build_pretest(wb):
    ws = wb.create_sheet("3. Pre-Test")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [42, 32, 16, 22]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 4, "PRE-TEST CHECKS")

    def check_section(ws, r, title, rows):
        r += 2
        h1(ws, r, 1, 4, title)
        r += 1
        for col, hdr in enumerate(["Check", "Expected / Action", "Result", "Notes"], 1):
            chdr(ws, r, col, hdr)
        for chk, exp in rows:
            r += 1
            label(ws, r, 1, chk)
            label(ws, r, 2, exp, bg=CCR)
            inp(ws, r, 3); inp(ws, r, 4)
        return r

    r = check_section(ws, r, "3.1  COMMUNICATIONS & IDENTIFICATION", [
        ("Connect to IED via PCM600 or web browser (Ethernet)",
         "PCM600 connects and IED tree populates"),
        ("Confirm IED name matches asset register",
         "LHMI: Main menu → Settings → IED name (TERMINALID)"),
        ("Confirm firmware version",
         "LHMI: Main menu → Diagnostics → IED status → Version"),
        ("Confirm serial number",
         "LHMI: Main menu → Diagnostics → IED status → Product info"),
        ("Normal LED (green, fixed) illuminated",
         "Green LED ON = IED healthy and in service"),
        ("No active alarms on LHMI alarm LEDs",
         "All alarm LEDs off (or known / accepted conditions only)"),
        ("Active setting group confirmed",
         "LHMI: Main menu → Settings → Setting group (ACTVGRP)"),
    ])

    # 3.2  INTERRSIG Self-Supervision (identical across Relion 670 family)
    r += 2
    h1(ws, r, 1, 4,
       "3.2  IED SELF-SUPERVISION  (INTERRSIG — LHMI: Main menu → Diagnostics → IED status → Signals)")
    r += 1
    note_row(ws, r, 1, 4,
             "The INTERRSIG function block provides continuous hardware and software self-monitoring. "
             "All signals should be NORMAL (0). Any FAIL signal will assert the INTERNAL FAIL contact on the PSM module. "
             "Navigate LHMI: Main menu → Diagnostics → IED status → Internal events to review the event log (last 40 events, FIFO).",
             height=44)
    r += 1
    for col, hdr in enumerate(["INTERRSIG Signal", "Description / What it monitors",
                                "State (NORMAL / FAIL / WARN)", "Action if not NORMAL"], 1):
        chdr(ws, r, col, hdr)
    for sig, desc in [
        ("FAIL",           "General IED failure — INTERNAL FAIL contact asserted on PSM"),
        ("WARNING",        "Warning condition — degraded operation, not yet a failure"),
        ("RTCERROR",       "Real-time clock error — timekeeping or battery issue"),
        ("TIMESYNCHERROR", "Time synchronisation error — IRIG-B / IEEE 1588 / NTP loss"),
        ("RTEERROR",       "Real-time execution error — protection task overrun"),
        ("IEC61850ERROR",  "IEC 61850 communication error — GOOSE or MMS failure"),
        ("WATCHDOG",       "Watchdog timer failure — CPU or task hang"),
        ("LMDERROR",       "LDCM (Line Data Communication Module) error — if fitted"),
        ("APPERROR",       "Application configuration error — PCM600 config mismatch"),
    ]:
        r += 1
        label(ws, r, 1, sig, bg=CLB, bold=True)
        label(ws, r, 2, desc, bg=CCR)
        inp(ws, r, 3); inp(ws, r, 4)

    r += 1
    h1(ws, r, 1, 4, "3.2 (cont.)  HARDWARE MODULE STATUS  (LHMI: Main menu → Diagnostics → Hardware)")
    r += 1
    for col, hdr in enumerate(["Module", "Description", "Status (OK / FAIL)", "Notes"], 1):
        chdr(ws, r, col, hdr)
    for mod, desc in [
        ("PSM",    "Power Supply Module — DC supply health and INTERNAL FAIL relay contact"),
        ("ADOne",  "A/D converter on TRM (Transformer Input Module) — CT/VT channel offsets"),
        ("ADTwo",  "Second A/D converter (if TRM has 2 A/D boards)"),
        ("BIM",    "Binary Input Module — all BIM cards fitted"),
        ("BOM",    "Binary Output Module — all BOM cards fitted"),
        ("IOM",    "Input/Output Module — combined BIM/BOM cards"),
        ("MIM",    "mA Input Module — if fitted for analogue supervision"),
        ("LDCM",   "Line Data Communication Module — if fitted"),
    ]:
        r += 1
        label(ws, r, 1, mod, bg=CLB, bold=True)
        label(ws, r, 2, desc, bg=CCR)
        inp(ws, r, 3); inp(ws, r, 4)

    r = check_section(ws, r, "3.3  TIME SYNCHRONISATION", [
        ("LHMI: Main menu → Diagnostics → IED status → Time",
         "IED time within 1 s of reference clock"),
        ("Time source active (IRIG-B / IEEE 1588 / NTP / manual)",
         "TIMESYNCHERROR = NORMAL; source shown as active"),
        ("Confirm date and time correct",
         "Check against substation master clock or GPS reference"),
    ])

    r = check_section(ws, r, "3.4  EVENT RECORDS (pre-test baseline)", [
        ("LHMI: Main menu → Events — review disturbance records",
         "No unexplained trips or alarms since last test"),
        ("Note most recent event number",
         "Baseline for post-test comparison"),
        ("Review internal event list (INTERRSIG)",
         "No historic FAIL or WARNING events outstanding"),
        ("Confirm disturbance recorder cleared / downloaded",
         "Oscillography available for any recent events"),
    ])

    r = check_section(ws, r, "3.5  SETTINGS VERIFICATION  (PCM600 → Upload from IED)", [
        ("Upload settings via PCM600 — compare to approved file",
         "Settings match last approved PCM600 project"),
        ("CTprim/CTsec stator CT correct",                "As per Cover sheet"),
        ("CTprim/CTsec Y-side (87G) CT correct",          "As per Cover sheet"),
        ("CTprim/CTsec neutral CT correct",               "As per Cover sheet"),
        ("VTprim/VTsec main VT correct",                  "As per Cover sheet"),
        ("VTprim/VTsec neutral VT correct",               "As per Cover sheet"),
        ("GENPDIF (87G) differential settings correct",   "Match approved coordination study"),
        ("GBASVAL (Ur, Ir, Sr) correct",                  "Match generator nameplate"),
        ("Active setting group as expected (ACTVGRP)",    "Group 1 or as applicable"),
        ("No unauthorised changes since last test",       "Settings unchanged"),
    ])

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 – ANALOGUE INPUTS
# ═══════════════════════════════════════════════════════════════════════════════
def build_analogue(wb):
    ws = wb.create_sheet("4. Analogue Inputs")
    ws.sheet_view.showGridLines = False
    col_widths = [24, 14, 13, 16, 16, 13, 10, 10, 8, 8, 8, 8, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    r = 1
    title_row(ws, r, 1, 13, "ANALOGUE INPUT VERIFICATION — PRIMARY VALUE CHECK")
    r += 1
    note_row(ws, r, 1, 13,
             "Verify the IED reports the correct PRIMARY value (Injected Secondary × CT or VT ratio). "
             "Use LHMI: Main menu → Measurements → [function block] to read primary values. "
             "All IED readings are in primary units (CTprim × secondary, VTprim × secondary).", 36)

    HDRS = ["Channel", "Injected Sec.\n(A or V)", "Inj. Angle\n(°)",
            "Expected Primary\n(auto)", "IED Primary\nReading",
            "IED Angle\nReading", "Mag\nErr %", "Angle\nErr °",
            "Mag\nTol %", "Ang\nTol °", "", "", "P/F"]

    def analogue_block(ws, r, title, ratio_name, tol_name, channels):
        r += 2
        h1(ws, r, 1, 13, title)
        r += 1
        label(ws, r, 1,
              f"Ratio: {ratio_name}  |  Tolerance: {tol_name}  (defined names on Cover sheet)",
              bg=CLB, cols=13)
        r += 1
        for col, hdr in enumerate(HDRS, 1):
            chdr(ws, r, col, hdr)
        for ch in channels:
            r += 1
            label(ws, r, 1, ch, bg=CLB)
            inp(ws, r, 2); inp(ws, r, 3)
            c = ws.cell(row=r, column=4,
                        value=f'=IFERROR(IF(B{r}="","",B{r}*{ratio_name}),"—")')
            c.fill = _fill(CGR); c.font = _font(bold=True)
            c.border = _thin(); c.alignment = _align("center", "center")
            inp(ws, r, 5); inp(ws, r, 6)
            c = ws.cell(row=r, column=7,
                        value=f'=IFERROR(IF(OR(D{r}="",E{r}=""),"",(E{r}-D{r})/D{r}*100),"—")')
            c.fill = _fill(CGR); c.border = _thin()
            c.number_format = "0.00"; c.alignment = _align("center", "center")
            c = ws.cell(row=r, column=8,
                        value=f'=IFERROR(IF(OR(C{r}="",F{r}=""),"",F{r}-C{r}),"—")')
            c.fill = _fill(CGR); c.border = _thin()
            c.number_format = "0.0"; c.alignment = _align("center", "center")
            c = ws.cell(row=r, column=9, value=f"={tol_name}")
            c.fill = _fill(CY); c.border = _thin(); c.alignment = _align("center", "center")
            c = ws.cell(row=r, column=10, value="=TolAngle_deg")
            c.fill = _fill(CY); c.border = _thin(); c.alignment = _align("center", "center")
            blank(ws, r, 11); blank(ws, r, 12)
            pf = (f'=IFERROR(IF(AND(B{r}="",E{r}=""),"—",'
                  f'IF(OR(AND(ISNUMBER(G{r}),ABS(G{r})>{tol_name}),'
                  f'AND(ISNUMBER(H{r}),ABS(H{r})>TolAngle_deg)),"FAIL","PASS")),"—")')
            c = ws.cell(row=r, column=13, value=pf)
            c.font = _font(bold=True); c.border = _thin(); c.alignment = _align("center", "center")
        return r

    r = analogue_block(ws, r,
        "4.1  STATOR PHASE CURRENTS  (IAX, IBX, ICX — stator / X-side CTs)  [Ratio_CT_stator]",
        "Ratio_CT_stator", "TolI_pct",
        ["IAX — Stator Phase A (X-side)", "IBX — Stator Phase B (X-side)", "ICX — Stator Phase C (X-side)"])

    r = analogue_block(ws, r,
        "4.2  Y-SIDE PHASE CURRENTS  (IAY, IBY, ICY — 87G differential input CTs)  [Ratio_CT_Y]",
        "Ratio_CT_Y", "TolI_pct",
        ["IAY — Phase A (Y-side / 87G CT)", "IBY — Phase B (Y-side / 87G CT)", "ICY — Phase C (Y-side / 87G CT)"])

    r = analogue_block(ws, r,
        "4.3  NEUTRAL CURRENT  (IN — generator neutral CT)  [Ratio_CT_neut]",
        "Ratio_CT_neut", "TolI_pct",
        ["IN — Generator neutral current"])

    r = analogue_block(ws, r,
        "4.4  MAIN VT VOLTAGES  (VAX, VBX, VCX — main terminal VTs)  [Ratio_VT_main]",
        "Ratio_VT_main", "TolV_pct",
        ["VAX — Phase A voltage (main VT)", "VBX — Phase B voltage", "VCX — Phase C voltage"])

    r = analogue_block(ws, r,
        "4.5  NEUTRAL VT VOLTAGE  (VN — stator neutral VT for 64S ground protection)  [Ratio_VT_neut]",
        "Ratio_VT_neut", "TolV_pct",
        ["VN — Generator neutral voltage (L-G primary)"])

    # 4.6  Differential balance check
    r += 2
    h1(ws, r, 1, 13,
       "4.6  DIFFERENTIAL BALANCE CHECK  "
       "(LHMI: Main menu → Measurements → GENPDIF → IOP, IRT)")
    r += 1
    note_row(ws, r, 1, 13,
             "Under balanced load: IOP (operate current) ≈ 0.00 pu, IRT (restraint) ≈ load current in pu of Ir. "
             "If IOP/IRT > 0.05 pu investigate CT polarity, CT_WyePoint direction settings, and CT connections. "
             "Currents are displayed in pu of GBASVAL Ir. Accept: IOP ≤ 0.05 pu.", 44)
    r += 1
    for col, hdr in enumerate(["Diff Function", "IOP\n(pu of Ir)", "IRT\n(pu of Ir)",
                                "MM = IOP/IRT", "Accept if\nIOP ≤ 0.05 pu",
                                "2nd Harm\n(pu)", "5th Harm\n(pu)",
                                "", "", "", "", "", "Result"], 1):
        chdr(ws, r, col, hdr)
    for elem in ["GENPDIF (87G) — Generator differential",
                 "T2WPDIF (87T) — Unit transformer differential (if fitted)"]:
        r += 1
        label(ws, r, 1, elem, bg=CLB)
        inp(ws, r, 2); inp(ws, r, 3)
        c = ws.cell(row=r, column=4,
                    value=f'=IFERROR(IF(OR(B{r}="",C{r}=""),"",B{r}/C{r}),"—")')
        c.fill = _fill(CGR); c.border = _thin()
        c.number_format = "0.000"; c.alignment = _align("center", "center")
        label(ws, r, 5, "IOP ≤ 0.05 pu", bg=CCR)
        inp(ws, r, 6); inp(ws, r, 7)
        blank(ws, r, 8); blank(ws, r, 9); blank(ws, r, 10); blank(ws, r, 11); blank(ws, r, 12)
        c = ws.cell(row=r, column=13,
                    value=f'=IFERROR(IF(B{r}="","—",IF(AND(ISNUMBER(B{r}),B{r}<=0.05),"PASS","FAIL")),"—")')
        c.font = _font(bold=True); c.border = _thin(); c.alignment = _align("center", "center")

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 – BINARY I/O
# ═══════════════════════════════════════════════════════════════════════════════
def build_binary(wb):
    ws = wb.create_sheet("5. Binary IO")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", [12, 30, 22, 26, 12]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 5, "BINARY I/O VERIFICATION")

    r += 2
    h1(ws, r, 1, 5, "5.1  BINARY INPUTS — Pickup Verification")
    r += 1
    note_row(ws, r, 1, 5,
             "Apply rated control voltage to each binary input. Navigate LHMI: Main menu → Test → "
             "Binary inputs — verify status changes 0→1. Alternatively view in PCM600 Signal Monitor.", 30)
    r += 1
    for col, hdr in enumerate(["Input No.", "Function Label (edit from schematics)",
                                "Voltage Applied", "LHMI / PCM600 Verification", "Result"], 1):
        chdr(ws, r, col, hdr)
    for i in range(1, 9):
        r += 1
        label(ws, r, 1, f"BI{i:02d}", bg=CLB, bold=True)
        inp(ws, r, 2, f"Binary Input {i} — edit label")
        inp(ws, r, 3)
        label(ws, r, 4, "LHMI: Test → Binary inputs", bg=CCR)
        inp(ws, r, 5)

    r += 2
    h1(ws, r, 1, 5, "5.2  BINARY OUTPUTS — Continuity Test")
    r += 1
    note_row(ws, r, 1, 5,
             "Use LHMI: Main menu → Test → Binary outputs to force each output. "
             "Verify contact continuity through wiring to trip coil / alarm circuit. "
             "Restore all outputs to normal after test — exit Test mode.", 30)
    r += 1
    for col, hdr in enumerate(["Output No.", "Function Label",
                                "LHMI Force Method", "Expected Result", "P/F"], 1):
        chdr(ws, r, col, hdr)
    for i in range(1, 9):
        r += 1
        label(ws, r, 1, f"BO{i:02d}", bg=CLB, bold=True)
        inp(ws, r, 2, f"Binary Output {i} — edit label")
        label(ws, r, 3, "LHMI: Test → Binary outputs", bg=CCR)
        label(ws, r, 4, "Contact closes — verify continuity", bg=CCR)
        inp(ws, r, 5)

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6 – ELEMENT TRACKER
# ═══════════════════════════════════════════════════════════════════════════════
def build_tracker(wb):
    ws = wb.create_sheet("6. Element Tracker")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [6, 22, 44, 14, 16, 12]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 6, "PROTECTION ELEMENT TRACKER")
    r += 2
    for col, hdr in enumerate(["#", "ANSI / IEC 61850", "Description",
                                "Enabled\n(Y / N / NA)", "Test Result Row", "Result"], 1):
        chdr(ws, r, col, hdr)

    elements = [
        # ── Main generator differential ───────────────────────────────────
        ("GENPDIF (87G)",    "Generator differential — restrained  (IDiff pickup, slope)"),
        ("GENPDIF-U (87G)",  "Generator differential — unrestrained  (high-set IDiff)"),
        ("T2WPDIF (87T)",    "Unit transformer differential (if unit transformer fitted)"),
        ("HZPDIF (87)",      "High impedance differential — single phase (if fitted)"),
        ("REFPDIF (87N)",    "Restricted earth fault — low impedance (if fitted)"),
        # ── Stator / rotor earth fault ─────────────────────────────────────
        ("STTIPHIZ (64S)",   "100% Stator earth fault — 20 Hz injection method (if fitted)"),
        ("ROTIPHIZ (64R)",   "Rotor earth fault — injection method (if fitted)"),
        # ── Loss of excitation / pole slip ────────────────────────────────
        ("LEXPDIS (40)",     "Loss of excitation — impedance characteristic (mho / offset mho)"),
        ("PSPPPAM (78)",     "Pole slip protection — lens characteristic"),
        # ── Overexcitation / V/Hz ─────────────────────────────────────────
        ("OEXPVPH (24)",     "Overexcitation / Volts-per-Hertz protection (if VT fitted)"),
        # ── Inadvertent energisation ──────────────────────────────────────
        ("AEGPVOC (50AE)",   "Accidental / inadvertent energisation protection"),
        # ── Backup overcurrent ─────────────────────────────────────────────
        ("VRPVOC (51V)",     "Voltage-restrained time overcurrent (backup stator OC)"),
        ("OC4PTOC (50/51P)", "Phase overcurrent backup — 4 stages"),
        ("EF4PTOC (50N/51N)","Residual / earth fault overcurrent backup — 4 stages"),
        ("NS4PTOC (46)",     "Negative-sequence overcurrent (current unbalance protection)"),
        # ── Thermal overload ───────────────────────────────────────────────
        ("GSPTTR (49S)",     "Stator thermal overload — I²t model (2 time constants)"),
        ("GRPTTR (49R)",     "Rotor thermal overload — I²t model (if fitted)"),
        # ── Voltage protection ─────────────────────────────────────────────
        ("UV2PTUV (27)",     "Undervoltage — 2 steps (if VT fitted)"),
        ("OV2PTOV (59)",     "Overvoltage — 2 steps (if VT fitted)"),
        ("ROV2PTOV (59N)",   "Residual overvoltage — 2 steps (if VT fitted)"),
        # ── Power / frequency ─────────────────────────────────────────────
        ("GOPPDOP (32)",     "Directional overpower (reverse power) protection"),
        ("GUPPDUP (37)",     "Directional underpower protection"),
        ("SAPTUF (81U)",     "Underfrequency — up to 10 stages (if VT fitted)"),
        ("SAPTOF (81O)",     "Overfrequency — up to 6 stages (if VT fitted)"),
        ("SAPFRC (81R)",     "Rate-of-change of frequency ROCOF (if VT fitted)"),
        # ── Breaker / circuit breaker ──────────────────────────────────────
        ("CCRBRF (50BF)",    "Breaker failure protection — 3-phase"),
        ("CCPDSC (52PD)",    "CB pole discordance protection"),
        # ── Synchronism / supervision ──────────────────────────────────────
        ("SESRSYN (25)",     "Synchronism check and energising check (if fitted)"),
        ("FUFSPVC (60)",     "Fuse failure / VT failure supervision"),
        ("CCSSPVC (CCS)",    "Current circuit supervision — open CT detection"),
    ]

    for i, (eid, desc) in enumerate(elements, 1):
        r += 1
        bg = CAL if i % 2 == 0 else CWH
        c = ws.cell(row=r, column=1, value=i)
        c.font = _font(bold=True); c.border = _thin(); c.fill = _fill(bg)
        c.alignment = _align("center", "center")
        label(ws, r, 2, eid, bg=CLB, bold=True)
        label(ws, r, 3, desc, bg=bg)
        inp(ws, r, 4); inp(ws, r, 5); inp(ws, r, 6)

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 7 – ELEMENT RESULTS
# ═══════════════════════════════════════════════════════════════════════════════
def build_results(wb):
    ws = wb.create_sheet("7. Element Results")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFG", [14, 40, 14, 14, 12, 14, 10]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 7, "PROTECTION ELEMENT TEST RESULTS")
    r += 1
    note_row(ws, r, 1, 7,
             "Setting column = relay pickup setting (pu, A secondary, V secondary, or %). "
             "Measured Pickup = value at which IED output asserted during injection. "
             "Pickup Err % = auto-calculated. Result = PASS if |Err| ≤ TolPickup_pct.", 32)
    r += 1
    for col, hdr in enumerate(["Element", "Description",
                                "Pickup Setting\n(pu, A sec, V sec, %)",
                                "Measured\nPickup",
                                "Pickup\nErr %",
                                "Measured\nTime (s)", "Result"], 1):
        chdr(ws, r, col, hdr)

    test_elements = [
        ("GENPDIF-R",   "87G restrained differential pickup — IDiff (pu of Ir)"),
        ("GENPDIF-U",   "87G unrestrained differential pickup — IDiff high-set (pu)"),
        ("T2WPDIF-R",   "87T unit transformer differential restrained (pu, if fitted)"),
        ("REFPDIF",     "87N restricted earth fault pickup (A secondary, if fitted)"),
        ("STTIPHIZ",    "64S 100% stator earth fault pickup — fundamental component (V)"),
        ("ROTIPHIZ",    "64R rotor earth fault pickup (Ω or pu, if fitted)"),
        ("LEXPDIS",     "40 loss of excitation — zone 1 reach (% impedance)"),
        ("LEXPDIS-2",   "40 loss of excitation — zone 2 reach (% impedance, if 2-zone)"),
        ("PSPPPAM",     "78 pole slip — lens inner boundary (% impedance)"),
        ("OEXPVPH",     "24 overexcitation pickup — V/Hz (% of rated, if VT fitted)"),
        ("AEGPVOC",     "50AE accidental energisation pickup (A secondary)"),
        ("VRPVOC",      "51V voltage-restrained OC pickup (A secondary)"),
        ("OC4PTOC-1",   "50/51P phase OC — stage 1 pickup (A secondary)"),
        ("OC4PTOC-2",   "50/51P phase OC — stage 2 pickup (A secondary)"),
        ("EF4PTOC-1",   "50N/51N residual OC — stage 1 pickup (A secondary)"),
        ("NS4PTOC-1",   "46 neg-seq OC — stage 1 pickup (A secondary)"),
        ("GSPTTR",      "49S stator thermal overload — trip level (A secondary or pu)"),
        ("UV2PTUV-1",   "27 undervoltage — stage 1 pickup (V secondary, if VT fitted)"),
        ("OV2PTOV-1",   "59 overvoltage — stage 1 pickup (V secondary, if VT fitted)"),
        ("ROV2PTOV-1",  "59N residual OV — stage 1 pickup (V secondary, if VT fitted)"),
        ("GOPPDOP",     "32 directional overpower pickup (W primary, if VT fitted)"),
        ("SAPTUF-1",    "81U underfrequency — stage 1 pickup (Hz, if VT fitted)"),
        ("SAPTOF-1",    "81O overfrequency — stage 1 pickup (Hz, if VT fitted)"),
        ("SAPFRC-1",    "81R ROCOF — stage 1 pickup (Hz/s, if VT fitted)"),
        ("CCRBRF",      "50BF breaker failure — current pickup + BFD timer (s)"),
        ("SESRSYN",     "25 synchronism check — closing window (degrees, if fitted)"),
    ]

    for i, (eid, desc) in enumerate(test_elements, 1):
        r += 1
        bg = CAL if i % 2 == 0 else CWH
        label(ws, r, 1, eid, bg=CLB, bold=True)
        label(ws, r, 2, desc, bg=bg)
        inp(ws, r, 3); inp(ws, r, 4)
        c = ws.cell(row=r, column=5,
                    value=f'=IFERROR(IF(OR(C{r}="",D{r}=""),"",(D{r}-C{r})/C{r}*100),"—")')
        c.fill = _fill(CGR); c.border = _thin()
        c.number_format = "0.00"; c.alignment = _align("center", "center")
        inp(ws, r, 6)
        result = (f'=IFERROR(IF(AND(C{r}="",D{r}=""),"—",'
                  f'IF(AND(ISNUMBER(E{r}),ABS(E{r})<=TolPickup_pct),"PASS","FAIL")),"—")')
        c = ws.cell(row=r, column=7, value=result)
        c.font = _font(bold=True); c.border = _thin(); c.alignment = _align("center", "center")

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 8 – POST-TEST
# ═══════════════════════════════════════════════════════════════════════════════
def build_posttest(wb):
    ws = wb.create_sheet("8. Post-Test")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [42, 30, 14, 20]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 4, "POST-TEST CHECKS")

    def check_section(ws, r, title, rows):
        r += 2
        h1(ws, r, 1, 4, title)
        r += 1
        for col, hdr in enumerate(["Check", "Expected / Action", "Result", "Notes"], 1):
            chdr(ws, r, col, hdr)
        for chk, exp in rows:
            r += 1
            label(ws, r, 1, chk)
            label(ws, r, 2, exp, bg=CCR)
            inp(ws, r, 3); inp(ws, r, 4)
        return r

    r = check_section(ws, r, "8.1  SETTINGS RESTORE & VERIFICATION", [
        ("Re-upload settings via PCM600 — verify unchanged",       "Match approved project file"),
        ("Exit Test mode on LHMI",                                 "LHMI: Main menu → Test → Exit test mode"),
        ("All forced binary outputs restored to normal",           "BOM cards de-energised"),
        ("Active setting group restored to normal",                "Group 1 or as applicable"),
        ("Normal LED (green, fixed) illuminated",                  "IED in service"),
    ])

    r += 2
    h1(ws, r, 1, 4,
       "8.2  FINAL SELF-SUPERVISION CHECK  "
       "(LHMI: Main menu → Diagnostics → IED status → Signals → INTERRSIG)")
    r += 1
    for col, hdr in enumerate(["Signal", "Expected", "Status", "Notes"], 1):
        chdr(ws, r, col, hdr)
    for sig in ["FAIL", "WARNING", "RTCERROR", "TIMESYNCHERROR",
                "RTEERROR", "IEC61850ERROR", "WATCHDOG", "APPERROR"]:
        r += 1
        label(ws, r, 1, sig, bg=CLB, bold=True)
        label(ws, r, 2, "NORMAL (0)", bg=CCR)
        inp(ws, r, 3); inp(ws, r, 4)

    r += 1
    h1(ws, r, 1, 4, "8.2 (cont.)  INTERNAL FAIL CONTACT CHECK")
    r += 1
    label(ws, r, 1, "INTERNAL FAIL contact (PSM terminal) — measure across contact")
    label(ws, r, 2, "Contact OPEN = IED healthy (de-energised in fault)", bg=CCR)
    inp(ws, r, 3); inp(ws, r, 4)

    r = check_section(ws, r, "8.3  PLANT RESTORATION", [
        ("Remove test connections from CT / VT terminals",         "Field CTs and VTs reconnected"),
        ("Restore control voltage wiring",                         "All control circuits normal"),
        ("Remove CT secondary shorting links — verify not open",   "CT circuits intact"),
        ("Verify IED measuring load current — LHMI Measurements",  "Compare to substation metering"),
        ("Differential balance — IOP ≤ 0.05 pu under load",       "GENPDIF IOP confirmed balanced"),
        ("64S stator ground neutral VT voltage reading normal",    "VN ≈ expected under load"),
        ("STTIPHIZ 64S injection source reconnected (if isolated)","20 Hz source back in service"),
        ("REX060 / 64R rotor earth fault module reconnected",      "Rotor earth fault module in service"),
        ("Cancel all outstanding LOTO permits",                    "All permits cleared"),
        ("Notify operations — IED returned to service",            "Operations informed"),
    ])

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 9 – FILES & SIGN-OFF
# ═══════════════════════════════════════════════════════════════════════════════
def build_signoff(wb, signed_row):
    ws = wb.create_sheet("9. Files & Sign-Off")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [28, 38, 20, 16]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 4, "FILES INDEX & SIGN-OFF")

    r += 2
    h1(ws, r, 1, 4, "OVERALL TEST RESULT (from Cover sheet)")
    r += 1
    label(ws, r, 1, "Signed Overall Result", bold=True)
    c = ws.cell(row=r, column=2, value=f"='1. Cover'!B{signed_row}")
    c.font = _font(bold=True, size=13); c.alignment = _align("center", "center")
    c.border = _thin(); c.fill = _fill(CY)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)

    r += 2
    h1(ws, r, 1, 4, "TEST ARTEFACTS INDEX")
    r += 1
    for col, hdr in enumerate(["Artefact", "File Name / SharePoint Location", "Date Saved", "Notes"], 1):
        chdr(ws, r, col, hdr)
    for art in ["PCM600 project file (.pcm) — as-left backup",
                "Pre-test disturbance records / internal event log",
                "Post-test disturbance records",
                "Completed results template (this file)",
                "Oscillography / disturbance records",
                "Photographs / site notes"]:
        r += 1
        label(ws, r, 1, art)
        inp(ws, r, 2); inp(ws, r, 3); inp(ws, r, 4)

    r += 2
    h1(ws, r, 1, 4, "DEFECTS / FINDINGS LOG")
    r += 1
    for col, hdr in enumerate(["#", "Defect Description", "Action Taken / Work Order", "Closed"], 1):
        chdr(ws, r, col, hdr)
    for i in range(1, 7):
        r += 1
        c = ws.cell(row=r, column=1, value=i)
        c.border = _thin(); c.alignment = _align("center", "center")
        for col in range(2, 5): inp(ws, r, col)

    r += 2
    h1(ws, r, 1, 4, "SIGN-OFF")
    r += 1
    for col, hdr in enumerate(["Role", "Name (print)", "Signature", "Date"], 1):
        chdr(ws, r, col, hdr)
    for role in ["Test Engineer", "Witness (Operations)", "Approving Engineer"]:
        r += 1
        ws.row_dimensions[r].height = 28
        label(ws, r, 1, role, bg=CLB, bold=True)
        inp(ws, r, 2); inp(ws, r, 3); inp(ws, r, 4)

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# DEFINED NAMES
# ═══════════════════════════════════════════════════════════════════════════════
def add_names(wb, ct_sta_pri, ct_sta_sec, ct_Y_pri, ct_Y_sec,
              ct_neut_pri, ct_neut_sec, vt_main_kV, vt_main_V,
              vt_neut_kV, vt_neut_V, tol_rows):
    sn = "'1. Cover'"
    defs = {
        "Ratio_CT_stator":  f"{sn}!$C${ct_sta_pri}/{sn}!$C${ct_sta_sec}",
        "Ratio_CT_Y":       f"{sn}!$C${ct_Y_pri}/{sn}!$C${ct_Y_sec}",
        "Ratio_CT_neut":    f"{sn}!$C${ct_neut_pri}/{sn}!$C${ct_neut_sec}",
        "Ratio_VT_main":    f"{sn}!$C${vt_main_kV}*1000/{sn}!$C${vt_main_V}",
        "Ratio_VT_neut":    f"{sn}!$C${vt_neut_kV}*1000/{sn}!$C${vt_neut_V}",
        "TolV_pct":         f"{sn}!$B${tol_rows['Voltage Magnitude Tolerance']}",
        "TolI_pct":         f"{sn}!$B${tol_rows['Current Magnitude Tolerance']}",
        "TolAngle_deg":     f"{sn}!$B${tol_rows['Angle Tolerance']}",
        "TolPickup_pct":    f"{sn}!$B${tol_rows['Pickup Tolerance']}",
        "TolTime_pct":      f"{sn}!$B${tol_rows['Timing Tolerance']}",
    }
    for name, attr in defs.items():
        dn = DefinedName(name=name, attr_text=attr)
        wb.defined_names[name] = dn


# ═══════════════════════════════════════════════════════════════════════════════
# CONDITIONAL FORMATTING
# ═══════════════════════════════════════════════════════════════════════════════
def add_cf(wb):
    pf = PatternFill("solid", fgColor=CPF); ff = PatternFill("solid", fgColor=CFL)
    pfont = Font(bold=True, color=CPFF, name="Calibri")
    ffont = Font(bold=True, color=CFFL, name="Calibri")
    for sheet_name, rng in [("4. Analogue Inputs", "M1:M300"),
                             ("7. Element Results",  "G1:G200"),
                             ("6. Element Tracker",  "F1:F200")]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.conditional_formatting.add(rng, CellIsRule("equal", ['"PASS"'], fill=pf, font=pfont))
            ws.conditional_formatting.add(rng, CellIsRule("equal", ['"FAIL"'], fill=ff, font=ffont))


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()

    (cover_ws, tol_rows,
     ct_sta_pri, ct_sta_sec, ct_Y_pri, ct_Y_sec,
     ct_neut_pri, ct_neut_sec, vt_main_kV, vt_main_V,
     vt_neut_kV, vt_neut_V, signed_row) = build_cover(wb)

    build_equip(wb)
    build_pretest(wb)
    build_analogue(wb)
    build_binary(wb)
    build_tracker(wb)
    build_results(wb)
    build_posttest(wb)
    build_signoff(wb, signed_row)

    add_names(wb, ct_sta_pri, ct_sta_sec, ct_Y_pri, ct_Y_sec,
              ct_neut_pri, ct_neut_sec, vt_main_kV, vt_main_V,
              vt_neut_kV, vt_neut_V, tol_rows)
    add_cf(wb)

    for ws in wb.worksheets:
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = "portrait"
        ws.page_margins.left = 0.4; ws.page_margins.right = 0.4
        ws.page_margins.top = 0.6;  ws.page_margins.bottom = 0.6

    out = "/home/michael/claudeProject /protection-relay-checksheet/REG670_Test_Results_Template.xlsx"
    wb.save(out)
    print(f"Saved: {out}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Signed result row: {signed_row}")
    print(f"Defined names: {list(wb.defined_names.keys())}")


if __name__ == "__main__":
    main()
