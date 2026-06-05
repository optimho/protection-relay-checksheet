#!/usr/bin/env python3
"""Generate Hitachi Relion 670 Biennial (2-Yearly) Visual Check Sheet (.xlsx).
   Single sheet: Relion 670 series (REG670, RET670) — graphical LHMI + PCM600.
   The 615 series is generated separately by generate_ABB615_biennial_check.py.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ── Colours ──────────────────────────────────────────────────────────────────
CB  = "1F4E79"; CLB = "D9E2F3"; CY  = "FFF2CC"; CCR = "FFF8E7"
CWH = "FFFFFF"; CAL = "F2F2F2"; CUI = "0000FF"

def _fill(c): return PatternFill("solid", fgColor=c)
def _font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

NCOLS = 5

def title_row(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = _fill(CB)
    c.font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28

def h1(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = _fill(CB); c.font = _font(bold=True, color="FFFFFF", size=10)
    c.alignment = _align("left", "center"); c.border = _thin()
    ws.row_dimensions[row].height = 18

def note(ws, row, text, bg=CY, height=32):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = _fill(bg); c.font = _font(italic=True, size=9)
    c.alignment = _align("left", "center", wrap=True); c.border = _thin()
    ws.row_dimensions[row].height = height

def lbl(ws, row, col, text, bg=CLB, bold=False, span=1, wrap=False):
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    c = ws.cell(row=row, column=col, value=text)
    c.fill = _fill(bg); c.font = _font(bold=bold)
    c.alignment = _align("left", "center", wrap=wrap); c.border = _thin()

def inp(ws, row, col, value=None, span=1, center=False, bg=CWH):
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    c = ws.cell(row=row, column=col, value=value)
    c.fill = _fill(bg); c.font = _font(color=CUI)
    c.alignment = _align("center" if center else "left", "center"); c.border = _thin()

def chdr(ws, row, col, text, span=1):
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    c = ws.cell(row=row, column=col, value=text)
    c.fill = _fill(CB); c.font = _font(bold=True, color="FFFFFF", size=9)
    c.alignment = _align("center", "center", wrap=True); c.border = _thin()
    ws.row_dimensions[row].height = 28

def set_col_widths(ws, widths):
    for col, w in zip("ABCDE", widths):
        ws.column_dimensions[col].width = w


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — RELION 670 (REG670, RET670)
# ═══════════════════════════════════════════════════════════════════════════════
def build_670_check(wb):
    ws = wb.active
    ws.title = "670 Check Sheet"
    ws.sheet_view.showGridLines = False
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 0
    ws.page_setup.fitToWidth = 1
    ws.page_margins.left = 0.5; ws.page_margins.right = 0.5
    ws.page_margins.top = 0.6;  ws.page_margins.bottom = 0.6
    set_col_widths(ws, [30, 34, 14, 12, 12])

    r = 1
    title_row(ws, r, "Hitachi RELION 670 IED — BIENNIAL (2-YEARLY) VISUAL CHECK SHEET  (REG670 / RET670)")

    # ── 1. IED / SITE INFORMATION ────────────────────────────────────────────
    r += 2
    h1(ws, r, "1.  IED AND SITE INFORMATION")
    for lbl_text, default in [
        ("IED Model",           "REG670  /  RET670  (circle one)"),
        ("Protected Equipment", ""),
        ("Site / Substation",   ""),
        ("IED Name (Station / Bay)", ""),
        ("Panel / Cubicle Tag", ""),
        ("Work Order No.",      ""),
        ("Date of Check",       ""),
        ("Technician Name",     ""),
        ("Next Check Due",      "(2 years from date above)"),
    ]:
        r += 1
        lbl(ws, r, 1, lbl_text, bold=True)
        inp(ws, r, 2, default, span=4)

    # ── 2. CONNECTION METHOD ─────────────────────────────────────────────────
    r += 2
    h1(ws, r, "2.  CONNECTION METHOD")
    r += 1
    note(ws, r,
         "OPTION A — LHMI only (no laptop): The RET/REG670 has a full graphical touchscreen or keypad LHMI. "
         "Navigate with arrow keys / touchscreen. Main menu accessible from the LHMI home screen.\n"
         "OPTION B — Ethernet + PCM600 laptop: Connect laptop Ethernet to IED front port (RJ45). "
         "Open PCM600 → Online > Signal Monitor for live values. IP address from LHMI: "
         "Main menu → Settings → Communication → Ethernet.\n"
         "OPTION C — Ethernet + web browser: Navigate to IED IP address in browser (if web server enabled).",
         bg=CLB, height=60)
    r += 1
    lbl(ws, r, 1, "Connection method used today", bold=True)
    inp(ws, r, 2, "LHMI only  /  Ethernet + PCM600  /  Ethernet + browser  (delete as applicable)", span=4)

    # ── 3. SELF-TEST / DEVICE STATUS ────────────────────────────────────────
    r += 2
    h1(ws, r, "3.  SELF-TEST AND IED STATUS  (INTERRSIG self-supervision)")
    r += 1
    note(ws, r,
         "LHMI path: Main menu → Diagnostics → IED status → Signals → INTERRSIG.  "
         "All signals should show NORMAL (0). A FAIL signal asserts the INTERNAL FAIL contact on the PSM. "
         "Hardware modules: Main menu → Diagnostics → Hardware → Module status.",
         bg=CLB, height=40)
    r += 1
    chdr(ws, r, 1, "Check Item")
    chdr(ws, r, 2, "LHMI Path / How to Read")
    chdr(ws, r, 3, "Expected")
    chdr(ws, r, 4, "Result")
    chdr(ws, r, 5, "Notes")

    status_checks = [
        ("Normal LED (green, fixed)",
         "Visual — front panel upper-left LED",
         "Green, solid"),
        ("No alarm LEDs active",
         "Visual — front panel LED indicators (15 programmable LEDs)",
         "All off (or known/accepted)"),
        ("INTERRSIG — FAIL signal",
         "LHMI: Main menu → Diagnostics → IED status → Signals",
         "NORMAL (0)"),
        ("INTERRSIG — WARNING signal",
         "Same path as above",
         "NORMAL (0)"),
        ("INTERRSIG — RTCERROR (clock error)",
         "Same path as above",
         "NORMAL (0)"),
        ("INTERRSIG — TIMESYNCHERROR",
         "Same path as above",
         "NORMAL (0)"),
        ("INTERRSIG — WATCHDOG",
         "Same path as above",
         "NORMAL (0)"),
        ("INTERNAL FAIL contact (PSM terminal)",
         "Measure across PSM INTERNAL FAIL output terminals",
         "Contact OPEN (IED healthy)"),
        ("PSM module status",
         "LHMI: Main menu → Diagnostics → Hardware",
         "OK"),
        ("BIM / BOM / TRM module status",
         "LHMI: Main menu → Diagnostics → Hardware",
         "OK"),
        ("Internal events — any recent FAILs?",
         "LHMI: Main menu → Diagnostics → IED status → Internal events",
         "No unexplained FAIL events"),
    ]
    for chk, how, exp in status_checks:
        r += 1
        lbl(ws, r, 1, chk)
        lbl(ws, r, 2, how, bg=CCR, wrap=True)
        lbl(ws, r, 3, exp, bg=CCR)
        inp(ws, r, 4, center=True)
        inp(ws, r, 5)

    # ── 4. ANALOGUE MEASUREMENTS ─────────────────────────────────────────────
    r += 2
    h1(ws, r, "4.  ANALOGUE MEASUREMENTS  (Primary Values)")
    r += 1
    note(ws, r,
         "LHMI: Main menu → Measurements → [function block, e.g. T2WPDIF, GENPDIF, OC4PTOC] → monitored data.  "
         "All values displayed in primary units. Record current load values below.  "
         "PCM600: Signal Monitor shows live primary values.",
         bg=CLB, height=36)
    r += 1
    chdr(ws, r, 1, "Quantity")
    chdr(ws, r, 2, "LHMI Path / Function Block")
    chdr(ws, r, 3, "Measured Value")
    chdr(ws, r, 4, "Unit")
    chdr(ws, r, 5, "Notes")

    measurements = [
        ("IA — Phase A Current (W1)",  "Measurements → [SMAI/CT function] → IA", "", "A primary"),
        ("IB — Phase B Current (W1)",  "Measurements → [SMAI/CT function] → IB", "", "A primary"),
        ("IC — Phase C Current (W1)",  "Measurements → [SMAI/CT function] → IC", "", "A primary"),
        ("IA — Phase A Current (W2 / Y-side)",  "Measurements → [SMAI W2] → IA (if 2nd CT set)", "", "A primary"),
        ("IB — Phase B Current (W2 / Y-side)",  "", "", "A primary"),
        ("IC — Phase C Current (W2 / Y-side)",  "", "", "A primary"),
        ("VA — Phase A Voltage",       "Measurements → [voltage function] → VA\n(if VT fitted)", "", "kV primary"),
        ("VB — Phase B Voltage",       "(if VT fitted)", "", "kV primary"),
        ("VC — Phase C Voltage",       "(if VT fitted)", "", "kV primary"),
        ("Frequency",                  "Measurements → [voltage function] → frequency", "", "Hz"),
    ]
    for qty, method, val, unit in measurements:
        r += 1
        lbl(ws, r, 1, qty)
        lbl(ws, r, 2, method, bg=CCR, wrap=True)
        inp(ws, r, 3, val, center=True)
        lbl(ws, r, 4, unit, bg=CCR)
        inp(ws, r, 5)

    # ── 5. CLOCK CHECK ───────────────────────────────────────────────────────
    r += 2
    h1(ws, r, "5.  CLOCK AND TIME SYNCHRONISATION")
    r += 1
    chdr(ws, r, 1, "Check")
    chdr(ws, r, 2, "LHMI Path")
    chdr(ws, r, 3, "IED Value")
    chdr(ws, r, 4, "Reference")
    chdr(ws, r, 5, "Difference / OK?")

    for chk, path, ref in [
        ("IED Date and Time",
         "Main menu → Diagnostics → IED status → Time\n(or Configuration → Time)",
         "Today's date / current time"),
        ("Time source (IRIG-B / IEEE 1588 / NTP)",
         "Main menu → Diagnostics → IED status → Time source",
         "IRIG-B or IEEE 1588 preferred"),
        ("TIMESYNCHERROR = NORMAL",
         "Main menu → Diagnostics → IED status → Signals",
         "NORMAL"),
        ("Time difference (accept ≤ 1 s with sync; ≤ 10 s internal)",
         "Compare LHMI time to GPS/phone reference",
         "≤ 1 s"),
    ]:
        r += 1
        lbl(ws, r, 1, chk)
        lbl(ws, r, 2, path, bg=CCR, wrap=True)
        inp(ws, r, 3)
        lbl(ws, r, 4, ref, bg=CCR)
        inp(ws, r, 5)

    # ── 6. RECENT EVENTS ─────────────────────────────────────────────────────
    r += 2
    h1(ws, r, "6.  RECENT EVENTS  (record most recent disturbance records / events)")
    r += 1
    note(ws, r,
         "LHMI: Main menu → Events → Disturbance records list (most recent first).  "
         "PCM600: Event viewer shows all logged events with timestamps.  "
         "Record the most recent events; flag any unexplained trips or alarms.",
         bg=CLB, height=32)
    r += 1
    chdr(ws, r, 1, "Event #\n(most recent = 1)")
    chdr(ws, r, 2, "Date")
    chdr(ws, r, 3, "Time")
    chdr(ws, r, 4, "Event Description / Type", span=2)
    for i in range(1, 13):
        r += 1
        bg = CAL if i % 2 == 0 else CWH
        c = ws.cell(row=r, column=1, value=i)
        c.font = _font(bold=True); c.border = _thin(); c.fill = _fill(bg)
        c.alignment = _align("center", "center")
        inp(ws, r, 2, bg=bg); inp(ws, r, 3, bg=bg)
        inp(ws, r, 4, span=2, bg=bg)

    r += 1
    lbl(ws, r, 1, "Total disturbance records stored (as displayed on LHMI)", bold=True)
    inp(ws, r, 2, span=4)

    # ── 7. DCS / 800XA CHECK ─────────────────────────────────────────────────
    r += 2
    h1(ws, r, "7.  DCS / ABB 800XA COMMUNICATION CHECK")
    r += 1
    note(ws, r,
         "Open the IED faceplate in the ABB 800XA control system. "
         "Verify live values (currents, IED status indicators) are updating. "
         "The 800XA communicates via IEC 61850 GOOSE / MMS. "
         "No yellow or red communication alarm objects should be present.",
         bg=CLB, height=36)
    r += 1
    chdr(ws, r, 1, "Check")
    chdr(ws, r, 2, "Description")
    chdr(ws, r, 3, "Result\n(Y / N)")
    chdr(ws, r, 4, "Notes", span=2)
    for chk, desc in [
        ("800XA faceplate opened",
         "Navigate to IED faceplate in 800XA Plant Explorer"),
        ("IED Status = HEALTHY / OK",
         "Green status indicator on faceplate — not alarm/fault colour"),
        ("Live current values updating",
         "Measured currents on faceplate match LHMI readings (not frozen)"),
        ("No communication alarm",
         "No yellow/red GOOSE or MMS communication alarm objects"),
        ("800XA object path (record below)",
         "e.g. 'Site / Substation / Relay Tag / Faceplate'"),
    ]:
        r += 1
        lbl(ws, r, 1, chk, bold=True)
        lbl(ws, r, 2, desc, bg=CCR)
        inp(ws, r, 3, center=True)
        inp(ws, r, 4, span=2)

    # ── 8. OPTIONAL — CLEAR EVENT BUFFERS ───────────────────────────────────
    r += 2
    h1(ws, r, "8.  OPTIONAL — CLEAR EVENT BUFFERS  (only if required by your procedure)")
    r += 1
    note(ws, r,
         "Only clear events if required by site procedure. "
         "Ensure all events have been downloaded/recorded BEFORE clearing. "
         "LHMI: Main menu → Events → [select event] → Clear  (or Clear all).  "
         "Internal events (INTERRSIG log) clear automatically after 40 events (FIFO).",
         bg=CY, height=36)
    r += 1
    chdr(ws, r, 1, "Action")
    chdr(ws, r, 2, "LHMI Path")
    chdr(ws, r, 3, "Performed?\n(Y / N / N/A)")
    chdr(ws, r, 4, "Notes", span=2)
    for action, path in [
        ("Download disturbance records", "PCM600 → Disturbance handling → Save to file"),
        ("Clear disturbance records",    "LHMI: Main menu → Events → Clear disturbance records"),
    ]:
        r += 1
        lbl(ws, r, 1, action, bold=True)
        lbl(ws, r, 2, path, bg=CCR)
        inp(ws, r, 3, center=True)
        inp(ws, r, 4, span=2)

    # ── 9. OBSERVATIONS & SIGN-OFF ───────────────────────────────────────────
    r += 2
    h1(ws, r, "9.  OBSERVATIONS AND CORRECTIVE ACTIONS")
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r+3, end_column=NCOLS)
    c = ws.cell(row=r, column=1); c.fill = _fill(CWH); c.border = _thin()
    c.alignment = _align("left", "top", wrap=True)
    for rr in range(r+1, r+4):
        ws.row_dimensions[rr].height = 18
        for col in range(1, NCOLS+1):
            ws.cell(row=rr, column=col).border = _thin()
    r += 4

    r += 1
    h1(ws, r, "10.  SIGN-OFF")
    r += 1
    chdr(ws, r, 1, "Role"); chdr(ws, r, 2, "Name (print)")
    chdr(ws, r, 3, "Signature"); chdr(ws, r, 4, "Date", span=2)
    for role in ["Technician", "Supervisor / Witness"]:
        r += 1
        ws.row_dimensions[r].height = 24
        lbl(ws, r, 1, role, bold=True)
        inp(ws, r, 2); inp(ws, r, 3); inp(ws, r, 4, span=2)

    return ws


def main():
    wb = openpyxl.Workbook()
    build_670_check(wb)

    out = "/home/michael/claudeProject /protection-relay-checksheet/HIT_670_Biennial_Visual_Check.xlsx"
    wb.save(out)
    print(f"Saved: {out}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
