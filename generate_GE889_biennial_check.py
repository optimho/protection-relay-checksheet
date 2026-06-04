#!/usr/bin/env python3
"""Generate GE Multilin 889 Generator Protection System Biennial (2-Yearly) Visual Check Sheet (.xlsx)."""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ── Colours ──────────────────────────────────────────────────────────────────
CB  = "1F4E79"   # deep blue   – headers
CLB = "D9E2F3"   # light blue  – label / instruction cells
CY  = "FFF2CC"   # yellow      – caution / optional
CCR = "FFF8E7"   # cream       – reference text
CWH = "FFFFFF"
CAL = "F2F2F2"
CUI = "0000FF"

def _fill(c): return PatternFill("solid", fgColor=c)
def _font(bold=False, color="000000", size=10, italic=False, name="Calibri"):
    return Font(bold=bold, color=color, size=size, italic=italic, name=name)
def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

NCOLS = 5   # A-E

def title_row(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = _fill(CB)
    c.font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28

def h1(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = _fill(CB)
    c.font = _font(bold=True, color="FFFFFF", size=10)
    c.alignment = _align("left", "center")
    c.border = _thin()
    ws.row_dimensions[row].height = 18

def note(ws, row, text, bg=CY, height=28):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = _fill(bg)
    c.font = _font(italic=True, size=9)
    c.alignment = _align("left", "center", wrap=True)
    c.border = _thin()
    ws.row_dimensions[row].height = height

def lbl(ws, row, col, text, bg=CLB, bold=False, span=1, wrap=False, size=10):
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    c = ws.cell(row=row, column=col, value=text)
    c.fill = _fill(bg)
    c.font = _font(bold=bold, size=size)
    c.alignment = _align("left", "center", wrap=wrap)
    c.border = _thin()

def inp(ws, row, col, value=None, span=1, center=False, bg=CWH, size=10):
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    c = ws.cell(row=row, column=col, value=value)
    c.fill = _fill(bg)
    c.font = _font(color=CUI, size=size)
    c.alignment = _align("center" if center else "left", "center")
    c.border = _thin()

def chdr(ws, row, col, text, span=1):
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    c = ws.cell(row=row, column=col, value=text)
    c.fill = _fill(CB)
    c.font = _font(bold=True, color="FFFFFF", size=9)
    c.alignment = _align("center", "center", wrap=True)
    c.border = _thin()
    ws.row_dimensions[row].height = 28


def build_check_sheet(wb):
    ws = wb.active
    ws.title = "GE 889 Biennial Check"
    ws.sheet_view.showGridLines = False
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 0
    ws.page_setup.fitToWidth = 1
    ws.page_margins.left = 0.5; ws.page_margins.right = 0.5
    ws.page_margins.top = 0.6;  ws.page_margins.bottom = 0.6

    # Column widths  A    B    C    D    E
    for col, w in zip("ABCDE", [30, 36, 14, 14, 12]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, "GE MULTILIN 889 GENERATOR PROTECTION SYSTEM — BIENNIAL (2-YEARLY) VISUAL CHECK SHEET")

    # ── 1. RELAY / SITE INFORMATION ─────────────────────────────────────────
    r += 2  # 3
    h1(ws, r, "1.  RELAY AND SITE INFORMATION")
    fields = [
        ("Relay Model",              "GE Multilin 889 Generator Protection System"),
        ("Protected Equipment",      ""),
        ("Site / Substation",        ""),
        ("Equipment Tag (relay)",    ""),
        ("Order Code / Part Number", ""),
        ("Serial Number",            ""),
        ("Firmware Version",         "(record from Status → Information → Main CPU)"),
        ("Panel / Cubicle Tag",      ""),
        ("Work Order No.",           ""),
        ("Date of Check",            ""),
        ("Technician Name",          ""),
        ("Next Check Due",           "(2 years from date above)"),
    ]
    for lbl_text, default in fields:
        r += 1
        lbl(ws, r, 1, lbl_text, bold=True)
        inp(ws, r, 2, default, span=4)

    # ── 2. CONNECTION METHOD ─────────────────────────────────────────────────
    r += 2
    h1(ws, r, "2.  CONNECTION METHOD")
    r += 1
    note(ws, r,
         "OPTION A — Front panel only (no laptop): Press any key to wake the graphical display. "
         "Navigate using the keypad and function keys.\n"
         "OPTION B — EnerVista 8 Series Setup Software via Ethernet (RJ45 port on rear): "
         "Launch EnerVista, add relay (File → New Device), enter relay IP address. "
         "Right-click device → Quick Connect.  Default Modbus TCP port: 502.\n"
         "OPTION C — USB local programming port (front panel): "
         "Connect USB cable; EnerVista detects via USB.",
         bg=CLB, height=60)
    r += 1
    lbl(ws, r, 1, "Connection method used today", bold=True)
    inp(ws, r, 2, "Front panel  /  EnerVista Ethernet  /  EnerVista USB  (delete as applicable)", span=4)

    # ── 3. SELF-TEST / DEVICE STATUS ────────────────────────────────────────
    r += 2
    h1(ws, r, "3.  SELF-TEST AND DEVICE STATUS")
    r += 1
    note(ws, r,
         "Front panel: Status → Device Status.  All items should show OK — no warnings or faults.  "
         "Normal LED (green, solid) must be illuminated.  Warning LED (amber) or Critical LED (red) "
         "indicates a condition requiring investigation before returning to service.  "
         "EnerVista: Actual Values → Status → Device Status.",
         bg=CLB, height=40)
    r += 1
    chdr(ws, r, 1, "Check Item")
    chdr(ws, r, 2, "How to Read / Navigation Path")
    chdr(ws, r, 3, "Expected")
    chdr(ws, r, 4, "Result")
    chdr(ws, r, 5, "Notes")

    status_checks = [
        ("Normal LED (green, solid)",
         "Visual — front panel top-left LED indicator",
         "ON solid green"),
        ("Warning LED (amber)",
         "Visual — front panel LED",
         "OFF — no warnings"),
        ("Critical LED (red)",
         "Visual — front panel LED",
         "OFF — no critical faults"),
        ("Device Status — no active faults",
         "Front panel: Status → Device Status\nEnerVista: Actual Values → Status → Device Status",
         "All OK"),
        ("Main CPU self-test",
         "Front panel: Status → Information → Main CPU",
         "No errors"),
        ("Comms CPU self-test",
         "Front panel: Status → Information → Comms CPU",
         "No errors"),
        ("Hardware module health",
         "Front panel: Status → Information → Hardware Versions",
         "All modules present"),
        ("Record firmware version",
         "Front panel: Status → Information → Main CPU → Firmware",
         "Record version"),
        ("Modbus TCP communications active",
         "Front panel: Status → Communications\nEnerVista: connected and updating",
         "Comms healthy"),
    ]
    for chk, how, exp in status_checks:
        r += 1
        lbl(ws, r, 1, chk)
        lbl(ws, r, 2, how, bg=CCR, wrap=True)
        lbl(ws, r, 3, exp, bg=CCR)
        inp(ws, r, 4, center=True)
        inp(ws, r, 5)

    # ── 4. ANALOGUE MEASUREMENTS ─────────────────────────────────────────────
    r += 2
    h1(ws, r, "4.  ANALOGUE MEASUREMENTS  (Primary Values — relay reports in primary units)")
    r += 1
    note(ws, r,
         "Front panel: Metering → Currents  (for phase and ground currents in primary A).  "
         "Metering → Voltages  (for phase voltages in primary kV).  "
         "Metering → Frequency  (Hz).  "
         "EnerVista: Actual Values → Metering → Currents / Voltages / Frequency.  "
         "Record measured values. No specific pass/fail for 2-yearly check — "
         "flag anything obviously wrong (e.g. one phase zero, both CT banks disagree).",
         bg=CLB, height=48)
    r += 1
    chdr(ws, r, 1, "Quantity")
    chdr(ws, r, 2, "Navigation Path")
    chdr(ws, r, 3, "Measured Value")
    chdr(ws, r, 4, "Unit")
    chdr(ws, r, 5, "Notes")

    measurements = [
        ("IA_J — Phase A (neutral-side CT, J1)",  "Metering → Currents", "", "A primary"),
        ("IB_J — Phase B (neutral-side CT, J1)",  "Metering → Currents", "", "A primary"),
        ("IC_J — Phase C (neutral-side CT, J1)",  "Metering → Currents", "", "A primary"),
        ("IA_K — Phase A (terminal-side CT, K1)", "Metering → Currents", "", "A primary"),
        ("IB_K — Phase B (terminal-side CT, K1)", "Metering → Currents", "", "A primary"),
        ("IC_K — Phase C (terminal-side CT, K1)", "Metering → Currents", "", "A primary"),
        ("IN — Ground / CBCT current",            "Metering → Currents", "", "A primary"),
        ("Va-n — Phase A voltage",                "Metering → Voltages", "", "kV primary"),
        ("Vb-n — Phase B voltage",                "Metering → Voltages", "", "kV primary"),
        ("Vc-n — Phase C voltage",                "Metering → Voltages", "", "kV primary"),
        ("Frequency",                             "Metering → Frequency", "", "Hz"),
        ("87G IOP (operate current, Phase A)",    "Metering → Generator → Percent Differential", "", "pu  (accept ≤ 0.05)"),
    ]
    for qty, method, val, unit in measurements:
        r += 1
        lbl(ws, r, 1, qty)
        lbl(ws, r, 2, method, bg=CCR, wrap=True)
        inp(ws, r, 3, val, center=True)
        lbl(ws, r, 4, unit, bg=CCR)
        inp(ws, r, 5)

    # 4 spare rows for site-specific quantities
    for i in range(4):
        r += 1
        bg = CAL if i % 2 == 0 else CWH
        lbl(ws, r, 1, "", bg=bg)
        lbl(ws, r, 2, "Metering → …", bg=CCR)
        inp(ws, r, 3, center=True, bg=bg)
        inp(ws, r, 4, bg=bg)
        inp(ws, r, 5, bg=bg)

    # ── 5. CLOCK / TIME CHECK ────────────────────────────────────────────────
    r += 2
    h1(ws, r, "5.  CLOCK AND TIME SYNCHRONISATION")
    r += 1
    chdr(ws, r, 1, "Check")
    chdr(ws, r, 2, "Navigation Path")
    chdr(ws, r, 3, "Relay Value")
    chdr(ws, r, 4, "Reference Value")
    chdr(ws, r, 5, "Difference / Result")

    time_checks = [
        ("Relay Date",
         "Front panel: Status → Clock\nEnerVista: Actual Values → Status → Clock",
         "", "Today's date"),
        ("Relay Time",
         "Front panel: Status → Clock\nEnerVista: Actual Values → Status → Clock",
         "", "GPS / NTP reference"),
        ("Time synchronisation source",
         "Front panel: Status → Clock\nor Status → PTP Status",
         "", "SNTP or IRIG-B preferred"),
        ("Time difference (accept ≤ 10 s)",
         "Relay time vs phone / GPS reference clock",
         "", "≤ 10 s"),
    ]
    for chk, how, relay_val, ref in time_checks:
        r += 1
        lbl(ws, r, 1, chk)
        lbl(ws, r, 2, how, bg=CCR, wrap=True)
        inp(ws, r, 3, relay_val, center=True)
        lbl(ws, r, 4, ref, bg=CCR)
        inp(ws, r, 5)

    # ── 6. RECENT EVENTS ─────────────────────────────────────────────────────
    r += 2
    h1(ws, r, "6.  RECENT EVENTS  (record most recent events — flag any unexplained trips or alarms)")
    r += 1
    note(ws, r,
         "Front panel: Records → Events.  Navigate to event list and scroll through entries.  "
         "EnerVista: Records → Events — export or view event list.  "
         "Record the most recent events. Flag any unexplained trips, alarms, or self-test events.",
         bg=CLB, height=36)
    r += 1
    chdr(ws, r, 1, "Event #\n(most recent = 1)")
    chdr(ws, r, 2, "Date")
    chdr(ws, r, 3, "Time")
    chdr(ws, r, 4, "Event Description / Type", span=2)
    for i in range(1, 13):
        r += 1
        bg = CAL if i % 2 == 0 else CWH
        c = ws.cell(row=r, column=1, value=i)
        c.font = _font(bold=True); c.border = _thin(); c.fill = _fill(bg)
        c.alignment = _align("center", "center")
        inp(ws, r, 2, bg=bg)
        inp(ws, r, 3, bg=bg)
        inp(ws, r, 4, span=2, bg=bg)

    r += 1
    lbl(ws, r, 1, "Total events since last check (approx.)", bold=True)
    inp(ws, r, 2, span=4)

    # ── 7. DCS / MODBUS TCP COMMS CHECK ──────────────────────────────────────
    r += 2
    h1(ws, r, "7.  DCS / MODBUS TCP COMMUNICATION CHECK  (800XA)")
    r += 1
    note(ws, r,
         "Open the relay faceplate in ABB 800XA. Verify that live values (currents, voltages, status) "
         "are updating and not frozen. Check for any Modbus communication alarm objects (yellow/red). "
         "The 889 communicates via Modbus TCP on port 502 (default). "
         "Record the 800XA object path or faceplate name below.",
         bg=CLB, height=40)
    r += 1
    chdr(ws, r, 1, "Check")
    chdr(ws, r, 2, "Description")
    chdr(ws, r, 3, "Result\n(Y / N)")
    chdr(ws, r, 4, "Notes", span=2)

    dcs_checks = [
        ("800XA faceplate opened",
         "Navigate to relay faceplate in ABB 800XA control system"),
        ("Live values updating",
         "Phase currents / voltages / status visible and updating with load"),
        ("No Modbus TCP communication alarm",
         "No yellow or red communication alarm objects on faceplate"),
        ("Status objects correct",
         "Relay status (Normal/Warning/Critical) visible and correct in 800XA"),
        ("Record 800XA object path / faceplate name",
         "e.g. \"Site/Substation/Relay Tag/Faceplate\""),
    ]
    for chk, desc in dcs_checks:
        r += 1
        lbl(ws, r, 1, chk, bold=True)
        lbl(ws, r, 2, desc, bg=CCR)
        inp(ws, r, 3, center=True)
        inp(ws, r, 4, span=2)

    # ── 8. OPTIONAL — CLEAR RECORDS ──────────────────────────────────────────
    r += 2
    h1(ws, r, "8.  OPTIONAL — CLEAR RECORDS  (only if required by procedure — confirm with engineer)")
    r += 1
    note(ws, r,
         "If your site procedure requires clearing event records after the biennial check, "
         "navigate to: Front panel: Records → Clear Records.  "
         "EnerVista: Records → Clear Records.  "
         "Ensure all events have been downloaded / recorded BEFORE clearing.",
         bg=CY, height=36)
    r += 1
    chdr(ws, r, 1, "Record Type")
    chdr(ws, r, 2, "Navigation Path")
    chdr(ws, r, 3, "Performed?\n(Y / N / N/A)")
    chdr(ws, r, 4, "Notes", span=2)
    for rec_type, path in [
        ("Events",
         "Front panel: Records → Clear Records → Events\nEnerVista: Records → Clear Records"),
        ("Transient Records  (if required)",
         "Front panel: Records → Clear Records → Transients"),
        ("Fault Reports  (if required)",
         "Front panel: Records → Clear Records → Fault Reports"),
    ]:
        r += 1
        lbl(ws, r, 1, rec_type, bg=CLB, bold=True)
        lbl(ws, r, 2, path, bg=CCR, wrap=True)
        inp(ws, r, 3, center=True)
        inp(ws, r, 4, span=2)

    # ── 9. OBSERVATIONS ──────────────────────────────────────────────────────
    r += 2
    h1(ws, r, "9.  OBSERVATIONS AND CORRECTIVE ACTIONS")
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 4, end_column=NCOLS)
    c = ws.cell(row=r, column=1)
    c.fill = _fill(CWH); c.border = _thin()
    c.alignment = _align("left", "top", wrap=True)
    ws.row_dimensions[r].height = 18
    for rr in range(r + 1, r + 5):
        ws.row_dimensions[rr].height = 18
        for col in range(1, NCOLS + 1):
            ws.cell(row=rr, column=col).border = _thin()
    r += 5

    # ── 10. SIGN-OFF ─────────────────────────────────────────────────────────
    r += 1
    h1(ws, r, "10.  SIGN-OFF")
    r += 1
    chdr(ws, r, 1, "Role")
    chdr(ws, r, 2, "Name (print)")
    chdr(ws, r, 3, "Signature")
    chdr(ws, r, 4, "Date", span=2)
    for role in ["Technician", "Supervisor / Witness"]:
        r += 1
        ws.row_dimensions[r].height = 24
        lbl(ws, r, 1, role, bold=True)
        inp(ws, r, 2); inp(ws, r, 3); inp(ws, r, 4, span=2)

    return ws


def main():
    wb = openpyxl.Workbook()
    build_check_sheet(wb)

    out = "/home/michael/claudeProject /protection-relay-checksheet/GE-889_Biennial_Visual_Check.xlsx"
    wb.save(out)
    print(f"Saved: {out}")


if __name__ == "__main__":
    main()
