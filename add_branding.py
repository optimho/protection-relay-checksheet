#!/usr/bin/env python3
"""Post-process all xlsx templates: add Contact Energy logo and footer.

Adds to every worksheet in every xlsx file (except New_Relay_Onboarding_Form.xlsx):
  - Contact Energy logo (PNG) anchored top-right of the title row
  - Footer: "Contact Energy" left | document name centre | page number right
"""

import glob
import os
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

PROJ    = "/home/michael/claudeProject /protection-relay-checksheet/"
LOGO    = PROJ + "Contact_2020_Master_RGB_Gradient.png"
SKIP    = {"New_Relay_Onboarding_Form.xlsx"}

# Logo display size — keep 2.18:1 aspect ratio, 40 px tall
LOGO_H  = 40
LOGO_W  = round(40 * (7331 / 3364))   # = 87 px

# Contact Energy deep blue for footer text
CE_BLUE = "1F4E79"


def logo_anchor(ws) -> str:
    """Return the cell address to anchor the logo (top-right of title row).

    Strategy: walk column widths from the right until we accumulate enough
    space for the logo (~90 character units ≈ 87 px at default zoom), then
    anchor one column further left so the logo sits just inside the right edge.
    Falls back to the penultimate column if widths are not set.
    """
    max_col = ws.max_column or 1
    # Collect column widths (default Excel column width is 8.43 characters)
    widths = []
    for i in range(1, max_col + 1):
        dim = ws.column_dimensions.get(get_column_letter(i))
        widths.append(dim.width if dim and dim.width else 8.43)

    # The logo is ~87 px wide; 1 character unit ≈ 7 px → logo ≈ 12.4 char units
    LOGO_CHARS = 12.5
    accumulated = 0.0
    anchor_idx = max_col   # fallback: last column
    for col_i in range(max_col, 0, -1):
        accumulated += widths[col_i - 1]
        if accumulated >= LOGO_CHARS:
            anchor_idx = col_i
            break

    return f"{get_column_letter(anchor_idx)}1"


def add_branding_to_file(fpath: str):
    fname = os.path.basename(fpath)
    # Use document name (no extension) for footer centre
    doc_name = os.path.splitext(fname)[0].replace("_", " ")

    wb = openpyxl.load_workbook(fpath)

    for ws in wb.worksheets:
        # ── Footer ────────────────────────────────────────────────────────────
        for footer in (ws.oddFooter, ws.evenFooter):
            footer.left.text  = "&BContact Energy"
            footer.left.size  = 8
            footer.left.color = CE_BLUE          # Contact Energy deep blue
            footer.center.text = doc_name
            footer.center.size = 8
            footer.right.text  = "Page &P of &N"
            footer.right.size  = 8

        # ── Print settings so footer appears ─────────────────────────────────
        ws.page_setup.paperSize     = ws.PAPERSIZE_A4
        ws.page_setup.orientation   = "portrait"
        ws.sheet_view.showGridLines = False   # already off in our sheets but safe to repeat

        # ── Logo (clear any previous run first) ───────────────────────────────
        ws._images = []
        img        = Image(LOGO)
        img.width  = LOGO_W
        img.height = LOGO_H
        anchor     = logo_anchor(ws)
        ws.add_image(img, anchor)

    wb.save(fpath)
    print(f"  ✓  {fname}  (sheets: {len(wb.worksheets)})")


def main():
    files = sorted(glob.glob(PROJ + "*.xlsx"))
    print(f"Contact Energy branding — logo {LOGO_W}×{LOGO_H} px\n")
    for fpath in files:
        fname = os.path.basename(fpath)
        if fname in SKIP:
            print(f"  –  {fname}  (skipped)")
            continue
        add_branding_to_file(fpath)
    print("\nDone — all files updated.")


if __name__ == "__main__":
    main()
