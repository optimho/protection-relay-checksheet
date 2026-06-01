#!/usr/bin/env python3
"""Generate SEL-787 Transformer Protection Relay Test Results Template (.xlsx)."""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import CellIsRule

# ── Colours ──────────────────────────────────────────────────────────────────
CB  = "1F4E79"   # deep blue  – headers
CLB = "D9E2F3"   # light blue – labels
CY  = "FFF2CC"   # yellow     – tolerance / caution
CCR = "FFF8E7"   # cream      – expected / reference
CPF = "C6EFCE"; CPFF = "006100"   # PASS
CFL = "FFC7CE"; CFFL = "9C0006"   # FAIL
CGR = "D9D9D9"   # grey       – auto-calculated
CAL = "F2F2F2"   # alt row
CUI = "0000FF"   # blue text  – user input
CWH = "FFFFFF"   # white

def _fill(c): return PatternFill("solid", fgColor=c)
def _font(bold=False, color="000000", size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def h1(ws, row, col1, col2, text):
    """Deep-blue section header spanning col1:col2."""
    ws.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    c = ws.cell(row=row, column=col1, value=text)
    c.fill = _fill(CB); c.font = _font(bold=True, color="FFFFFF", size=11)
    c.alignment = _align("left", "center"); c.border = _thin()

def label(ws, row, col, text, bg=CLB, bold=False, cols=1, wrap=False):
    if cols > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+cols-1)
    c = ws.cell(row=row, column=col, value=text)
    c.fill = _fill(bg); c.font = _font(bold=bold)
    c.alignment = _align("left", "center", wrap=wrap); c.border = _thin()

def inp(ws, row, col, value=None, cols=1):
    if cols > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+cols-1)
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(color=CUI); c.alignment = _align("left", "center")
    c.border = _thin(); c.fill = _fill(CWH)

def tol(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(color=CUI); c.alignment = _align("center", "center")
    c.border = _thin(); c.fill = _fill(CY)

def chdr(ws, row, col, text, bg=CB, fg="FFFFFF"):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = _fill(bg); c.font = _font(bold=True, color=fg, size=10)
    c.alignment = _align("center", "center", wrap=True); c.border = _thin()

def auto(ws, row, col, formula_or_value, fmt=None, bold=False):
    c = ws.cell(row=row, column=col, value=formula_or_value)
    c.fill = _fill(CGR); c.font = _font(bold=bold)
    c.alignment = _align("center", "center"); c.border = _thin()
    if fmt: c.number_format = fmt

def note_row(ws, row, col1, col2, text, height=32):
    ws.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    c = ws.cell(row=row, column=col1, value=text)
    c.fill = _fill(CY); c.font = _font(italic=True, size=10)
    c.alignment = _align("left", "center", wrap=True); c.border = _thin()
    ws.row_dimensions[row].height = height

def title_row(ws, row, col1, col2, text):
    ws.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    c = ws.cell(row=row, column=col1, value=text)
    c.fill = _fill(CB); c.font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 26

def blank(ws, row, col, bg=CWH):
    c = ws.cell(row=row, column=col)
    c.fill = _fill(bg); c.border = _thin()


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 – COVER
# ═══════════════════════════════════════════════════════════════════════════════
def build_cover(wb):
    ws = wb.active
    ws.title = "1. Cover"
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [30, 22, 14, 14, 14, 14]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 6, "SEL-787 TRANSFORMER PROTECTION RELAY — TEST RESULTS TEMPLATE")

    # ── Section 1: Relay ID ──────────────────────────────────────────────────
    r += 2                                          # r = 3
    h1(ws, r, 1, 6, "1.  RELAY IDENTIFICATION")
    rows_id = ["Relay Type / Model", "Relay ID (RID)", "Terminal ID (TID)",
               "Serial Number", "Firmware ID (FID)", "Checksum ID (CID)",
               "Protected Equipment (transformer tag)", "Site / Substation",
               "Test Date", "Test Engineer"]
    defaults = ["SEL-787"] + [""] * 9
    for lbl, dflt in zip(rows_id, defaults):
        r += 1
        label(ws, r, 1, lbl)
        inp(ws, r, 2, dflt, cols=5)

    # ── Section 2: Transformer / CT / VT data ───────────────────────────────
    r += 2                                          # r = 15
    h1(ws, r, 1, 6, "2.  TRANSFORMER AND CT/VT PARAMETERS")
    r += 1                                          # r = 16  column headers
    for col, hdr in enumerate(["Parameter", "Setting Label", "Value", "Units", "", ""], 1):
        chdr(ws, r, col, hdr)

    # (label, setting_label, default_value, unit)
    xfmr_data = [
        ("Transformer MVA Rating",          "MVA",              "",     "MVA"),
        ("Winding 1 Nominal Voltage (L-L)",  "VWDG1",            "",     "kV"),
        ("Winding 2 Nominal Voltage (L-L)",  "VWDG2",            "",     "kV"),
        ("Winding 1 CT Connection",          "W1CT",             "WYE",  "WYE / DELTA"),
        ("Winding 2 CT Connection",          "W2CT",             "WYE",  "WYE / DELTA"),
        ("W1 CT Ratio — Primary",            "CTR1 (primary)",   "",     "A primary"),
        ("W1 CT Ratio — Secondary",          "CTR1 (secondary)", "5",    "A secondary"),
        ("W2 CT Ratio — Primary",            "CTR2 (primary)",   "",     "A primary"),
        ("W2 CT Ratio — Secondary",          "CTR2 (secondary)", "5",    "A secondary"),
        ("Neutral CT Ratio — Primary",       "CTRN (primary)",   "",     "A primary"),
        ("Neutral CT Ratio — Secondary",     "CTRN (secondary)", "5",    "A secondary"),
        ("VT Ratio — Primary",               "PTR (primary)",    "",     "V primary"),
        ("VT Ratio — Secondary",             "PTR (secondary)",  "",     "V secondary"),
        ("VT Connection Type",               "DELTA_Y",          "WYE",  "WYE / DELTA"),
        ("Diff. TAP Winding 1",              "TAP1",             "",     "A secondary"),
        ("Diff. TAP Winding 2",              "TAP2",             "",     "A secondary"),
        ("W1 CT Connection Compensation",    "W1CTC",            "",     "0 = none"),
        ("W2 CT Connection Compensation",    "W2CTC",            "",     "0 = none"),
    ]
    # Note the row numbers for ratio cells (column C = col 3)
    # r=17: MVA, r=18: VWDG1, … r=22: CTR1 primary (index 5)
    data_start_row = r + 1   # = 17
    for lbl, slbl, dflt, unit in xfmr_data:
        r += 1
        label(ws, r, 1, lbl)
        label(ws, r, 2, slbl, bg=CCR)
        inp(ws, r, 3, dflt)
        label(ws, r, 4, unit, bg=CWH)
        blank(ws, r, 5); blank(ws, r, 6)

    # Row numbers for ratios: data_start_row + index
    # index 5 = CTR1 primary → row 17+5=22; index 6 = CTR1 sec → 23
    # index 7 = CTR2 primary → 24; index 8 = CTR2 sec → 25
    # index 9 = CTRN primary → 26; index 10 = CTRN sec → 27
    # index 11 = PTR primary → 28; index 12 = PTR sec → 29
    ctr1_pri = data_start_row + 5   # 22
    ctr1_sec = data_start_row + 6   # 23
    ctr2_pri = data_start_row + 7   # 24
    ctr2_sec = data_start_row + 8   # 25
    ctrn_pri = data_start_row + 9   # 26
    ctrn_sec = data_start_row + 10  # 27
    ptr_pri  = data_start_row + 11  # 28
    ptr_sec  = data_start_row + 12  # 29

    # ── Section 3: Tolerances ────────────────────────────────────────────────
    r += 2                                          # r = 36
    h1(ws, r, 1, 6, "3.  TEST TOLERANCES  (yellow cells are editable)")
    tol_rows = {}
    for name, dflt, unit in [
        ("Voltage Magnitude Tolerance",  1.0, "%"),
        ("Current Magnitude Tolerance",  1.0, "%"),
        ("Angle Tolerance",              2.0, "degrees"),
        ("Pickup Tolerance",             5.0, "%"),
        ("Timing Tolerance",             5.0, "%"),
    ]:
        r += 1
        label(ws, r, 1, name)
        tol(ws, r, 2, dflt)
        label(ws, r, 3, unit, bg=CWH)
        blank(ws, r, 4); blank(ws, r, 5); blank(ws, r, 6)
        tol_rows[name] = r
    # r = 37..41

    # ── Section 4: Overall result ────────────────────────────────────────────
    r += 2                                          # r = 43
    h1(ws, r, 1, 6, "4.  OVERALL TEST RESULT")
    r += 1                                          # r = 44  auto
    label(ws, r, 1, "Auto-calculated cross-check")
    auto_formula = (
        "=IF(COUNTIF('4. Analogue Inputs'!M:M,\"FAIL\")"
        "+COUNTIF('7. Element Results'!G:G,\"FAIL\")>0,\"FAIL\","
        "IF(COUNTIF('4. Analogue Inputs'!M:M,\"PASS\")"
        "+COUNTIF('7. Element Results'!G:G,\"PASS\")>0,\"PASS\",\"INCOMPLETE\"))"
    )
    c = ws.cell(row=r, column=2, value=auto_formula)
    c.font = _font(bold=True, size=12); c.alignment = _align("center", "center")
    c.border = _thin(); c.fill = _fill(CWH)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    label(ws, r, 5, "(auto only — see signed result)", bg=CWH, cols=2)

    r += 1                                          # r = 45  signed
    label(ws, r, 1, "SIGNED OVERALL RESULT", bold=True)
    signed_row = r
    c = ws.cell(row=r, column=2, value="PASS")
    c.font = _font(bold=True, size=12, color=CUI); c.alignment = _align("center", "center")
    c.border = _thin(); c.fill = _fill(CY)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    label(ws, r, 5, "PASS / PASS WITH NOTES / FAIL", bg=CWH, cols=2)

    dv = DataValidation(type="list", formula1='"PASS,PASS WITH NOTES,FAIL"')
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=signed_row, column=2))

    return ws, tol_rows, ctr1_pri, ctr1_sec, ctr2_pri, ctr2_sec, ctrn_pri, ctrn_sec, ptr_pri, ptr_sec, signed_row


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 – EQUIPMENT & REFERENCES
# ═══════════════════════════════════════════════════════════════════════════════
def build_equip(wb):
    ws = wb.create_sheet("2. Equip & Refs")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [30, 40, 22, 16]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 4, "EQUIPMENT & REFERENCES")

    r += 2
    h1(ws, r, 1, 4, "TEST EQUIPMENT USED")
    r += 1
    for col, hdr in enumerate(["Equipment", "Manufacturer / Model", "Asset / Serial No.", "Cal. Due"], 1):
        chdr(ws, r, col, hdr)
    for _ in range(6):
        r += 1
        for col in range(1, 5):
            inp(ws, r, col)

    r += 2
    h1(ws, r, 1, 4, "DOCUMENTS & DRAWINGS")
    r += 1
    for col, hdr in enumerate(["Document", "Title / Description", "Document No.", "Rev"], 1):
        chdr(ws, r, col, hdr)
    docs = ["Relay Settings File (.sel / QuickSet)",
            "AC Elementary / Schematic",
            "DC Control Schematic",
            "Relay Panel Wiring Diagram",
            "Protection Coordination Study",
            "Maintenance Work Order"]
    for doc in docs:
        r += 1
        label(ws, r, 1, doc)
        for col in range(2, 5):
            inp(ws, r, col)
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 – PRE-TEST
# ═══════════════════════════════════════════════════════════════════════════════
def build_pretest(wb):
    ws = wb.create_sheet("3. Pre-Test")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [40, 32, 16, 22]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 4, "PRE-TEST CHECKS")

    def check_section(ws, r, title, rows):
        r += 2
        h1(ws, r, 1, 4, title)
        r += 1
        for col, hdr in enumerate(["Check", "Expected / Action", "Result", "Notes"], 1):
            chdr(ws, r, col, hdr)
        for chk, exp in rows:
            r += 1
            label(ws, r, 1, chk)
            label(ws, r, 2, exp, bg=CCR)
            inp(ws, r, 3)
            inp(ws, r, 4)
        return r

    r = check_section(ws, r, "3.1  COMMUNICATIONS & IDENTIFICATION", [
        ("Connect serial cable (SEL-C662) or Ethernet to relay",
         "Terminal prompt '=' or '-' appears"),
        ("Issue QUIT <Enter> — verify response header",
         "RID and TID match asset register"),
        ("Issue ACC <Enter> + Level 1 password",
         "Access Level 1 granted"),
        ("ENABLED LED illuminated (green, front panel)",
         "LED ON = relay healthy and in service"),
        ("Firmware ID (FID) — record string",
         "Matches approved firmware version"),
        ("Checksum ID (CID) — record string",
         "Matches as-left checksum from last test"),
    ])

    # 3.2  STA self-test
    r += 2
    h1(ws, r, 1, 4, "3.2  RELAY SELF-TEST STATUS  (issue STA <Enter> at serial port)")
    r += 1
    note_row(ws, r, 1, 4,
             "All items should show OK. WARN items require STA C to clear; persistent FAIL items require relay "
             "replacement / SEL support. Record measured voltages for power supply rails.", height=36)
    r += 1
    for col, hdr in enumerate(["Self-Test Item", "Normal Range / Description",
                                "Status / Measured Value", "Action if WARN or FAIL"], 1):
        chdr(ws, r, col, hdr)

    sta_items = [
        ("FPGA",             "Mainboard FPGA — power-up and run-time CRC"),
        ("GPSB",             "Back-plane (board-to-board) communications"),
        ("HMI",              "Front-panel display / HMI interface"),
        ("RAM",              "External and internal RAM read/write test"),
        ("ROM / NON_VOL",    "Data flash checksum (non-volatile memory)"),
        ("CR_RAM",           "Critical RAM — active settings checksum"),
        ("CLOCK",            "Real-time clock chip communications"),
        ("CID_FILE",         "CID (configured IED description) file access"),
        ("CARD_C / CARD_D / CARD_E", "Option card ID register checks"),
        ("CT Board (CURRENT)",       "CT A/D board present; offset −50 to +50 mV"),
        ("VT Board (CARD_E)",        "VT A/D board present; offset −50 to +50 mV (if fitted)"),
        ("RTD module (if fitted)",   "Internal/external RTD comms OK"),
        ("+0.9 V supply",   "0.855 – 0.945 V"),
        ("+1.2 V supply",   "1.152 – 1.248 V"),
        ("+1.5 V supply",   "1.35 – 1.65 V"),
        ("+1.8 V supply",   "1.71 – 1.89 V"),
        ("+2.5 V supply",   "2.32 – 2.68 V"),
        ("+3.3 V supply",   "3.07 – 3.53 V"),
        ("+3.75 V supply",  "3.48 – 4.02 V"),
        ("+5.0 V supply",   "4.65 – 5.35 V"),
        ("−1.25 V supply",  "−1.16 to −1.34 V"),
        ("−5.0 V supply",   "−4.65 to −5.35 V"),
        ("BATT (clock battery)", "2.3 – 3.5 V"),
        ("HALARM relay word bit", "Should be 0 (not asserted)"),
    ]
    for item, desc in sta_items:
        r += 1
        label(ws, r, 1, item)
        label(ws, r, 2, desc, bg=CCR)
        inp(ws, r, 3)
        inp(ws, r, 4)

    r = check_section(ws, r, "3.3  CLOCK / TIME SYNCHRONISATION", [
        ("Issue TIM <Enter> — relay time",         "Within 5 s of reference clock"),
        ("Issue DAT <Enter> — relay date",         "Correct calendar date"),
        ("Time source shown (Internal / IRIG-B)",  "IRIG-B if time-code input connected"),
        ("Synchronise clock if out of tolerance",  "TIM hh:mm:ss   DAT mm/dd/yy"),
    ])

    r = check_section(ws, r, "3.4  EVENT RECORDS (pre-test baseline)", [
        ("Issue HIS <Enter> — review event history",     "No unexplained trips/alarms since last test"),
        ("Issue SER <Enter> — review SER log",           "No unexpected HALARM / SALARM events"),
        ("Issue SUM <Enter> — event summary",            "Review; record most recent event number"),
        ("Issue EVE <Enter> — most recent event report", "Review for unexpected operation"),
    ])

    r = check_section(ws, r, "3.5  SETTINGS VERIFICATION", [
        ("Issue SHOW <Enter> — download all settings",        "Settings match approved file"),
        ("CTR1 (W1 CT ratio) matches Cover sheet",            "As per relay data"),
        ("CTR2 (W2 CT ratio) matches Cover sheet",            "As per relay data"),
        ("CTRN (neutral CT ratio) matches Cover sheet",       "As per relay data"),
        ("PTR (VT ratio) matches Cover sheet (if VT fitted)", "As per relay data"),
        ("MVA, VWDG1, VWDG2 match transformer nameplate",    "As per nameplate"),
        ("W1CTC / W2CTC non-zero for grounded-WYE windings", "See commissioning test worksheet"),
        ("Active setting group correct",                      "Group 1 or as applicable"),
        ("Settings unchanged since last test",                "No unauthorised changes"),
    ])

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 – ANALOGUE INPUTS
# ═══════════════════════════════════════════════════════════════════════════════
def build_analogue(wb):
    ws = wb.create_sheet("4. Analogue Inputs")
    ws.sheet_view.showGridLines = False
    col_widths = [22, 14, 13, 16, 16, 13, 10, 10, 8, 8, 8, 8, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    r = 1
    title_row(ws, r, 1, 13, "ANALOGUE INPUT VERIFICATION — PRIMARY VALUE CHECK")
    r += 1
    note_row(ws, r, 1, 13,
             "Verify the relay reports the correct PRIMARY value (Injected Secondary × CT or VT ratio). "
             "Use METER <Enter> or front-panel MAIN > Meters > Fundamental. "
             "Compare to Expected Primary calculated here from the ratios set on Cover sheet.", height=36)

    HDRS = ["Channel", "Injected Sec.\n(A or V)", "Inj. Angle\n(°)",
            "Expected Primary\n(auto)", "Relay Primary\nReading",
            "Relay Angle\nReading", "Mag\nErr %", "Angle\nErr °",
            "Mag\nTol %", "Ang\nTol °", "", "", "P/F"]

    def analogue_block(ws, r, title, ratio_name, tol_mag_name, channels):
        r += 2
        h1(ws, r, 1, 13, title)
        r += 1
        label(ws, r, 1,
              f"Ratio: {ratio_name} defined on Cover sheet  |  "
              f"Tolerance: {tol_mag_name}", bg=CLB, cols=13)
        r += 1
        for col, hdr in enumerate(HDRS, 1):
            chdr(ws, r, col, hdr)
        for ch in channels:
            r += 1
            label(ws, r, 1, ch, bg=CLB)
            inp(ws, r, 2)   # injected secondary
            inp(ws, r, 3)   # injected angle
            # Expected Primary
            c = ws.cell(row=r, column=4,
                        value=f'=IFERROR(IF(B{r}="","",B{r}*{ratio_name}),"—")')
            c.fill = _fill(CGR); c.font = _font(bold=True)
            c.border = _thin(); c.alignment = _align("center", "center")
            inp(ws, r, 5)   # relay primary
            inp(ws, r, 6)   # relay angle
            # Mag Err %
            c = ws.cell(row=r, column=7,
                        value=f'=IFERROR(IF(OR(D{r}="",E{r}=""),"",(E{r}-D{r})/D{r}*100),"—")')
            c.fill = _fill(CGR); c.border = _thin()
            c.number_format = "0.00"; c.alignment = _align("center", "center")
            # Angle Err °
            c = ws.cell(row=r, column=8,
                        value=f'=IFERROR(IF(OR(C{r}="",F{r}=""),"",F{r}-C{r}),"—")')
            c.fill = _fill(CGR); c.border = _thin()
            c.number_format = "0.0"; c.alignment = _align("center", "center")
            # Mag tol ref
            c = ws.cell(row=r, column=9, value=f"={tol_mag_name}")
            c.fill = _fill(CY); c.border = _thin(); c.alignment = _align("center", "center")
            # Ang tol ref
            c = ws.cell(row=r, column=10, value="=TolAngle_deg")
            c.fill = _fill(CY); c.border = _thin(); c.alignment = _align("center", "center")
            blank(ws, r, 11); blank(ws, r, 12)
            # P/F
            pf = (f'=IFERROR(IF(AND(B{r}="",E{r}=""),"—",'
                  f'IF(OR(AND(ISNUMBER(G{r}),ABS(G{r})>{tol_mag_name}),'
                  f'AND(ISNUMBER(H{r}),ABS(H{r})>TolAngle_deg)),"FAIL","PASS")),"—")')
            c = ws.cell(row=r, column=13, value=pf)
            c.font = _font(bold=True); c.border = _thin()
            c.alignment = _align("center", "center")
        return r

    r = analogue_block(ws, r,
        "4.1  WINDING 1 PHASE CURRENTS  (IAW1, IBW1, ICW1)  [CTR1]",
        "Ratio_CTR1", "TolI_pct",
        ["IAW1 — A-Phase Winding 1", "IBW1 — B-Phase Winding 1", "ICW1 — C-Phase Winding 1"])

    r = analogue_block(ws, r,
        "4.2  WINDING 2 PHASE CURRENTS  (IAW2, IBW2, ICW2)  [CTR2]",
        "Ratio_CTR2", "TolI_pct",
        ["IAW2 — A-Phase Winding 2", "IBW2 — B-Phase Winding 2", "ICW2 — C-Phase Winding 2"])

    r = analogue_block(ws, r,
        "4.3  NEUTRAL CURRENT  (IN)  [CTRN]  — if neutral CT fitted",
        "Ratio_CTRN", "TolI_pct",
        ["IN — Neutral current"])

    r = analogue_block(ws, r,
        "4.4  VOLTAGE INPUTS  (VA, VB, VC)  [PTR]  — if VT card fitted",
        "Ratio_PTR", "TolV_pct",
        ["VA — Phase A voltage", "VB — Phase B voltage", "VC — Phase C voltage"])

    # 4.5 Differential Meter check
    r += 2
    h1(ws, r, 1, 13, "4.5  DIFFERENTIAL METER CHECK  (METER DIF <Enter> or MAIN > Meters > Differential)")
    r += 1
    note_row(ws, r, 1, 13,
             "Under balanced load: IOP (operate) should be ≈ 0.00, IRT (restraint) ≈ load in TAP units. "
             "If MM = IOP/IRT > 0.10 investigate CT polarity, W1CTC/W2CTC compensation, and CT connections.", 40)
    r += 1
    for col, hdr in enumerate(["Diff Element", "IOP\n(Operate, TAP)", "IRT\n(Restraint, TAP)",
                                "MM = IOP/IRT", "Accept if\nMM ≤ 0.10",
                                "2nd Harm\n(TAP)", "5th Harm\n(TAP)",
                                "", "", "", "", "", "Result"], 1):
        chdr(ws, r, col, hdr)
    for elem in ["87-1 (DIFF1)", "87-2 (DIFF2)", "87-3 (DIFF3) if used"]:
        r += 1
        label(ws, r, 1, elem, bg=CLB)
        inp(ws, r, 2); inp(ws, r, 3)
        c = ws.cell(row=r, column=4,
                    value=f'=IFERROR(IF(OR(B{r}="",C{r}=""),"",B{r}/C{r}),"—")')
        c.fill = _fill(CGR); c.border = _thin()
        c.number_format = "0.000"; c.alignment = _align("center", "center")
        label(ws, r, 5, "≤ 0.10", bg=CCR)
        inp(ws, r, 6); inp(ws, r, 7)
        blank(ws, r, 8); blank(ws, r, 9); blank(ws, r, 10); blank(ws, r, 11); blank(ws, r, 12)
        c = ws.cell(row=r, column=13,
                    value=f'=IFERROR(IF(D{r}="","—",IF(AND(ISNUMBER(D{r}),D{r}<=0.10),"PASS","FAIL")),"—")')
        c.font = _font(bold=True); c.border = _thin(); c.alignment = _align("center", "center")

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 – BINARY I/O
# ═══════════════════════════════════════════════════════════════════════════════
def build_binary(wb):
    ws = wb.create_sheet("5. Binary IO")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", [10, 30, 22, 24, 12]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 5, "BINARY I/O VERIFICATION")

    r += 2
    h1(ws, r, 1, 5, "5.1  DIGITAL INPUTS — Pickup Verification")
    r += 1
    note_row(ws, r, 1, 5,
             "Apply rated control voltage to each input. Use front-panel MAIN > Targets > Row 21 "
             "or serial TARGET command to verify IN101/IN102 state changes 0→1.", 30)
    r += 1
    for col, hdr in enumerate(["Input No.", "Function Label", "Control Voltage Applied", "Relay Word Bit", "Result"], 1):
        chdr(ws, r, col, hdr)
    inputs = [
        ("IN101", "Digital Input 1 — edit to match schematics", "IN101"),
        ("IN102", "Digital Input 2 — edit to match schematics", "IN102"),
        ("IN301", "Optional DI — Input 1 (if card fitted)",     "IN301"),
        ("IN302", "Optional DI — Input 2 (if card fitted)",     "IN302"),
        ("IN303", "Optional DI — Input 3 (if card fitted)",     "IN303"),
        ("IN304", "Optional DI — Input 4 (if card fitted)",     "IN304"),
    ]
    for inp_id, func, rwb in inputs:
        r += 1
        label(ws, r, 1, inp_id, bg=CLB, bold=True)
        inp(ws, r, 2, func)
        inp(ws, r, 3)
        label(ws, r, 4, rwb, bg=CCR)
        inp(ws, r, 5)

    r += 2
    h1(ws, r, 1, 5, "5.2  DIGITAL OUTPUTS — Continuity Test")
    r += 1
    note_row(ws, r, 1, 5,
             "Force each output via serial (OUT101:=1 etc.) or CONTROL menu. "
             "Verify continuity through to trip coil / alarm annunciator. "
             "Restore all SELOGIC equations after test.", 30)
    r += 1
    for col, hdr in enumerate(["Output No.", "Function Label", "Force Command", "Expected Result", "P/F"], 1):
        chdr(ws, r, col, hdr)
    outputs = [
        ("OUT101", "TRIP / Main trip (edit label)", "OUT101:=1", "Contact closes — continuity to trip coil"),
        ("OUT102", "Alarm / Close (edit label)",    "OUT102:=1", "Contact closes — continuity verified"),
        ("OUT103", "Auxiliary output (edit label)", "OUT103:=1", "Contact closes — continuity verified"),
        ("OUT301", "Optional DO — 1 (if card)",     "OUT301:=1", "Contact closes — continuity verified"),
        ("OUT302", "Optional DO — 2 (if card)",     "OUT302:=1", "Contact closes — continuity verified"),
    ]
    for out_id, func, cmd, exp in outputs:
        r += 1
        label(ws, r, 1, out_id, bg=CLB, bold=True)
        inp(ws, r, 2, func)
        label(ws, r, 3, cmd, bg=CCR)
        label(ws, r, 4, exp, bg=CCR)
        inp(ws, r, 5)

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6 – ELEMENT TRACKER
# ═══════════════════════════════════════════════════════════════════════════════
def build_tracker(wb):
    ws = wb.create_sheet("6. Element Tracker")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [6, 18, 40, 14, 16, 12]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 6, "PROTECTION ELEMENT TRACKER")
    r += 2
    for col, hdr in enumerate(["#", "Element ID", "Description", "Enabled\n(Y / N / NA)", "Test Result Row", "Result"], 1):
        chdr(ws, r, col, hdr)

    elements = [
        # Standard protection
        ("87-1",   "DIFF1 — Current Differential Element 1 (restrained + unrestrained)"),
        ("87-2",   "DIFF2 — Current Differential Element 2 (if fitted/enabled)"),
        ("87-3",   "DIFF3 — Current Differential Element 3 (if fitted/enabled)"),
        ("50P1",   "W1 Phase Instantaneous OC  (50P11, 50P12, 50P13, 50P14)"),
        ("50P2",   "W2 Phase Instantaneous OC  (50P21, 50P22, 50P23, 50P24)"),
        ("50G1",   "W1 Ground Instantaneous OC  (50G11, 50G12)"),
        ("50G2",   "W2 Ground Instantaneous OC  (50G21, 50G22)"),
        ("50Q1",   "W1 Neg-Seq Instantaneous OC  (50Q11, 50Q12)"),
        ("50Q2",   "W2 Neg-Seq Instantaneous OC  (50Q21, 50Q22)"),
        ("51P1",   "W1 Phase Time OC  (51P1 — U1..U5, C1..C5 or user curve)"),
        ("51P2",   "W2 Phase Time OC  (51P2)"),
        ("51G1",   "W1 Ground Time OC  (51G1)"),
        ("51G2",   "W2 Ground Time OC  (51G2)"),
        ("51Q1",   "W1 Neg-Seq Time OC  (51Q1)"),
        ("51Q2",   "W2 Neg-Seq Time OC  (51Q2)"),
        ("BF1",    "Breaker Failure — Winding 1  (BFI1, BFT1)"),
        ("BF2",    "Breaker Failure — Winding 2  (BFI2, BFT2)"),
        # Optional
        ("50N1",   "Neutral Instantaneous OC 1  (50N11, 50N12) — optional neutral CT"),
        ("50N2",   "Neutral Instantaneous OC 2  (50N21, 50N22) — optional"),
        ("51N1",   "Neutral Time OC  (51N1) — optional"),
        ("REF",    "Restricted Earth Fault  (REF1P, REF1F) — optional"),
        ("27P1",   "Undervoltage 1  (27P1, 27P1T) — requires VT card"),
        ("27P2",   "Undervoltage 2  (27P2, 27P2T) — requires VT card"),
        ("59P1",   "Overvoltage 1  (59P1, 59P1T) — requires VT card"),
        ("59P2",   "Overvoltage 2  (59P2, 59P2T) — requires VT card"),
        ("59Q",    "Neg-Seq Overvoltage  (59Q1, 59Q1T) — requires VT card"),
        ("24D1",   "Volts/Hertz Definite-Time  (24D1, 24D1T) — requires VT card"),
        ("24C2",   "Volts/Hertz Inverse-Time  (24C2, 24C2T) — requires VT card"),
        ("32",     "Directional Power  (3PWR1, 3PWR2) — requires VT card"),
        ("81D1",   "Frequency — element 1  (81D1T) — requires VT card"),
        ("81D2",   "Frequency — element 2  (81D2T) — requires VT card"),
        ("81D3",   "Frequency — element 3  (81D3T) — requires VT card"),
        ("81D4",   "Frequency — element 4  (81D4T) — requires VT card"),
        ("RTD",    "RTD Thermal Protection  (RTDT, RTDFLT) — requires RTD card"),
    ]

    for i, (eid, desc) in enumerate(elements, 1):
        r += 1
        bg = CAL if i % 2 == 0 else CWH
        c = ws.cell(row=r, column=1, value=i)
        c.font = _font(bold=True); c.border = _thin(); c.fill = _fill(bg)
        c.alignment = _align("center", "center")
        label(ws, r, 2, eid, bg=CLB, bold=True)
        label(ws, r, 3, desc, bg=bg)
        inp(ws, r, 4)   # Y/N/NA
        inp(ws, r, 5)   # row ref
        inp(ws, r, 6)   # result

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 7 – ELEMENT RESULTS
# ═══════════════════════════════════════════════════════════════════════════════
def build_results(wb):
    ws = wb.create_sheet("7. Element Results")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFG", [12, 36, 14, 14, 12, 14, 10]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 7, "PROTECTION ELEMENT TEST RESULTS")
    r += 1
    note_row(ws, r, 1, 7,
             "Setting column = relay pickup setting (secondary A, V, or %). "
             "Measured Pickup = value at which relay output asserted during injection. "
             "Pickup Err % = auto-calculated. Result = PASS if |Err| ≤ TolPickup_pct.", 32)
    r += 1
    for col, hdr in enumerate(["Element", "Description",
                                "Pickup Setting\n(sec A, V, %)",
                                "Measured\nPickup",
                                "Pickup\nErr %",
                                "Measured\nTime (s)", "Result"], 1):
        chdr(ws, r, col, hdr)

    test_elements = [
        ("87-1 R",   "Differential restrained pickup — 87AP × TAP1  (A secondary)"),
        ("87-1 U",   "Differential unrestrained pickup — U87P × TAP1  (A secondary)"),
        ("87-2 R",   "Differential restrained pickup — 87AP × TAP2  (if enabled)"),
        ("50P11",    "W1 Phase Inst OC — Stage 1  (A secondary)"),
        ("50P12",    "W1 Phase Inst OC — Stage 2  (A secondary)"),
        ("50P21",    "W2 Phase Inst OC — Stage 1  (A secondary)"),
        ("50P22",    "W2 Phase Inst OC — Stage 2  (A secondary)"),
        ("50G11",    "W1 Ground Inst OC — Stage 1  (A secondary)"),
        ("50G21",    "W2 Ground Inst OC — Stage 1  (A secondary)"),
        ("50Q11",    "W1 Neg-Seq Inst OC — Stage 1  (A secondary)"),
        ("50Q21",    "W2 Neg-Seq Inst OC — Stage 1  (A secondary)"),
        ("51P1",     "W1 Phase TOC — pickup  (A secondary)"),
        ("51G1",     "W1 Ground TOC — pickup  (A secondary)"),
        ("51Q1",     "W1 Neg-Seq TOC — pickup  (A secondary)"),
        ("51P2",     "W2 Phase TOC — pickup  (A secondary)"),
        ("51G2",     "W2 Ground TOC — pickup  (A secondary)"),
        ("51Q2",     "W2 Neg-Seq TOC — pickup  (A secondary)"),
        ("50N11",    "Neutral Inst OC 1  (A secondary, if enabled)"),
        ("51N1",     "Neutral TOC — pickup  (A secondary, if enabled)"),
        ("REF1",     "Restricted Earth Fault  (A secondary, if enabled)"),
        ("27P1",     "Undervoltage 1  (V secondary, if enabled)"),
        ("27P2",     "Undervoltage 2  (V secondary, if enabled)"),
        ("59P1",     "Overvoltage 1  (V secondary, if enabled)"),
        ("59P2",     "Overvoltage 2  (V secondary, if enabled)"),
        ("24D1",     "Volts/Hertz definite-time  (%, if enabled)"),
        ("24C2",     "Volts/Hertz inverse-time  (%, if enabled)"),
        ("32-1",     "Directional Power 1  (kVA secondary, if enabled)"),
        ("81D1",     "Frequency 1  (Hz, if enabled)"),
        ("81D2",     "Frequency 2  (Hz, if enabled)"),
        ("RTD",      "RTD trip threshold  (°C, if fitted)"),
        ("BFD1",     "Breaker Failure 1 — BFD1 timer  (s)"),
        ("BFD2",     "Breaker Failure 2 — BFD2 timer  (s)"),
    ]

    for i, (eid, desc) in enumerate(test_elements, 1):
        r += 1
        bg = CAL if i % 2 == 0 else CWH
        label(ws, r, 1, eid, bg=CLB, bold=True)
        label(ws, r, 2, desc, bg=bg)
        inp(ws, r, 3)   # setting
        inp(ws, r, 4)   # measured pickup
        c = ws.cell(row=r, column=5,
                    value=f'=IFERROR(IF(OR(C{r}="",D{r}=""),"",(D{r}-C{r})/C{r}*100),"—")')
        c.fill = _fill(CGR); c.border = _thin()
        c.number_format = "0.00"; c.alignment = _align("center", "center")
        inp(ws, r, 6)   # time measured
        result = (f'=IFERROR(IF(AND(C{r}="",D{r}=""),"—",'
                  f'IF(AND(ISNUMBER(E{r}),ABS(E{r})<=TolPickup_pct),"PASS","FAIL")),"—")')
        c = ws.cell(row=r, column=7, value=result)
        c.font = _font(bold=True); c.border = _thin(); c.alignment = _align("center", "center")

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 8 – POST-TEST
# ═══════════════════════════════════════════════════════════════════════════════
def build_posttest(wb):
    ws = wb.create_sheet("8. Post-Test")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [40, 30, 14, 20]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 4, "POST-TEST CHECKS")

    def check_section(ws, r, title, rows):
        r += 2
        h1(ws, r, 1, 4, title)
        r += 1
        for col, hdr in enumerate(["Check", "Expected / Action", "Result", "Notes"], 1):
            chdr(ws, r, col, hdr)
        for chk, exp in rows:
            r += 1
            label(ws, r, 1, chk)
            label(ws, r, 2, exp, bg=CCR)
            inp(ws, r, 3)
            inp(ws, r, 4)
        return r

    r = check_section(ws, r, "8.1  SETTINGS RESTORE & VERIFICATION", [
        ("Issue SHOW <Enter> — verify settings unchanged",    "Match approved settings file"),
        ("All test-forced SELOGIC equations restored",        "OUT101:=, SV variables correct"),
        ("Active setting group restored to normal",           "Group 1 or as applicable"),
        ("ENABLED LED illuminated (green)",                   "Relay in service"),
    ])

    r = check_section(ws, r, "8.2  FINAL SELF-TEST  (STA Command)", [
        ("Issue STA <Enter> — all items OK",                  "No WARN or FAIL items"),
        ("HALARM = 0 (no hardware alarm)",                    "Relay word bit confirmed 0"),
        ("SALARM = 0 (no software alarm)",                    "Normal operating condition"),
        ("Relay Enabled confirmed",                           "'Relay Enabled' shown in STA output"),
    ])

    # 8.3 Clear data buffers
    r += 2
    h1(ws, r, 1, 4, "8.3  CLEAR TEST DATA BUFFERS")
    r += 1
    note_row(ws, r, 1, 4,
             "Issue the commands below at Access Level 1 or higher to clear data "
             "generated during testing, preventing contamination of operational records.", 30)
    r += 1
    for col, hdr in enumerate(["Serial Command", "Action Performed", "Completed", "Notes"], 1):
        chdr(ws, r, col, hdr)
    for cmd, action in [
        ("SUM R <Enter>", "Reset event report and summary command buffers"),
        ("SER R <Enter>", "Reset Sequential Events Record buffer"),
        ("LDP C <Enter>", "Clear Load Profile Data"),
        ("HIS C <Enter>", "Clear event history (if required by procedure)"),
    ]:
        r += 1
        label(ws, r, 1, cmd, bg=CLB, bold=True)
        label(ws, r, 2, action, bg=CCR)
        inp(ws, r, 3); inp(ws, r, 4)

    r = check_section(ws, r, "8.4  PLANT RESTORATION", [
        ("Remove test connections from CT / VT terminals",        "Field CTs and VTs reconnected"),
        ("Restore control voltage wiring",                        "All control circuits normal"),
        ("Remove CT secondary shorting links",                    "CT circuits not open-circuited"),
        ("Verify relay measuring load current — METER command",   "Compare to substation panel meters"),
        ("METER DIF check — IOP/IRT ≤ 0.10 under load",          "Differential balance confirmed on load"),
        ("Cancel all outstanding LOTO permits",                   "All permits cleared"),
        ("Notify operations — relay returned to service",         "Operations informed and recorded"),
    ])

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 9 – FILES & SIGN-OFF
# ═══════════════════════════════════════════════════════════════════════════════
def build_signoff(wb, signed_row):
    ws = wb.create_sheet("9. Files & Sign-Off")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [28, 38, 20, 16]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 4, "FILES INDEX & SIGN-OFF")

    r += 2
    h1(ws, r, 1, 4, "OVERALL TEST RESULT (from Cover sheet)")
    r += 1
    label(ws, r, 1, "Signed Overall Result", bold=True)
    c = ws.cell(row=r, column=2, value=f"='1. Cover'!B{signed_row}")
    c.font = _font(bold=True, size=13); c.alignment = _align("center", "center")
    c.border = _thin(); c.fill = _fill(CY)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)

    r += 2
    h1(ws, r, 1, 4, "TEST ARTEFACTS INDEX")
    r += 1
    for col, hdr in enumerate(["Artefact", "File Name / SharePoint Location", "Date Saved", "Notes"], 1):
        chdr(ws, r, col, hdr)
    for art in ["Relay settings file (.sel / QuickSet project)",
                "Pre-test event records (HIS, SER, EVE, SUM)",
                "Post-test event records",
                "Completed results template (this file)",
                "Oscillography / disturbance records",
                "Test photographs / site notes"]:
        r += 1
        label(ws, r, 1, art)
        inp(ws, r, 2); inp(ws, r, 3); inp(ws, r, 4)

    r += 2
    h1(ws, r, 1, 4, "DEFECTS / FINDINGS LOG")
    r += 1
    for col, hdr in enumerate(["#", "Defect Description", "Action Taken / Work Order", "Closed"], 1):
        chdr(ws, r, col, hdr)
    for i in range(1, 7):
        r += 1
        c = ws.cell(row=r, column=1, value=i)
        c.border = _thin(); c.alignment = _align("center", "center")
        for col in range(2, 5): inp(ws, r, col)

    r += 2
    h1(ws, r, 1, 4, "SIGN-OFF")
    r += 1
    for col, hdr in enumerate(["Role", "Name (print)", "Signature", "Date"], 1):
        chdr(ws, r, col, hdr)
    for role in ["Test Engineer", "Witness (Operations)", "Approving Engineer"]:
        r += 1
        ws.row_dimensions[r].height = 28
        label(ws, r, 1, role, bg=CLB, bold=True)
        inp(ws, r, 2); inp(ws, r, 3); inp(ws, r, 4)

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# DEFINED NAMES
# ═══════════════════════════════════════════════════════════════════════════════
def add_names(wb, ctr1_pri, ctr1_sec, ctr2_pri, ctr2_sec,
              ctrn_pri, ctrn_sec, ptr_pri, ptr_sec, tol_rows):
    sn = "'1. Cover'"
    defs = {
        "Ratio_CTR1":     f"{sn}!$C${ctr1_pri}/{sn}!$C${ctr1_sec}",
        "Ratio_CTR2":     f"{sn}!$C${ctr2_pri}/{sn}!$C${ctr2_sec}",
        "Ratio_CTRN":     f"{sn}!$C${ctrn_pri}/{sn}!$C${ctrn_sec}",
        "Ratio_PTR":      f"{sn}!$C${ptr_pri}/{sn}!$C${ptr_sec}",
        "TolV_pct":       f"{sn}!$B${tol_rows['Voltage Magnitude Tolerance']}",
        "TolI_pct":       f"{sn}!$B${tol_rows['Current Magnitude Tolerance']}",
        "TolAngle_deg":   f"{sn}!$B${tol_rows['Angle Tolerance']}",
        "TolPickup_pct":  f"{sn}!$B${tol_rows['Pickup Tolerance']}",
        "TolTime_pct":    f"{sn}!$B${tol_rows['Timing Tolerance']}",
    }
    for name, attr in defs.items():
        dn = DefinedName(name=name, attr_text=attr)
        wb.defined_names[name] = dn


# ═══════════════════════════════════════════════════════════════════════════════
# CONDITIONAL FORMATTING
# ═══════════════════════════════════════════════════════════════════════════════
def add_cf(wb):
    pf = PatternFill("solid", fgColor=CPF)
    ff = PatternFill("solid", fgColor=CFL)
    pfont = Font(bold=True, color=CPFF, name="Calibri")
    ffont = Font(bold=True, color=CFFL, name="Calibri")

    targets = {
        "4. Analogue Inputs": "M1:M300",
        "7. Element Results": "G1:G200",
        "6. Element Tracker": "F1:F200",
    }
    for sheet_name, rng in targets.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.conditional_formatting.add(rng, CellIsRule("equal", ['"PASS"'], fill=pf, font=pfont))
            ws.conditional_formatting.add(rng, CellIsRule("equal", ['"FAIL"'], fill=ff, font=ffont))


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()

    (cover_ws, tol_rows,
     ctr1_pri, ctr1_sec, ctr2_pri, ctr2_sec,
     ctrn_pri, ctrn_sec, ptr_pri, ptr_sec,
     signed_row) = build_cover(wb)

    build_equip(wb)
    build_pretest(wb)
    build_analogue(wb)
    build_binary(wb)
    build_tracker(wb)
    build_results(wb)
    build_posttest(wb)
    build_signoff(wb, signed_row)

    add_names(wb, ctr1_pri, ctr1_sec, ctr2_pri, ctr2_sec,
              ctrn_pri, ctrn_sec, ptr_pri, ptr_sec, tol_rows)
    add_cf(wb)

    for ws in wb.worksheets:
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = "portrait"
        ws.page_margins.left = 0.4
        ws.page_margins.right = 0.4
        ws.page_margins.top = 0.6
        ws.page_margins.bottom = 0.6

    out = "/home/michael/claudeProject /protection-relay-checksheet/SEL-787_Test_Results_Template.xlsx"
    wb.save(out)
    print(f"Saved: {out}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Cover signed result row: {signed_row}")
    print(f"CTR1 ratio rows: C{ctr1_pri} / C{ctr1_sec}")
    print(f"Defined names: {list(wb.defined_names.keys())}")


if __name__ == "__main__":
    main()
