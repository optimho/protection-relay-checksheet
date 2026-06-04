#!/usr/bin/env python3
"""Generate GE Grid Solutions 889 Generator Protection System Test Results Template (.xlsx)."""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import CellIsRule

# ── Colours ──────────────────────────────────────────────────────────────────
CB  = "1F4E79"   # deep blue  – headers
CLB = "D9E2F3"   # light blue – labels
CY  = "FFF2CC"   # yellow     – tolerance / caution
CCR = "FFF8E7"   # cream      – expected / reference
CPF = "C6EFCE"; CPFF = "006100"   # PASS
CFL = "FFC7CE"; CFFL = "9C0006"   # FAIL
CGR = "D9D9D9"   # grey       – auto-calculated
CAL = "F2F2F2"   # alt row
CUI = "0000FF"   # blue text  – user input
CWH = "FFFFFF"

def _fill(c): return PatternFill("solid", fgColor=c)
def _font(bold=False, color="000000", size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def h1(ws, row, col1, col2, text):
    ws.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    c = ws.cell(row=row, column=col1, value=text)
    c.fill = _fill(CB); c.font = _font(bold=True, color="FFFFFF", size=11)
    c.alignment = _align("left", "center"); c.border = _thin()

def label(ws, row, col, text, bg=CLB, bold=False, cols=1, wrap=False):
    if cols > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+cols-1)
    c = ws.cell(row=row, column=col, value=text)
    c.fill = _fill(bg); c.font = _font(bold=bold)
    c.alignment = _align("left", "center", wrap=wrap); c.border = _thin()

def inp(ws, row, col, value=None, cols=1):
    if cols > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+cols-1)
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(color=CUI); c.alignment = _align("left", "center")
    c.border = _thin(); c.fill = _fill(CWH)

def tol(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(color=CUI); c.alignment = _align("center", "center")
    c.border = _thin(); c.fill = _fill(CY)

def chdr(ws, row, col, text, bg=CB, fg="FFFFFF"):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = _fill(bg); c.font = _font(bold=True, color=fg, size=10)
    c.alignment = _align("center", "center", wrap=True); c.border = _thin()

def auto(ws, row, col, formula_or_value, fmt=None, bold=False):
    c = ws.cell(row=row, column=col, value=formula_or_value)
    c.fill = _fill(CGR); c.font = _font(bold=bold)
    c.alignment = _align("center", "center"); c.border = _thin()
    if fmt: c.number_format = fmt

def note_row(ws, row, col1, col2, text, height=32):
    ws.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    c = ws.cell(row=row, column=col1, value=text)
    c.fill = _fill(CY); c.font = _font(italic=True, size=10)
    c.alignment = _align("left", "center", wrap=True); c.border = _thin()
    ws.row_dimensions[row].height = height

def title_row(ws, row, col1, col2, text):
    ws.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    c = ws.cell(row=row, column=col1, value=text)
    c.fill = _fill(CB); c.font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 26

def blank(ws, row, col, bg=CWH):
    c = ws.cell(row=row, column=col)
    c.fill = _fill(bg); c.border = _thin()


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 – COVER
# ═══════════════════════════════════════════════════════════════════════════════
def build_cover(wb):
    ws = wb.active
    ws.title = "1. Cover"
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [30, 22, 14, 14, 14, 14]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 6, "GE MULTILIN 889 GENERATOR PROTECTION SYSTEM — TEST RESULTS TEMPLATE")

    # ── Section 1: Relay ID ──────────────────────────────────────────────────
    r += 2  # r = 3
    h1(ws, r, 1, 6, "1.  RELAY IDENTIFICATION")
    rows_id = ["Relay Type / Model", "Order Code / Part Number", "Serial Number",
               "Firmware Version", "Settings File (EnerVista .8SS)",
               "Protected Equipment (generator tag)", "Site / Substation",
               "Equipment Tag (relay panel)", "Test Date", "Test Engineer"]
    defaults = ["GE Multilin 889"] + [""] * 9
    for lbl_text, dflt in zip(rows_id, defaults):
        r += 1
        label(ws, r, 1, lbl_text)
        inp(ws, r, 2, dflt, cols=5)

    # ── Section 2: Generator / CT / VT data ─────────────────────────────────
    r += 2  # r = 15
    h1(ws, r, 1, 6, "2.  GENERATOR AND CT/VT PARAMETERS")
    r += 1  # r = 16  column headers
    for col, hdr in enumerate(["Parameter", "Setting Label", "Value", "Units", "", ""], 1):
        chdr(ws, r, col, hdr)

    # Note row explaining CT banks
    r += 1
    note_row(ws, r, 1, 6,
             "J1 = Neutral-side (generator) CTs (Slot J Bank 1).  "
             "K1 = Terminal-side (line/bus) CTs (Slot K Bank 1) — used for 87G differential.  "
             "GND = Ground/CBCT input (Slot K).  "
             "Aux VT (Vx): Slot J Bank 2 auxiliary voltage input — enter 0 if not fitted.  "
             "Neutral VT: enter 0 if 64TN not fitted.", 44)

    data_start_row = r + 1  # = 18

    gen_data = [
        ("Generator MVA Rating",                  "MVA Rating",          "",    "MVA"),
        ("Generator Terminal Voltage (L-L)",       "Generator Voltage",   "",    "kV (primary)"),
        ("Phase Rotation",                         "Phase Rotation",      "ABC", "ABC / ACB"),
        ("J1 (Neutral-side) CT Primary",           "CTR J1 Primary",      "",    "A primary"),
        ("J1 (Neutral-side) CT Secondary",         "CTR J1 Secondary",    "1",   "A secondary"),
        ("K1 (Terminal-side) CT Primary",          "CTR K1 Primary",      "",    "A primary"),
        ("K1 (Terminal-side) CT Secondary",        "CTR K1 Secondary",    "1",   "A secondary"),
        ("Ground CT (CBCT) Primary",               "CTR GND Primary",     "",    "A primary"),
        ("Ground CT (CBCT) Secondary",             "CTR GND Secondary",   "1",   "A secondary"),
        ("Main VT Primary",                        "PTR Primary",         "",    "kV primary"),
        ("Main VT Secondary",                      "PTR Secondary",       "",    "V secondary"),
        ("Auxiliary VT (Vx) Primary",              "PTR Aux Primary",     "",    "kV primary  (0 = not fitted)"),
        ("Auxiliary VT (Vx) Secondary",            "PTR Aux Secondary",   "",    "V secondary"),
        ("Neutral VT Primary  (64TN)",             "PTR Neut Primary",    "",    "kV primary  (0 = not fitted)"),
        ("Neutral VT Secondary  (64TN)",           "PTR Neut Secondary",  "",    "V secondary"),
    ]

    for lbl_text, slbl, dflt, unit in gen_data:
        r += 1
        label(ws, r, 1, lbl_text)
        label(ws, r, 2, slbl, bg=CCR)
        inp(ws, r, 3, dflt)
        label(ws, r, 4, unit, bg=CWH)
        blank(ws, r, 5); blank(ws, r, 6)

    # Row numbers for ratio defined names (column C):
    # index 3  = J1 primary  → data_start_row + 3
    ct_j1_pri  = data_start_row + 3   # 21
    ct_j1_sec  = data_start_row + 4   # 22
    ct_k1_pri  = data_start_row + 5   # 23
    ct_k1_sec  = data_start_row + 6   # 24
    ct_gnd_pri = data_start_row + 7   # 25
    ct_gnd_sec = data_start_row + 8   # 26
    vt_main_pri = data_start_row + 9  # 27
    vt_main_sec = data_start_row + 10 # 28
    vt_aux_pri  = data_start_row + 11 # 29
    vt_aux_sec  = data_start_row + 12 # 30
    vt_neut_pri = data_start_row + 13 # 31
    vt_neut_sec = data_start_row + 14 # 32

    # ── Section 3: Tolerances ────────────────────────────────────────────────
    r += 2  # r = 35
    h1(ws, r, 1, 6, "3.  TEST TOLERANCES  (yellow cells are editable)")
    tol_rows = {}
    for name, dflt, unit in [
        ("Voltage Magnitude Tolerance",  1.0, "%"),
        ("Current Magnitude Tolerance",  1.0, "%"),
        ("Angle Tolerance",              2.0, "degrees"),
        ("Pickup Tolerance",             5.0, "%"),
        ("Timing Tolerance",             5.0, "%"),
    ]:
        r += 1
        label(ws, r, 1, name)
        tol(ws, r, 2, dflt)
        label(ws, r, 3, unit, bg=CWH)
        blank(ws, r, 4); blank(ws, r, 5); blank(ws, r, 6)
        tol_rows[name] = r

    # ── Section 4: Overall result ────────────────────────────────────────────
    r += 2  # r = 43
    h1(ws, r, 1, 6, "4.  OVERALL TEST RESULT")
    r += 1  # r = 44
    label(ws, r, 1, "Auto-calculated cross-check")
    auto_formula = (
        "=IF(COUNTIF('4. Analogue Inputs'!M:M,\"FAIL\")"
        "+COUNTIF('7. Element Results'!G:G,\"FAIL\")>0,\"FAIL\","
        "IF(COUNTIF('4. Analogue Inputs'!M:M,\"PASS\")"
        "+COUNTIF('7. Element Results'!G:G,\"PASS\")>0,\"PASS\",\"INCOMPLETE\"))"
    )
    c = ws.cell(row=r, column=2, value=auto_formula)
    c.font = _font(bold=True, size=12); c.alignment = _align("center", "center")
    c.border = _thin(); c.fill = _fill(CWH)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    label(ws, r, 5, "(auto only — see signed result)", bg=CWH, cols=2)

    r += 1  # r = 45
    label(ws, r, 1, "SIGNED OVERALL RESULT", bold=True)
    signed_row = r
    c = ws.cell(row=r, column=2, value="PASS")
    c.font = _font(bold=True, size=12, color=CUI); c.alignment = _align("center", "center")
    c.border = _thin(); c.fill = _fill(CY)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    label(ws, r, 5, "PASS / PASS WITH NOTES / FAIL", bg=CWH, cols=2)

    dv = DataValidation(type="list", formula1='"PASS,PASS WITH NOTES,FAIL"')
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=signed_row, column=2))

    return (ws, tol_rows,
            ct_j1_pri, ct_j1_sec, ct_k1_pri, ct_k1_sec,
            ct_gnd_pri, ct_gnd_sec,
            vt_main_pri, vt_main_sec, vt_aux_pri, vt_aux_sec,
            vt_neut_pri, vt_neut_sec, signed_row)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 – EQUIPMENT & REFERENCES
# ═══════════════════════════════════════════════════════════════════════════════
def build_equip(wb):
    ws = wb.create_sheet("2. Equip & Refs")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [30, 40, 22, 16]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 4, "EQUIPMENT & REFERENCES")

    r += 2
    h1(ws, r, 1, 4, "TEST EQUIPMENT USED")
    r += 1
    for col, hdr in enumerate(["Equipment", "Manufacturer / Model", "Asset / Serial No.", "Cal. Due"], 1):
        chdr(ws, r, col, hdr)
    for _ in range(6):
        r += 1
        for col in range(1, 5): inp(ws, r, col)

    r += 2
    h1(ws, r, 1, 4, "DOCUMENTS & DRAWINGS")
    r += 1
    for col, hdr in enumerate(["Document", "Title / Description", "Document No.", "Rev"], 1):
        chdr(ws, r, col, hdr)
    for doc in ["Relay Settings File (EnerVista .8SS project file)",
                "AC Elementary / Schematic",
                "DC Control Schematic",
                "Generator Protection Single-Line Diagram",
                "Protection Coordination Study",
                "Maintenance Work Order"]:
        r += 1
        label(ws, r, 1, doc)
        for col in range(2, 5): inp(ws, r, col)
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 – PRE-TEST
# ═══════════════════════════════════════════════════════════════════════════════
def build_pretest(wb):
    ws = wb.create_sheet("3. Pre-Test")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [40, 36, 16, 22]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 4, "PRE-TEST CHECKS")

    def check_section(ws, r, title, rows):
        r += 2
        h1(ws, r, 1, 4, title)
        r += 1
        for col, hdr in enumerate(["Check", "Expected / Action", "Result", "Notes"], 1):
            chdr(ws, r, col, hdr)
        for chk, exp in rows:
            r += 1
            label(ws, r, 1, chk)
            label(ws, r, 2, exp, bg=CCR)
            inp(ws, r, 3); inp(ws, r, 4)
        return r

    r = check_section(ws, r, "3.1  COMMUNICATIONS & IDENTIFICATION", [
        ("Connect EnerVista 8 Series via Ethernet (RJ45) or USB",
         "EnerVista connects — relay appears in device list"),
        ("EnerVista: right-click device → Quick Connect",
         "Connection successful — no comms error"),
        ("Record relay Order Code from nameplate / EnerVista",
         "Matches asset register entry"),
        ("Record Firmware Version (EnerVista: Status → Information → Main CPU)",
         "Matches approved firmware version"),
        ("Record Hardware Modules (Status → Information → Hardware Versions)",
         "All modules present and reporting correctly"),
        ("Front panel Status LEDs — Normal (green) illuminated",
         "Green LED solid = relay healthy and in service"),
    ])

    # 3.2 Device Status
    r += 2
    h1(ws, r, 1, 4, "3.2  RELAY SELF-TEST STATUS  (front panel: Status → Device Status)")
    r += 1
    note_row(ws, r, 1, 4,
             "Navigate: Status → Device Status.  All items should show OK / no faults.  "
             "If any WARNING or FAIL present, investigate before proceeding.  "
             "EnerVista: Actual Values → Status → Device Status.", 36)
    r += 1
    for col, hdr in enumerate(["Self-Test Item", "Normal Indication", "Status / Value", "Action if Fault"], 1):
        chdr(ws, r, col, hdr)

    status_items = [
        ("Normal LED (green, front panel)",
         "Solid green = no faults"),
        ("Warning LED (amber, front panel)",
         "OFF = no warnings"),
        ("Critical LED (red, front panel)",
         "OFF = no critical faults"),
        ("Device Status — no active errors",
         "Status → Device Status: all OK"),
        ("Main CPU self-test",
         "Status → Information → Main CPU: no errors"),
        ("Comms CPU self-test",
         "Status → Information → Comms CPU: no errors"),
        ("Hardware module health",
         "Status → Information → Hardware Versions: all modules present"),
        ("Power supply status",
         "Status → Device Status: PSU OK"),
        ("RTD module status  (if fitted)",
         "No open/shorted RTD alarms"),
        ("Analog I/O module status  (if fitted)",
         "Analog I/O module healthy"),
        ("Clock / PTP synchronisation",
         "Status → PTP Status or Clock: SNTP/IRIG-B locked"),
    ]
    for item, desc in status_items:
        r += 1
        label(ws, r, 1, item)
        label(ws, r, 2, desc, bg=CCR)
        inp(ws, r, 3); inp(ws, r, 4)

    r = check_section(ws, r, "3.3  CLOCK / TIME SYNCHRONISATION", [
        ("Front panel: Status → Clock — record relay date",    "Correct calendar date"),
        ("Front panel: Status → Clock — record relay time",    "Within 10 s of reference clock"),
        ("Time synchronisation source",                        "SNTP or IRIG-B (preferred over Internal)"),
        ("Synchronise if out of tolerance",                    "EnerVista: Setpoints → Device → Real-time Clock"),
    ])

    r = check_section(ws, r, "3.4  EVENT RECORDS (pre-test baseline)", [
        ("Front panel: Records → Events — review recent events",
         "No unexplained trips or alarms since last test"),
        ("Records → Fault Reports — review last fault report",
         "No unexpected generator trips"),
        ("Records → Transients — review last transient record",
         "No unexplained disturbances"),
        ("Note total event count for comparison at post-test",
         "Record event count before testing begins"),
    ])

    r = check_section(ws, r, "3.5  SETTINGS VERIFICATION", [
        ("EnerVista: download settings from relay (Get Settings from Device)",
         "Settings file received without error"),
        ("J1 CT primary/secondary (Setpoints → System → Current Sensing)",
         "Matches Cover sheet — J1 CT ratio"),
        ("K1 CT primary/secondary (Setpoints → System → Current Sensing)",
         "Matches Cover sheet — K1 CT ratio"),
        ("GND CT primary/secondary",
         "Matches Cover sheet — GND CT ratio"),
        ("VT primary/secondary (Setpoints → System → Voltage Sensing)",
         "Matches Cover sheet — Main VT ratio"),
        ("Generator rated MVA and voltage (Setpoints → System → Generator)",
         "Matches generator nameplate data"),
        ("87G enabled (Setpoints → Protection → Generator Elements)",
         "As per protection scheme"),
        ("64TN/27TN settings (Setpoints → Protection → Generator Elements)",
         "As per protection scheme"),
        ("Active Setpoint Group (Setpoints → Device → Setpoint Group)",
         "Group 1 or as per operating procedure"),
        ("Settings unchanged since last test (compare to approved file)",
         "No unauthorised changes"),
    ])

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 – ANALOGUE INPUTS
# ═══════════════════════════════════════════════════════════════════════════════
def build_analogue(wb):
    ws = wb.create_sheet("4. Analogue Inputs")
    ws.sheet_view.showGridLines = False
    col_widths = [22, 14, 13, 16, 16, 13, 10, 10, 8, 8, 8, 8, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    r = 1
    title_row(ws, r, 1, 13, "ANALOGUE INPUT VERIFICATION — PRIMARY VALUE CHECK")
    r += 1
    note_row(ws, r, 1, 13,
             "Verify the relay reports the correct PRIMARY value (Injected Secondary × CT or VT ratio). "
             "Front panel: Metering → Currents / Voltages.  "
             "EnerVista: Actual Values → Metering → Currents / Voltages.  "
             "All relay readings are in PRIMARY units.", 36)

    HDRS = ["Channel", "Injected Sec.\n(A or V)", "Inj. Angle\n(°)",
            "Expected Primary\n(auto)", "Relay Primary\nReading",
            "Relay Angle\nReading", "Mag\nErr %", "Angle\nErr °",
            "Mag\nTol %", "Ang\nTol °", "", "", "P/F"]

    def analogue_block(ws, r, title, ratio_name, tol_mag_name, channels, nav_note=None):
        r += 2
        h1(ws, r, 1, 13, title)
        r += 1
        nav = nav_note or f"Ratio: {ratio_name}  |  Tolerance: {tol_mag_name}  (defined names on Cover sheet)"
        label(ws, r, 1, nav, bg=CLB, cols=13)
        r += 1
        for col, hdr in enumerate(HDRS, 1):
            chdr(ws, r, col, hdr)
        for ch in channels:
            r += 1
            label(ws, r, 1, ch, bg=CLB)
            inp(ws, r, 2); inp(ws, r, 3)
            c = ws.cell(row=r, column=4,
                        value=f'=IFERROR(IF(B{r}="","",B{r}*{ratio_name}),"—")')
            c.fill = _fill(CGR); c.font = _font(bold=True)
            c.border = _thin(); c.alignment = _align("center", "center")
            inp(ws, r, 5); inp(ws, r, 6)
            c = ws.cell(row=r, column=7,
                        value=f'=IFERROR(IF(OR(D{r}="",E{r}=""),"",(E{r}-D{r})/D{r}*100),"—")')
            c.fill = _fill(CGR); c.border = _thin()
            c.number_format = "0.00"; c.alignment = _align("center", "center")
            c = ws.cell(row=r, column=8,
                        value=f'=IFERROR(IF(OR(C{r}="",F{r}=""),"",F{r}-C{r}),"—")')
            c.fill = _fill(CGR); c.border = _thin()
            c.number_format = "0.0"; c.alignment = _align("center", "center")
            c = ws.cell(row=r, column=9, value=f"={tol_mag_name}")
            c.fill = _fill(CY); c.border = _thin(); c.alignment = _align("center", "center")
            c = ws.cell(row=r, column=10, value="=TolAngle_deg")
            c.fill = _fill(CY); c.border = _thin(); c.alignment = _align("center", "center")
            blank(ws, r, 11); blank(ws, r, 12)
            pf = (f'=IFERROR(IF(AND(B{r}="",E{r}=""),"—",'
                  f'IF(OR(AND(ISNUMBER(G{r}),ABS(G{r})>{tol_mag_name}),'
                  f'AND(ISNUMBER(H{r}),ABS(H{r})>TolAngle_deg)),"FAIL","PASS")),"—")')
            c = ws.cell(row=r, column=13, value=pf)
            c.font = _font(bold=True); c.border = _thin(); c.alignment = _align("center", "center")
        return r

    r = analogue_block(ws, r,
        "4.1  J1 PHASE CURRENTS  (IA_J, IB_J, IC_J — neutral/generator-side CTs, Slot J Bank 1)  [Ratio_CT_J1]",
        "Ratio_CT_J1", "TolI_pct",
        ["IA_J — Phase A (neutral side)", "IB_J — Phase B (neutral side)", "IC_J — Phase C (neutral side)"],
        "Front panel: Metering → Currents  |  EnerVista: Actual Values → Metering → Currents  |  Ratio: Ratio_CT_J1")

    r = analogue_block(ws, r,
        "4.2  K1 PHASE CURRENTS  (IA_K, IB_K, IC_K — terminal/line-side CTs, Slot K Bank 1)  [Ratio_CT_K1]",
        "Ratio_CT_K1", "TolI_pct",
        ["IA_K — Phase A (terminal side)", "IB_K — Phase B (terminal side)", "IC_K — Phase C (terminal side)"],
        "Front panel: Metering → Currents  |  EnerVista: Actual Values → Metering → Currents  |  Ratio: Ratio_CT_K1")

    r = analogue_block(ws, r,
        "4.3  GROUND CURRENT  (IN — CBCT or zero-seq CT, Slot K)  [Ratio_CT_GND]",
        "Ratio_CT_GND", "TolI_pct",
        ["IN — Ground / neutral current"],
        "Front panel: Metering → Currents  |  EnerVista: Actual Values → Metering → Currents  |  Ratio: Ratio_CT_GND")

    r = analogue_block(ws, r,
        "4.4  PHASE VOLTAGES  (Va-n, Vb-n, Vc-n — Slot J VT inputs)  [Ratio_VT_main]",
        "Ratio_VT_main", "TolV_pct",
        ["Va-n — Phase A to neutral", "Vb-n — Phase B to neutral", "Vc-n — Phase C to neutral"],
        "Front panel: Metering → Voltages  |  EnerVista: Actual Values → Metering → Voltages  |  Ratio: Ratio_VT_main")

    r = analogue_block(ws, r,
        "4.5  AUXILIARY VOLTAGE (Vx — Slot J Bank 2 or open-delta VT)  [Ratio_VT_aux]  (complete if Vx fitted)",
        "Ratio_VT_aux", "TolV_pct",
        ["Vx — Auxiliary voltage input"],
        "Front panel: Metering → Voltages  |  EnerVista: Actual Values → Metering → Voltages  |  Ratio: Ratio_VT_aux")

    r = analogue_block(ws, r,
        "4.6  NEUTRAL VOLTAGE (Vn — neutral-connected VT for 64TN)  [Ratio_VT_neut]  (complete if neutral VT fitted)",
        "Ratio_VT_neut", "TolV_pct",
        ["Vn — Generator neutral voltage"],
        "Front panel: Metering → Voltages  |  EnerVista: Actual Values → Metering → Voltages  |  Ratio: Ratio_VT_neut")

    # 4.7 Differential meter check — 87G
    r += 2
    h1(ws, r, 1, 13,
       "4.7  DIFFERENTIAL METER CHECK — 87G  (front panel: Metering → Generator → Percent Differential Current)")
    r += 1
    note_row(ws, r, 1, 13,
             "Under balanced load: IOP (operate current) should be ≈ 0.00 pu; "
             "IRT (restraint current) ≈ load current in pu.  "
             "Accept IOP ≤ 0.05 pu under normal load conditions.  "
             "If IOP/IRT > 0.05, investigate CT polarity, ratio balance, and CT connections before testing.", 44)
    r += 1
    for col, hdr in enumerate(["Diff Element", "IOP\n(Operate, pu)", "IRT\n(Restraint, pu)",
                                "IOP/IRT Ratio\n(auto)", "Accept if\nIOP ≤ 0.05 pu",
                                "", "", "", "", "", "", "", "Result"], 1):
        chdr(ws, r, col, hdr)
    for elem in ["87G Phase A", "87G Phase B", "87G Phase C"]:
        r += 1
        label(ws, r, 1, elem, bg=CLB)
        inp(ws, r, 2); inp(ws, r, 3)
        c = ws.cell(row=r, column=4,
                    value=f'=IFERROR(IF(OR(B{r}="",C{r}=""),"",B{r}/C{r}),"—")')
        c.fill = _fill(CGR); c.border = _thin()
        c.number_format = "0.000"; c.alignment = _align("center", "center")
        label(ws, r, 5, "IOP ≤ 0.05 pu", bg=CCR)
        for col in range(6, 13): blank(ws, r, col)
        c = ws.cell(row=r, column=13,
                    value=f'=IFERROR(IF(B{r}="","—",IF(AND(ISNUMBER(B{r}),B{r}<=0.05),"PASS","FAIL")),"—")')
        c.font = _font(bold=True); c.border = _thin(); c.alignment = _align("center", "center")

    # 4.8 Differential meter check — 87O (optional)
    r += 2
    h1(ws, r, 1, 13,
       "4.8  DIFFERENTIAL METER CHECK — 87O  (Overall Gen-Xfmr differential — complete if 87O fitted)")
    r += 1
    note_row(ws, r, 1, 13,
             "Complete this section only if the relay is ordered with the 87O (Overall Differential) option.  "
             "Front panel: Metering → Generator → Overall Differential Current.  "
             "Accept IOP ≤ 0.05 pu under normal load conditions.", 36)
    r += 1
    for col, hdr in enumerate(["Diff Element", "IOP\n(Operate, pu)", "IRT\n(Restraint, pu)",
                                "IOP/IRT Ratio\n(auto)", "Accept if\nIOP ≤ 0.05 pu",
                                "", "", "", "", "", "", "", "Result"], 1):
        chdr(ws, r, col, hdr)
    for elem in ["87O Phase A", "87O Phase B", "87O Phase C"]:
        r += 1
        label(ws, r, 1, elem, bg=CLB)
        inp(ws, r, 2); inp(ws, r, 3)
        c = ws.cell(row=r, column=4,
                    value=f'=IFERROR(IF(OR(B{r}="",C{r}=""),"",B{r}/C{r}),"—")')
        c.fill = _fill(CGR); c.border = _thin()
        c.number_format = "0.000"; c.alignment = _align("center", "center")
        label(ws, r, 5, "IOP ≤ 0.05 pu", bg=CCR)
        for col in range(6, 13): blank(ws, r, col)
        c = ws.cell(row=r, column=13,
                    value=f'=IFERROR(IF(B{r}="","—",IF(AND(ISNUMBER(B{r}),B{r}<=0.05),"PASS","FAIL")),"—")')
        c.font = _font(bold=True); c.border = _thin(); c.alignment = _align("center", "center")

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 – BINARY I/O
# ═══════════════════════════════════════════════════════════════════════════════
def build_binary(wb):
    ws = wb.create_sheet("5. Binary IO")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", [10, 30, 22, 24, 12]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 5, "BINARY I/O VERIFICATION")

    r += 2
    h1(ws, r, 1, 5, "5.1  CONTACT INPUTS — Pickup Verification")
    r += 1
    note_row(ws, r, 1, 5,
             "Apply rated control voltage to each contact input terminal. "
             "Verify state change via front panel: Status → Contact Inputs  "
             "or EnerVista: Actual Values → Status → Contact Inputs.  "
             "Standard Slot F provides 7 inputs (Contact IN_1 to IN_7).  "
             "Optional Slot H provides up to 10 additional inputs. Add rows for fitted build.", 36)
    r += 1
    for col, hdr in enumerate(["Terminal", "Function Label\n(site-specific)", "Control Voltage Applied", "Relay Status Indication", "Result"], 1):
        chdr(ws, r, col, hdr)

    inputs = [
        ("Contact IN_1", "52a — Generator breaker closed  (edit to match schematics)"),
        ("Contact IN_2", "52b — Generator breaker open  (edit to match schematics)"),
        ("Contact IN_3", "Digital Input 3  (edit to match schematics)"),
        ("Contact IN_4", "Digital Input 4  (edit to match schematics)"),
        ("Contact IN_5", "Digital Input 5  (edit to match schematics)"),
        ("Contact IN_6", "Digital Input 6  (edit to match schematics)"),
        ("Contact IN_7", "Digital Input 7  (edit to match schematics)"),
    ]
    for term, func in inputs:
        r += 1
        label(ws, r, 1, term, bg=CLB, bold=True)
        inp(ws, r, 2, func)
        inp(ws, r, 3)
        label(ws, r, 4, "Status → Contact Inputs", bg=CCR)
        inp(ws, r, 5)

    r += 2
    h1(ws, r, 1, 5, "5.2  OUTPUT RELAYS — Continuity Test")
    r += 1
    note_row(ws, r, 1, 5,
             "Force each output relay via EnerVista: Setpoints → Testing → Output Relays  "
             "or front panel: Setpoints → Testing → Output Relays.  "
             "Verify contact continuity through wiring to trip coil / close coil / alarm.  "
             "Note: FA_3 (Critical Fail Relay, terminals 22–24) is normally energised — "
             "it de-energises on loss of control power or critical relay failure.", 40)
    r += 1
    for col, hdr in enumerate(["Terminal", "Function Label", "Force Method", "Expected Result", "P/F"], 1):
        chdr(ws, r, col, hdr)

    outputs = [
        ("FA_1  (Slot F terminals 1-3)",
         "TRIP",
         "EnerVista Testing → Output Relays → FA_1",
         "Form-A closes — continuity to trip coil verified"),
        ("FA_2  (Slot F terminals 4-6)",
         "CLOSE / AUX",
         "EnerVista Testing → Output Relays → FA_2",
         "Form-A closes — continuity verified"),
        ("FC_1  (Slot F terminals 7-9)",
         "AUX 3  (edit to match scheme)",
         "EnerVista Testing → Output Relays → FC_1",
         "Form-C operates — continuity verified"),
        ("FC_2  (Slot F terminals 10-12)",
         "AUX 4  (edit to match scheme)",
         "EnerVista Testing → Output Relays → FC_2",
         "Form-C operates — continuity verified"),
        ("FC_3  (Slot F terminals 22-24)",
         "CRITICAL FAIL RELAY  (normally energised)",
         "De-energise control power momentarily",
         "Form-C de-energises — critical fail contact opens"),
    ]
    for term, func, cmd, exp in outputs:
        r += 1
        label(ws, r, 1, term, bg=CLB, bold=True)
        inp(ws, r, 2, func)
        label(ws, r, 3, cmd, bg=CCR)
        label(ws, r, 4, exp, bg=CCR)
        inp(ws, r, 5)

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6 – ELEMENT SIMULATOR
# ═══════════════════════════════════════════════════════════════════════════════
def build_tracker(wb):
    ws = wb.create_sheet("6. Element Simulation")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [6, 16, 46, 14, 16, 12]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 6, "PROTECTION ELEMENT SIMULATION  (Built-in Simulation / External Test Set)")

    # ── Simulation procedure note ────────────────────────────────────────────
    r += 2
    h1(ws, r, 1, 6, "HOW TO USE THE BUILT-IN SIMULATION  (Setpoints → Testing → Simulation)")
    r += 1
    note_row(ws, r, 1, 6,
             "The 889 can simulate current and voltage phasors internally — no external test set needed for "
             "functional element verification.  When active, the relay suspends reading actual CT/VT inputs "
             "and substitutes the programmed values.  TEST MODE LED illuminates.  "
             "All simulation settings revert to Disabled at power-up (safe — cannot be left in test mode accidentally).", 40)
    r += 1
    sim_steps = [
        ("STEP 1 — Enter Pre-Fault state",
         "Setpoints → Testing → Simulation → Setup → Simulation State = Prefault State\n"
         "Relay now reads simulated inputs.  TEST MODE LED ON."),
        ("STEP 2 — Set Pre-Fault values",
         "Setpoints → Testing → Simulation → Pre-Fault\n"
         "Enter healthy load V (sec V) and I (× CT primary).  "
         "Values must be BELOW any trip setting.  "
         "Typical: Van=63.5 V∠0°, Vbn=63.5 V∠−120°, Vcn=63.5 V∠−240°, Ia=0.5×CTpri∠−30°  "
         "(angles entered −359.9° to 0.0° only — use −240° not +120°)"),
        ("STEP 3 — Set Fault values",
         "Setpoints → Testing → Simulation → Fault\n"
         "Enter values that should assert the element under test.  "
         "Example for 27P: reduce Van below pickup threshold.  "
         "Example for 51G: increase Ig above pickup setting."),
        ("STEP 4 — Trigger Fault state",
         "Setpoints → Testing → Simulation → Setup → Simulation State = Fault State\n"
         "Or assign a FlexLogic operand to Pre-Fault to Fault Trigger setting."),
        ("STEP 5 — Observe operation",
         "PICKUP LED asserts when element picks up.  TRIP LED + target message when element operates.  "
         "Relay auto-transitions to Postfault State after trip.  "
         "Verify operated element via Targets → Generator."),
        ("STEP 6 — Record and reset",
         "Record result in the table below.  Press RESET to clear latched targets/LEDs.  "
         "Return to Fault State to repeat test, or Pre-Fault for the next element."),
        ("STEP 7 — Restore when done",
         "Setpoints → Testing → Simulation → Setup → Simulation State = Disabled\n"
         "Confirm IN SERVICE LED returns to green.  Relay resumes reading actual CT/VT inputs."),
    ]
    for col, hdr in enumerate(["Step", "Procedure / Navigation Path", "", "", "", ""], 1):
        chdr(ws, r, col, hdr)
    for step, proc in sim_steps:
        r += 1
        ws.row_dimensions[r].height = 32
        label(ws, r, 1, step, bg=CLB, bold=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        c = ws.cell(row=r, column=2, value=proc)
        c.fill = _fill(CCR); c.font = _font(size=10)
        c.alignment = _align("left", "center", wrap=True); c.border = _thin()

    r += 2
    note_row(ws, r, 1, 6,
             "ELEMENTS NOT TESTABLE VIA BUILT-IN SIMULATION — use Omicron CMC or equivalent:  "
             "RTD-based elements (38 Bearing temp, 49 Thermal) — need RTD simulator or actual temperature source.  "
             "Bearing vibration (39) — needs dcmA signal source.  "
             "Arc Flash (50AF) — needs light/fiber sensor trigger.  "
             "For precise pickup accuracy and timing across all elements, Omicron CMC is recommended.  "
             "Simulation is suitable for functional verification and commissioning checks.", 44)

    # ── Element table ─────────────────────────────────────────────────────────
    r += 2
    for col, hdr in enumerate(["#", "Element ID", "Description", "Enabled\n(Y / N / NA)", "Test Method\n(Sim / Omicron)", "Result"], 1):
        chdr(ws, r, col, hdr)

    elements = [
        # Generator differential / stator earth fault
        ("87G",    "Generator Stator Differential  (Percent Differential, Slot J1 vs K1 CTs)"),
        ("87O",    "Overall Unit Differential  (Gen-Xfmr combined) — if 87O order option fitted"),
        ("87GD",   "Restricted Ground Fault  (Restricted Earth Fault)"),
        ("64TN",   "100% Stator Ground — 3rd Harmonic Voltage Differential  (requires neutral VT)"),
        ("27TN",   "Third Harmonic Neutral Undervoltage  (part of 64TN scheme)"),
        # Inadvertent energisation
        ("50/27",  "Inadvertent Energisation  (50 O/C + 27 UV combined logic)"),
        # Loss of excitation / field
        ("40-1",   "Loss of Excitation — Circle 1  (negative mho impedance circle)"),
        ("40-2",   "Loss of Excitation — Circle 2  (negative mho impedance circle)"),
        ("40Q",    "Reactive Power  (absorbing leading vars — loss of excitation alarm/trip)"),
        ("76",     "Excitation Current Protection  (dcmA field current input — if fitted)"),
        # Unbalance / thermal
        ("46",     "Generator Unbalance  (I²t negative-sequence current, trip)"),
        ("46A",    "Generator Unbalance  (I²t negative-sequence current, alarm)"),
        ("49",     "Thermal Protection  (RTD temperature — requires RTD module) ⚑ Omicron / RTD sim required"),
        ("49TOL",  "Thermal Overload  (IEC hot/cold curve — if monitoring option ordered)"),
        ("38",     "Bearing Overtemperature  (RTD — requires RTD module) ⚑ Omicron / RTD sim required"),
        ("39",     "Bearing Vibration  (dcmA transducer input — if Slot G/L I/O fitted) ⚑ dcmA source required"),
        # Frequency / voltage elements
        ("24D1",   "Volts/Hz — Definite Time element 1  (over-excitation)"),
        ("24D2",   "Volts/Hz — Definite Time element 2  (over-excitation)"),
        ("24C",    "Volts/Hz — Inverse Time curve  (over-excitation)"),
        ("47",     "Phase Reversal  (negative sequence voltage detection)"),
        ("27P",    "Phase Undervoltage  (any one, two, or all three phases)"),
        ("27X",    "Auxiliary Undervoltage  (Vx input — if Aux VT fitted)"),
        ("59P",    "Phase Overvoltage  (fundamental phasor magnitude)"),
        ("59N",    "Neutral Overvoltage  (calculated 3V0 or measured)"),
        ("59X",    "Auxiliary Overvoltage  (Vx input — if Aux VT fitted)"),
        ("59_2",   "Negative Sequence Overvoltage  (V2)"),
        ("81U",    "Underfrequency  (multiple stages available)"),
        ("81O",    "Overfrequency  (multiple stages available)"),
        ("81R",    "Frequency Rate of Change  (df/dt)"),
        ("81A",    "Frequency Out-of-Band accumulation"),
        # Overcurrent / backup
        ("51V",    "Voltage Restrained Time Overcurrent  (backup to 87G)"),
        ("67P",    "Phase Directional Overcurrent  (90° quadrature connection)"),
        ("50P",    "Phase Instantaneous Overcurrent  (Slots J or K)"),
        ("51N",    "Neutral Time Overcurrent"),
        ("50N",    "Neutral Instantaneous Overcurrent"),
        ("67N",    "Neutral Directional Overcurrent  (V0 or Vx polarised)"),
        ("51G",    "Ground Time Overcurrent  (CT bank ground input)"),
        ("50G",    "Ground Instantaneous Overcurrent"),
        ("67G",    "Ground Directional Overcurrent  (V0 or Vx polarised)"),
        ("51SG",   "Sensitive Ground Time Overcurrent  (CBCT/50:0.025 input)"),
        ("50SG",   "Sensitive Ground Instantaneous Overcurrent  (CBCT input)"),
        ("67SG",   "Sensitive Ground Directional Overcurrent"),
        ("50_2",   "Negative Sequence Instantaneous Overcurrent  (I2)"),
        ("50OFL",  "Offline Overcurrent  (generator not connected to system)"),
        ("50OL",   "Overload Alarm  (phase current RMS)"),
        # Power / impedance
        ("32",     "Directional Power  (reverse / low forward power)"),
        ("78",     "Out-of-Step  (single blinder + mho supervisory)"),
        # Control / breaker
        ("25",     "Synchrocheck  (frequency, angle, and voltage difference)"),
        ("50BF",   "Breaker Failure  (3-pole, current supervised)"),
        ("VTFF",   "VT Fuse Failure  (loss of potential supervision)"),
    ]

    for i, (eid, desc) in enumerate(elements, 1):
        r += 1
        bg = CAL if i % 2 == 0 else CWH
        c = ws.cell(row=r, column=1, value=i)
        c.font = _font(bold=True); c.border = _thin(); c.fill = _fill(bg)
        c.alignment = _align("center", "center")
        label(ws, r, 2, eid, bg=CLB, bold=True)
        label(ws, r, 3, desc, bg=bg)
        inp(ws, r, 4); inp(ws, r, 5); inp(ws, r, 6)

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 7 – ELEMENT RESULTS
# ═══════════════════════════════════════════════════════════════════════════════
def build_results(wb):
    ws = wb.create_sheet("7. Element Results")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFG", [12, 42, 14, 14, 12, 14, 10]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 7, "PROTECTION ELEMENT TEST RESULTS")
    r += 1
    note_row(ws, r, 1, 7,
             "Setting column = relay pickup setting (secondary A, V, pu, or %). "
             "Measured Pickup = value at which relay output asserted (from Omicron or test set). "
             "Pickup Err % = auto-calculated. Result = PASS if |Err| ≤ TolPickup_pct.", 32)
    r += 1
    for col, hdr in enumerate(["Element", "Description",
                                "Pickup Setting\n(sec A, V, Hz, or pu)",
                                "Measured\nPickup",
                                "Pickup\nErr %",
                                "Measured\nTime (s)", "Result"], 1):
        chdr(ws, r, col, hdr)

    test_elements = [
        ("87G",    "Stator differential restrained pickup  (% of rated / pu CT secondary)"),
        ("87G U",  "Stator differential unrestrained pickup  (pu CT secondary)"),
        ("87O",    "Overall differential restrained pickup  (if 87O fitted, pu)"),
        ("87GD",   "Restricted Ground Fault pickup  (A secondary)"),
        ("64TN",   "100% Stator Ground 3rd harmonic pickup  (V secondary)"),
        ("27TN",   "3rd Harmonic Neutral Undervoltage pickup  (V secondary)"),
        ("50/27",  "Inadvertent Energisation — 50P threshold  (A secondary)"),
        ("40-1",   "Loss of Excitation circle 1 diameter  (Ω secondary)"),
        ("40-2",   "Loss of Excitation circle 2 diameter  (Ω secondary)"),
        ("40Q",    "Reactive Power pickup  (kvar or × rated)"),
        ("76",     "Excitation Current Protection pickup  (mA DC)"),
        ("46",     "Unbalance I²t pickup — trip  (% FLA or A negative-seq)"),
        ("49TOL",  "Thermal Overload pickup  (× FLA)"),
        ("38",     "Bearing RTD trip temperature  (°C)"),
        ("24D1",   "Volts/Hz element 1 pickup  (pu — V/Hz relative to rated)"),
        ("24D2",   "Volts/Hz element 2 pickup  (pu)"),
        ("24C",    "Volts/Hz inverse curve pickup  (pu)"),
        ("47",     "Phase Reversal pickup  (% negative-sequence voltage)"),
        ("27P",    "Phase Undervoltage pickup  (V secondary or × VT)"),
        ("27X",    "Auxiliary Undervoltage pickup  (V secondary)"),
        ("59P",    "Phase Overvoltage pickup  (V secondary or × VT)"),
        ("59N",    "Neutral Overvoltage pickup  (V secondary)"),
        ("59X",    "Auxiliary Overvoltage pickup  (V secondary)"),
        ("59_2",   "Negative Sequence Overvoltage pickup  (V secondary)"),
        ("81U",    "Underfrequency pickup  (Hz)"),
        ("81O",    "Overfrequency pickup  (Hz)"),
        ("81R",    "Rate of Change of Frequency pickup  (Hz/s)"),
        ("51V",    "Voltage Restrained TOC pickup  (A secondary)"),
        ("50P",    "Phase Instantaneous OC pickup  (A secondary)"),
        ("67P",    "Phase Directional OC pickup  (A secondary)"),
        ("51N",    "Neutral TOC pickup  (A secondary)"),
        ("50N",    "Neutral Instantaneous OC pickup  (A secondary)"),
        ("51G",    "Ground TOC pickup  (A secondary)"),
        ("50G",    "Ground Instantaneous OC pickup  (A secondary)"),
        ("51SG",   "Sensitive Ground TOC pickup  (A secondary)"),
        ("50SG",   "Sensitive Ground Instantaneous OC pickup  (A secondary)"),
        ("50_2",   "Negative Sequence Instantaneous OC pickup  (A secondary)"),
        ("32",     "Directional Power pickup  (kW or × rated MW)"),
        ("78",     "Out-of-Step forward reach  (Ω secondary)"),
        ("25",     "Synchrocheck maximum angle difference  (degrees)"),
        ("50BF",   "Breaker Failure timer delay  (s)"),
    ]

    for i, (eid, desc) in enumerate(test_elements, 1):
        r += 1
        bg = CAL if i % 2 == 0 else CWH
        label(ws, r, 1, eid, bg=CLB, bold=True)
        label(ws, r, 2, desc, bg=bg)
        inp(ws, r, 3); inp(ws, r, 4)
        c = ws.cell(row=r, column=5,
                    value=f'=IFERROR(IF(OR(C{r}="",D{r}=""),"",(D{r}-C{r})/C{r}*100),"—")')
        c.fill = _fill(CGR); c.border = _thin()
        c.number_format = "0.00"; c.alignment = _align("center", "center")
        inp(ws, r, 6)
        result = (f'=IFERROR(IF(AND(C{r}="",D{r}=""),"—",'
                  f'IF(AND(ISNUMBER(E{r}),ABS(E{r})<=TolPickup_pct),"PASS","FAIL")),"—")')
        c = ws.cell(row=r, column=7, value=result)
        c.font = _font(bold=True); c.border = _thin(); c.alignment = _align("center", "center")

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 8 – POST-TEST
# ═══════════════════════════════════════════════════════════════════════════════
def build_posttest(wb):
    ws = wb.create_sheet("8. Post-Test")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [40, 34, 14, 20]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 4, "POST-TEST CHECKS")

    def check_section(ws, r, title, rows):
        r += 2
        h1(ws, r, 1, 4, title)
        r += 1
        for col, hdr in enumerate(["Check", "Expected / Action", "Result", "Notes"], 1):
            chdr(ws, r, col, hdr)
        for chk, exp in rows:
            r += 1
            label(ws, r, 1, chk)
            label(ws, r, 2, exp, bg=CCR)
            inp(ws, r, 3); inp(ws, r, 4)
        return r

    r = check_section(ws, r, "8.1  SETTINGS RESTORE & VERIFICATION", [
        ("EnerVista: Get Settings from Device — compare to approved file",
         "Settings match approved EnerVista .8SS settings file"),
        ("All forced testing outputs restored (EnerVista → Testing → Output Relays)",
         "All output relays returned to normal logic control"),
        ("Active Setpoint Group restored to normal operating group",
         "Setpoints → Device → Setpoint Group: Group 1 or as applicable"),
        ("Normal LED (green) illuminated — relay in service",
         "Green LED solid, no warning/critical LEDs active"),
    ])

    r = check_section(ws, r, "8.2  FINAL DEVICE STATUS CHECK", [
        ("Status → Device Status: no active errors",
         "All self-test items OK — no faults remaining"),
        ("Status → Information: firmware version unchanged",
         "Version matches pre-test record"),
        ("Normal LED ON; Warning and Critical LEDs OFF",
         "Relay healthy and in service"),
    ])

    # 8.3 Clear test data
    r += 2
    h1(ws, r, 1, 4, "8.3  CLEAR TEST DATA  (front panel or EnerVista)")
    r += 1
    note_row(ws, r, 1, 4,
             "Clear records generated during testing to prevent test data contaminating "
             "operational event records. Confirm with engineer before clearing.", 28)
    r += 1
    for col, hdr in enumerate(["Action", "Navigation Path", "Completed", "Notes"], 1):
        chdr(ws, r, col, hdr)
    for action, path in [
        ("Clear Event Records",
         "Front panel: Records → Clear Records → Events  |  EnerVista: Records → Clear Records"),
        ("Clear Transient Records  (if required by procedure)",
         "Front panel: Records → Clear Records → Transients"),
        ("Clear Fault Reports  (if required by procedure)",
         "Front panel: Records → Clear Records → Fault Reports"),
        ("Clear Data Logger  (if required by procedure)",
         "Front panel: Records → Clear Records → Data Logger"),
    ]:
        r += 1
        label(ws, r, 1, action, bg=CLB, bold=True)
        label(ws, r, 2, path, bg=CCR)
        inp(ws, r, 3); inp(ws, r, 4)

    r = check_section(ws, r, "8.4  PLANT RESTORATION", [
        ("Remove test connections from CT / VT secondary terminals",
         "Field CTs and VTs reconnected — no open CT circuits"),
        ("Remove CT secondary shorting links",
         "All CT circuits reconnected"),
        ("Restore control voltage wiring",
         "All control circuits normal — trip and close circuits healthy"),
        ("Verify relay measuring load current (Metering → Currents)",
         "Values agree with substation metering — both CT banks plausible"),
        ("87G differential check — IOP ≤ 0.05 pu under load",
         "Metering → Generator → Percent Differential Current"),
        ("DCS / Modbus TCP comms: 800XA faceplate shows live values",
         "Values updating in 800XA — no comms alarms"),
        ("Cancel all outstanding LOTO permits",
         "All permits cleared and returned"),
        ("Notify operations — relay returned to service",
         "Operations personnel informed"),
    ])

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 9 – FILES & SIGN-OFF
# ═══════════════════════════════════════════════════════════════════════════════
def build_signoff(wb, signed_row):
    ws = wb.create_sheet("9. Files & Sign-Off")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [28, 38, 20, 16]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, 1, 4, "FILES INDEX & SIGN-OFF")

    r += 2
    h1(ws, r, 1, 4, "OVERALL TEST RESULT (from Cover sheet)")
    r += 1
    label(ws, r, 1, "Signed Overall Result", bold=True)
    c = ws.cell(row=r, column=2, value=f"='1. Cover'!B{signed_row}")
    c.font = _font(bold=True, size=13); c.alignment = _align("center", "center")
    c.border = _thin(); c.fill = _fill(CY)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)

    r += 2
    h1(ws, r, 1, 4, "TEST ARTEFACTS INDEX")
    r += 1
    for col, hdr in enumerate(["Artefact", "File Name / SharePoint Location", "Date Saved", "Notes"], 1):
        chdr(ws, r, col, hdr)
    for art in ["Relay settings file (EnerVista .8SS project file)",
                "Pre-test event records (Records → Events export)",
                "Post-test event records",
                "Completed results template (this file)",
                "Oscillography / transient records",
                "Fault reports",
                "Photographs / site notes"]:
        r += 1
        label(ws, r, 1, art)
        inp(ws, r, 2); inp(ws, r, 3); inp(ws, r, 4)

    r += 2
    h1(ws, r, 1, 4, "DEFECTS / FINDINGS LOG")
    r += 1
    for col, hdr in enumerate(["#", "Defect Description", "Action Taken / Work Order", "Closed"], 1):
        chdr(ws, r, col, hdr)
    for i in range(1, 7):
        r += 1
        c = ws.cell(row=r, column=1, value=i)
        c.border = _thin(); c.alignment = _align("center", "center")
        for col in range(2, 5): inp(ws, r, col)

    r += 2
    h1(ws, r, 1, 4, "SIGN-OFF")
    r += 1
    for col, hdr in enumerate(["Role", "Name (print)", "Signature", "Date"], 1):
        chdr(ws, r, col, hdr)
    for role in ["Test Engineer", "Witness (Operations)", "Approving Engineer"]:
        r += 1
        ws.row_dimensions[r].height = 28
        label(ws, r, 1, role, bg=CLB, bold=True)
        inp(ws, r, 2); inp(ws, r, 3); inp(ws, r, 4)

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# DEFINED NAMES
# ═══════════════════════════════════════════════════════════════════════════════
def add_names(wb,
              ct_j1_pri, ct_j1_sec, ct_k1_pri, ct_k1_sec,
              ct_gnd_pri, ct_gnd_sec,
              vt_main_pri, vt_main_sec, vt_aux_pri, vt_aux_sec,
              vt_neut_pri, vt_neut_sec, tol_rows):
    sn = "'1. Cover'"
    defs = {
        "Ratio_CT_J1":    f"{sn}!$C${ct_j1_pri}/{sn}!$C${ct_j1_sec}",
        "Ratio_CT_K1":    f"{sn}!$C${ct_k1_pri}/{sn}!$C${ct_k1_sec}",
        "Ratio_CT_GND":   f"{sn}!$C${ct_gnd_pri}/{sn}!$C${ct_gnd_sec}",
        "Ratio_VT_main":  f"{sn}!$C${vt_main_pri}*1000/{sn}!$C${vt_main_sec}",
        "Ratio_VT_aux":   f"{sn}!$C${vt_aux_pri}*1000/{sn}!$C${vt_aux_sec}",
        "Ratio_VT_neut":  f"{sn}!$C${vt_neut_pri}*1000/{sn}!$C${vt_neut_sec}",
        "TolV_pct":       f"{sn}!$B${tol_rows['Voltage Magnitude Tolerance']}",
        "TolI_pct":       f"{sn}!$B${tol_rows['Current Magnitude Tolerance']}",
        "TolAngle_deg":   f"{sn}!$B${tol_rows['Angle Tolerance']}",
        "TolPickup_pct":  f"{sn}!$B${tol_rows['Pickup Tolerance']}",
        "TolTime_pct":    f"{sn}!$B${tol_rows['Timing Tolerance']}",
    }
    for name, attr in defs.items():
        dn = DefinedName(name=name, attr_text=attr)
        wb.defined_names[name] = dn


# ═══════════════════════════════════════════════════════════════════════════════
# CONDITIONAL FORMATTING
# ═══════════════════════════════════════════════════════════════════════════════
def add_cf(wb):
    pf = PatternFill("solid", fgColor=CPF)
    ff = PatternFill("solid", fgColor=CFL)
    pfont = Font(bold=True, color=CPFF, name="Calibri")
    ffont = Font(bold=True, color=CFFL, name="Calibri")
    for sheet_name, rng in [("4. Analogue Inputs", "M1:M300"),
                             ("7. Element Results",  "G1:G200"),
                             ("6. Element Simulation", "F1:F200")]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.conditional_formatting.add(rng, CellIsRule("equal", ['"PASS"'], fill=pf, font=pfont))
            ws.conditional_formatting.add(rng, CellIsRule("equal", ['"FAIL"'], fill=ff, font=ffont))


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()

    (cover_ws, tol_rows,
     ct_j1_pri, ct_j1_sec, ct_k1_pri, ct_k1_sec,
     ct_gnd_pri, ct_gnd_sec,
     vt_main_pri, vt_main_sec, vt_aux_pri, vt_aux_sec,
     vt_neut_pri, vt_neut_sec, signed_row) = build_cover(wb)

    build_equip(wb)
    build_pretest(wb)
    build_analogue(wb)
    build_binary(wb)
    build_tracker(wb)
    build_results(wb)
    build_posttest(wb)
    build_signoff(wb, signed_row)

    add_names(wb,
              ct_j1_pri, ct_j1_sec, ct_k1_pri, ct_k1_sec,
              ct_gnd_pri, ct_gnd_sec,
              vt_main_pri, vt_main_sec, vt_aux_pri, vt_aux_sec,
              vt_neut_pri, vt_neut_sec, tol_rows)
    add_cf(wb)

    for ws in wb.worksheets:
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = "portrait"
        ws.page_margins.left = 0.4
        ws.page_margins.right = 0.4
        ws.page_margins.top = 0.6
        ws.page_margins.bottom = 0.6

    out = "/home/michael/claudeProject /protection-relay-checksheet/GE-889_Test_Results_Template.xlsx"
    wb.save(out)
    print(f"Saved: {out}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Signed result row: {signed_row}")
    print(f"CT J1 rows: C{ct_j1_pri} / C{ct_j1_sec}")
    print(f"CT K1 rows: C{ct_k1_pri} / C{ct_k1_sec}")
    print(f"CT GND rows: C{ct_gnd_pri} / C{ct_gnd_sec}")
    print(f"VT main rows: C{vt_main_pri} / C{vt_main_sec}")
    print(f"Defined names: {list(wb.defined_names.keys())}")


if __name__ == "__main__":
    main()
