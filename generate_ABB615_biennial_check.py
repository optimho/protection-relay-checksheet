#!/usr/bin/env python3
"""Generate ABB Relion 615 Series Biennial (2-Yearly) Visual Check Sheet (.xlsx).
   Covers: REF615 (feeder), RET615 (transformer), REM615 (motor), REU615 (voltage/freq).
   Tailored from Technical Manual 1YHT530004D05 Rev E (product version 5.0 FP1).
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ── Colours ──────────────────────────────────────────────────────────────────
CB  = "1F4E79"; CLB = "D9E2F3"; CY  = "FFF2CC"; CCR = "FFF8E7"
CWH = "FFFFFF"; CAL = "F2F2F2"; CUI = "0000FF"
CGRN = "E8F5E9"   # light green for healthy indicator

NCOLS = 5

def _fill(c): return PatternFill("solid", fgColor=c)
def _font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def title_row(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = _fill(CB)
    c.font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28

def h1(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = _fill(CB); c.font = _font(bold=True, color="FFFFFF", size=10)
    c.alignment = _align("left", "center"); c.border = _thin()
    ws.row_dimensions[row].height = 18

def note(ws, row, text, bg=CY, height=32):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = _fill(bg); c.font = _font(italic=True, size=9)
    c.alignment = _align("left", "center", wrap=True); c.border = _thin()
    ws.row_dimensions[row].height = height

def lbl(ws, row, col, text, bg=CLB, bold=False, span=1, wrap=False):
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    c = ws.cell(row=row, column=col, value=text)
    c.fill = _fill(bg); c.font = _font(bold=bold)
    c.alignment = _align("left", "center", wrap=wrap); c.border = _thin()

def inp(ws, row, col, value=None, span=1, center=False, bg=CWH):
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    c = ws.cell(row=row, column=col, value=value)
    c.fill = _fill(bg); c.font = _font(color=CUI)
    c.alignment = _align("center" if center else "left", "center"); c.border = _thin()

def chdr(ws, row, col, text, span=1):
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    c = ws.cell(row=row, column=col, value=text)
    c.fill = _fill(CB); c.font = _font(bold=True, color="FFFFFF", size=9)
    c.alignment = _align("center", "center", wrap=True); c.border = _thin()
    ws.row_dimensions[row].height = 28


def build_check_sheet(wb):
    ws = wb.active
    ws.title = "615 Biennial Check"
    ws.sheet_view.showGridLines = False
    ws.page_setup.paperSize  = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage  = True
    ws.page_setup.fitToHeight = 0
    ws.page_setup.fitToWidth  = 1
    ws.page_margins.left = 0.5; ws.page_margins.right = 0.5
    ws.page_margins.top  = 0.6; ws.page_margins.bottom = 0.6

    for col, w in zip("ABCDE", [30, 34, 14, 14, 12]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r,
              "ABB RELION 615 SERIES — BIENNIAL (2-YEARLY) VISUAL CHECK SHEET  "
              "(REF615 / RET615 / REM615 / REU615)")

    # ── 1. IED / SITE INFORMATION ────────────────────────────────────────────
    r += 2
    h1(ws, r, "1.  IED AND SITE INFORMATION")
    for lbl_text, default in [
        ("IED Model",              "REF615  /  RET615  /  REM615  /  REU615  (circle one)"),
        ("IED Configuration (A/B/C…)", "e.g. REF615 C — from nameplate"),
        ("Product Version",         "From LHMI: Information → Version  (e.g. 5.0 FP1)"),
        ("Protected Equipment",     ""),
        ("Site / Substation",       ""),
        ("IED Name / Bay",          "From LHMI home screen header"),
        ("Panel / Cubicle Tag",     ""),
        ("Work Order No.",          ""),
        ("Date of Check",           ""),
        ("Technician Name",         ""),
        ("Next Check Due",          "(2 years from date above)"),
    ]:
        r += 1
        lbl(ws, r, 1, lbl_text, bold=True)
        inp(ws, r, 2, default, span=4)

    # ── 2. CONNECTION METHOD ─────────────────────────────────────────────────
    r += 2
    h1(ws, r, "2.  CONNECTION METHOD")
    r += 1
    note(ws, r,
         "OPTION A — LHMI only (no laptop): Navigate with keypad (ESC, arrows, Enter, Clear). "
         "Press any key to wake the display. Home screen shows the single-line diagram (SLD).\n"
         "OPTION B — Web HMI (WHMI) via laptop: Connect laptop Ethernet to front RJ-45 port (100Base-TX). "
         "Open web browser → navigate to IED IP address (find it at LHMI: Configuration → Communication → Ethernet). "
         "Login with VIEWER or OPERATOR credentials. "
         "NOTE: WHMI is disabled by default — confirm with site engineers before use.\n"
         "OPTION C — PCM600: Connect via front RJ-45 or substation Ethernet. "
         "Open PCM600, select relay in Plant Structure, use Online monitoring for live data.",
         bg=CLB, height=64)
    r += 1
    lbl(ws, r, 1, "Connection method used today", bold=True)
    inp(ws, r, 2, "LHMI only  /  Web HMI (WHMI)  /  PCM600  (delete as applicable)", span=4)

    # ── 3. LED STATUS CHECK ───────────────────────────────────────────────────
    r += 2
    h1(ws, r, "3.  FRONT PANEL LED STATUS  (visual check before connecting)")
    r += 1
    note(ws, r,
         "The 615 series has three fixed protection indicator LEDs above the display:  "
         "READY (green) — START (yellow) — TRIP (red).  "
         "Additionally, 11 programmable alarm LEDs on the right side of the display.  "
         "IMPORTANT: READY LED flashing green = Internal fault (protection disabled). "
         "READY LED solid green = IED healthy and in service.",
         bg=CLB, height=44)
    r += 1
    chdr(ws, r, 1, "LED / Indicator")
    chdr(ws, r, 2, "Expected State (healthy IED)")
    chdr(ws, r, 3, "Observed State")
    chdr(ws, r, 4, "Result\n(OK / FAULT)")
    chdr(ws, r, 5, "Notes")
    for led, expected, note_text in [
        ("READY LED (green)",
         "Solid green — continuous, not flashing",
         ""),
        ("START LED (yellow)",
         "Off (or on if a protection element is picked up under load — may be normal)",
         ""),
        ("TRIP LED (red)",
         "Off (steady) — on only if latched trip not yet cleared",
         ""),
        ("Programmable alarm LEDs (right side of display, LEDs 1–11)",
         "All off, or known/accepted alarm conditions only",
         ""),
        ("LHMI display — any 'Internal Fault' message visible?",
         "No Internal Fault message on display",
         ""),
        ("LHMI display — any 'Warning' message visible?",
         "No Warning message, or known/accepted warning",
         ""),
    ]:
        r += 1
        lbl(ws, r, 1, led)
        lbl(ws, r, 2, expected, bg=CCR, wrap=True)
        inp(ws, r, 3)
        inp(ws, r, 4, center=True)
        inp(ws, r, 5)

    # ── 4. SELF-SUPERVISION CHECK ────────────────────────────────────────────
    r += 2
    h1(ws, r, "4.  SELF-SUPERVISION  (Internal Faults and Warnings)")
    r += 1
    note(ws, r,
         "INTERNAL FAULT: Ready LED flashes. Protection is DISABLED. "
         "LHMI shows 'Internal Fault' with a numeric code (e.g. Code 80 = RAM error, Code 116 = COM card error). "
         "Record the code — relay requires restart or ABB support.\n"
         "WARNING: Ready LED stays solid green. Protection continues (functions affected by the fault may be impaired). "
         "LHMI shows 'Warning' with a name and code (e.g. Code 20 = IEC 61850 error, Code 35 = Comm. channel down). "
         "Record the code — warning message can be manually cleared after investigating.",
         bg=CY, height=60)
    r += 1
    chdr(ws, r, 1, "Check Item")
    chdr(ws, r, 2, "How to Check / LHMI Path")
    chdr(ws, r, 3, "Expected")
    chdr(ws, r, 4, "Result / Code")
    chdr(ws, r, 5, "Action")
    for chk, how, exp in [
        ("Any 'Internal Fault' text on LHMI display?",
         "LHMI home screen or press ESC to return to top level — fault message has highest LHMI priority",
         "No Internal Fault message"),
        ("If fault: record fault code and description",
         "Code is displayed alongside 'Internal Fault' message  (Table 25 in Tech Manual for code list)",
         "N/A if no fault"),
        ("Any 'Warning' text on LHMI display?",
         "LHMI home screen — Warning message appears with name and code",
         "No Warning message"),
        ("If warning: record warning code and name",
         "Code displayed alongside 'Warning' message  (Table 26 in Tech Manual for code list)",
         "N/A if no warning"),
        ("IRF (Internal fault relay) contact — X100 terminals 3-5",
         "Measure continuity across X100 terminals 3-5 (closed-circuit principle: CLOSED = healthy, OPEN = fault)",
         "Contact CLOSED (continuity present)"),
    ]:
        r += 1
        lbl(ws, r, 1, chk)
        lbl(ws, r, 2, how, bg=CCR, wrap=True)
        lbl(ws, r, 3, exp, bg=CCR)
        inp(ws, r, 4)
        inp(ws, r, 5)

    # ── 5. ANALOGUE MEASUREMENTS (PRIMARY VALUES) ─────────────────────────────
    r += 2
    h1(ws, r, "5.  ANALOGUE MEASUREMENTS  (Primary Values — relay reports in primary units)")
    r += 1
    note(ws, r,
         "LHMI: Navigate to home screen (SLD). Live values may be shown on SLD (IL2, U12, P, Q).  "
         "For full measurement view: LHMI → Monitoring → Measurements → Current measurements (or Voltage measurements).  "
         "WHMI: Monitoring → Measurements (same data available via browser).  "
         "Record measured primary values. Flag anything obviously abnormal (e.g. one phase zero while others are loaded).",
         bg=CLB, height=44)
    r += 1
    chdr(ws, r, 1, "Quantity")
    chdr(ws, r, 2, "LHMI Path")
    chdr(ws, r, 3, "Measured Value")
    chdr(ws, r, 4, "Unit")
    chdr(ws, r, 5, "Notes")
    measurements = [
        ("IL1 — Phase L1 Current (primary)", "Monitoring → Measurements → Current measurements → IL1", "A primary"),
        ("IL2 — Phase L2 Current (primary)", "Monitoring → Measurements → Current measurements → IL2", "A primary"),
        ("IL3 — Phase L3 Current (primary)", "Monitoring → Measurements → Current measurements → IL3", "A primary"),
        ("I0 / IN — Residual Current",        "Monitoring → Measurements → Current measurements → I0",  "A primary"),
        ("UL1-L2 — Line Voltage (primary)",   "Monitoring → Measurements → Voltage measurements → U12\n(if VT fitted)", "kV primary"),
        ("UL2-L3 — Line Voltage (primary)",   "Monitoring → Measurements → Voltage measurements → U23\n(if VT fitted)", "kV primary"),
        ("UL3-L1 — Line Voltage (primary)",   "Monitoring → Measurements → Voltage measurements → U31\n(if VT fitted)", "kV primary"),
        ("U0 — Residual Voltage",             "Monitoring → Measurements → Voltage measurements → U0\n(if VT fitted)", "V primary"),
        ("Frequency",                          "Monitoring → Measurements → Voltage measurements → f\n(if VT fitted)", "Hz"),
        ("P — Active Power",                   "Monitoring → Measurements → Power measurements → P\n(if VT fitted)", "kW primary"),
        ("Q — Reactive Power",                 "Monitoring → Measurements → Power measurements → Q\n(if VT fitted)", "kVAr primary"),
    ]
    for qty, path, unit in measurements:
        r += 1
        lbl(ws, r, 1, qty)
        lbl(ws, r, 2, path, bg=CCR, wrap=True)
        inp(ws, r, 3, center=True)
        lbl(ws, r, 4, unit, bg=CCR)
        inp(ws, r, 5)

    # 5 spare rows for additional quantities
    for i in range(5):
        r += 1
        bg = CAL if i % 2 == 0 else CWH
        lbl(ws, r, 1, "", bg=bg)
        lbl(ws, r, 2, "Monitoring → Measurements", bg=CCR)
        inp(ws, r, 3, center=True, bg=bg)
        inp(ws, r, 4, bg=bg)
        inp(ws, r, 5, bg=bg)

    # ── 6. CLOCK / TIME CHECK ─────────────────────────────────────────────────
    r += 2
    h1(ws, r, "6.  CLOCK AND TIME SYNCHRONISATION")
    r += 1
    note(ws, r,
         "LHMI: Configuration → Time → Date and Time (to view current date/time).  "
         "Time sync source: Configuration → Time → Synch source (SNTP, IRIG-B, IEEE 1588 v2, Modbus, IEC 103, DNP or None).  "
         "48-hour capacitor backup keeps RTC running during power outages.  "
         "If sync source = 'None', time is free-running and may drift — set manually if >10 s out.",
         bg=CLB, height=44)
    r += 1
    chdr(ws, r, 1, "Check")
    chdr(ws, r, 2, "LHMI Path")
    chdr(ws, r, 3, "IED Value")
    chdr(ws, r, 4, "Reference")
    chdr(ws, r, 5, "Difference / OK?")
    for chk, path, ref in [
        ("IED Date",
         "Configuration → Time → Date",
         "Today's date"),
        ("IED Time",
         "Configuration → Time → Time",
         "Current time (GPS/phone reference)"),
        ("Time sync source",
         "Configuration → Time → Synch source",
         "SNTP / IRIG-B / IEEE 1588 preferred over 'None'"),
        ("Time sync status (GNRLTMS function block)",
         "Monitoring → I/O status → GNRLTMS → ALARM / WARNING output signals",
         "ALARM = False, WARNING = False"),
        ("Time difference (accept ≤ 10 s if internal; ≤ 1 s if SNTP/IRIG-B)",
         "Compare LHMI time display to GPS/phone reference",
         "≤ 10 s"),
    ]:
        r += 1
        lbl(ws, r, 1, chk)
        lbl(ws, r, 2, path, bg=CCR, wrap=True)
        inp(ws, r, 3)
        lbl(ws, r, 4, ref, bg=CCR)
        inp(ws, r, 5)

    # ── 7. RECENT EVENT LOG ───────────────────────────────────────────────────
    r += 2
    h1(ws, r, "7.  RECENT EVENTS  (record most recent events — flag any unexplained trips or alarms)")
    r += 1
    note(ws, r,
         "LHMI: Navigate to Events (press Events button or use menu → Events → Event list). "
         "Events are stored newest first (FIFO — 1024 process events, 2048 audit trail events in nonvolatile memory). "
         "WHMI: Events menu at top of screen — same data, easier to scroll.  "
         "PCM600: Event Viewer for full event list including audit trail.  "
         "Flag any unexplained TRIP, OPERATE or FAULT events for investigation.",
         bg=CLB, height=40)
    r += 1
    chdr(ws, r, 1, "Event #\n(most recent = 1)")
    chdr(ws, r, 2, "Date")
    chdr(ws, r, 3, "Time")
    chdr(ws, r, 4, "Event Description / Function", span=2)
    for i in range(1, 13):
        r += 1
        bg = CAL if i % 2 == 0 else CWH
        c = ws.cell(row=r, column=1, value=i)
        c.font = _font(bold=True); c.border = _thin(); c.fill = _fill(bg)
        c.alignment = _align("center", "center")
        inp(ws, r, 2, bg=bg); inp(ws, r, 3, bg=bg)
        inp(ws, r, 4, span=2, bg=bg)

    r += 1
    lbl(ws, r, 1, "Total events in event list (approximate)", bold=True)
    inp(ws, r, 2, span=4)

    # ── 8. DCS / 800XA COMMUNICATION CHECK ───────────────────────────────────
    r += 2
    h1(ws, r, "8.  DCS / ABB 800XA COMMUNICATION CHECK")
    r += 1
    note(ws, r,
         "Open the IED faceplate in the ABB 800XA control system. "
         "Verify live values (currents, IED status) are updating and not frozen. "
         "The 615 series communicates via IEC 61850 (primary), IEC 60870-5-103 or Modbus to the 800XA.  "
         "Check for any yellow or red communication alarm objects on the 800XA faceplate.  "
         "Record the 800XA object path for reference.",
         bg=CLB, height=40)
    r += 1
    chdr(ws, r, 1, "Check")
    chdr(ws, r, 2, "Description")
    chdr(ws, r, 3, "Result\n(Y / N)")
    chdr(ws, r, 4, "Notes", span=2)
    for chk, desc in [
        ("800XA faceplate opened",
         "Navigate to IED faceplate in 800XA Plant Explorer"),
        ("IED Status = HEALTHY / OK on faceplate",
         "Status indicator green — not alarm/fault colour"),
        ("Live phase current values updating",
         "Measured currents on faceplate match LHMI readings (not frozen or stale)"),
        ("No IEC 61850 / Modbus communication alarm",
         "No yellow or red comms alarm objects on 800XA faceplate"),
        ("Record 800XA object path / faceplate name",
         "e.g. 'Site / Substation / Bay / Relay faceplate'"),
    ]:
        r += 1
        lbl(ws, r, 1, chk, bold=True)
        lbl(ws, r, 2, desc, bg=CCR)
        inp(ws, r, 3, center=True)
        inp(ws, r, 4, span=2)

    # ── 9. OPTIONAL — CLEAR EVENTS ────────────────────────────────────────────
    r += 2
    h1(ws, r, "9.  OPTIONAL — CLEAR EVENTS / INDICATIONS  (only if required by site procedure)")
    r += 1
    note(ws, r,
         "Requires ENGINEER or ADMINISTRATOR access level. "
         "Download / record all events BEFORE clearing. "
         "LHMI: Clear menu (bottom of main menu or dedicated Clear button) → Event list → Clear all.  "
         "WHMI: Events menu → select events → Clear.  "
         "Trip/start indications: LHMI Clear menu → Trip indications / Start indications.  "
         "WARNING indications: can be manually cleared from LHMI display when Warning message is shown.",
         bg=CY, height=44)
    r += 1
    chdr(ws, r, 1, "Action")
    chdr(ws, r, 2, "LHMI / WHMI Path")
    chdr(ws, r, 3, "Performed?\n(Y / N / N/A)")
    chdr(ws, r, 4, "Notes", span=2)
    for action, path in [
        ("Download event list before clearing",
         "WHMI: Events → Export, or PCM600: Event Viewer → Save"),
        ("Clear event list",
         "LHMI: Clear → Event list → Clear all  (requires ENGINEER access)"),
        ("Clear trip / start LED indications",
         "LHMI: Clear → Trip indications  OR press Clear button on keypad"),
        ("Clear warning indication (if present)",
         "Press any key when Warning message displayed, or LHMI: Clear → Indications"),
        ("Clear disturbance records (if required)",
         "LHMI: Clear → Disturbance records  (requires ENGINEER access)"),
    ]:
        r += 1
        lbl(ws, r, 1, action, bold=True)
        lbl(ws, r, 2, path, bg=CCR, wrap=True)
        inp(ws, r, 3, center=True)
        inp(ws, r, 4, span=2)

    # ── 10. OBSERVATIONS ──────────────────────────────────────────────────────
    r += 2
    h1(ws, r, "10.  OBSERVATIONS AND CORRECTIVE ACTIONS")
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r+4, end_column=NCOLS)
    c = ws.cell(row=r, column=1); c.fill = _fill(CWH); c.border = _thin()
    c.alignment = _align("left", "top", wrap=True)
    for rr in range(r+1, r+5):
        ws.row_dimensions[rr].height = 18
        for col in range(1, NCOLS+1):
            ws.cell(row=rr, column=col).border = _thin()
    r += 5

    # ── 11. SIGN-OFF ───────────────────────────────────────────────────────────
    r += 1
    h1(ws, r, "11.  SIGN-OFF")
    r += 1
    chdr(ws, r, 1, "Role"); chdr(ws, r, 2, "Name (print)")
    chdr(ws, r, 3, "Signature"); chdr(ws, r, 4, "Date", span=2)
    for role in ["Technician", "Supervisor / Witness"]:
        r += 1
        ws.row_dimensions[r].height = 24
        lbl(ws, r, 1, role, bold=True)
        inp(ws, r, 2); inp(ws, r, 3); inp(ws, r, 4, span=2)

    return ws


def main():
    wb = openpyxl.Workbook()
    build_check_sheet(wb)

    out = ("/home/michael/claudeProject /protection-relay-checksheet/"
           "ABB_615_Biennial_Visual_Check.xlsx")
    wb.save(out)
    print(f"Saved: {out}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
