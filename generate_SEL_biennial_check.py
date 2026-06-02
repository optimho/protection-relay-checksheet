#!/usr/bin/env python3
"""Generate SEL Relay Biennial (2-Yearly) Visual Check Sheet (.xlsx).
   Covers: SEL-300G, SEL-311L, SEL-787, SEL-700G
   All use the same serial ASCII interface — one generic check form.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# ── Colours ──────────────────────────────────────────────────────────────────
CB  = "1F4E79"   # deep blue   – headers
CLB = "D9E2F3"   # light blue  – label / instruction cells
CY  = "FFF2CC"   # yellow      – caution / optional
CCR = "FFF8E7"   # cream       – reference text
CGR = "E8F5E9"   # light green – OK / Pass reference
CPF = "C6EFCE"; CPFF = "006100"
CFL = "FFC7CE"; CFFL = "9C0006"
CWH = "FFFFFF"
CAL = "F2F2F2"
CUI = "0000FF"

def _fill(c): return PatternFill("solid", fgColor=c)
def _font(bold=False, color="000000", size=10, italic=False, name="Calibri"):
    return Font(bold=bold, color=color, size=size, italic=italic, name=name)
def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)
def _thick_bottom():
    return Border(bottom=Side(style="medium"))

NCOLS = 5   # A-E

def title_row(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = _fill(CB)
    c.font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28

def h1(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = _fill(CB)
    c.font = _font(bold=True, color="FFFFFF", size=10)
    c.alignment = _align("left", "center")
    c.border = _thin()
    ws.row_dimensions[row].height = 18

def note(ws, row, text, bg=CY, height=28):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = _fill(bg)
    c.font = _font(italic=True, size=9)
    c.alignment = _align("left", "center", wrap=True)
    c.border = _thin()
    ws.row_dimensions[row].height = height

def lbl(ws, row, col, text, bg=CLB, bold=False, span=1, wrap=False, size=10):
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    c = ws.cell(row=row, column=col, value=text)
    c.fill = _fill(bg)
    c.font = _font(bold=bold, size=size)
    c.alignment = _align("left", "center", wrap=wrap)
    c.border = _thin()

def inp(ws, row, col, value=None, span=1, center=False, bg=CWH, size=10):
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    c = ws.cell(row=row, column=col, value=value)
    c.fill = _fill(bg)
    c.font = _font(color=CUI, size=size)
    c.alignment = _align("center" if center else "left", "center")
    c.border = _thin()

def chdr(ws, row, col, text, span=1):
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    c = ws.cell(row=row, column=col, value=text)
    c.fill = _fill(CB)
    c.font = _font(bold=True, color="FFFFFF", size=9)
    c.alignment = _align("center", "center", wrap=True)
    c.border = _thin()
    ws.row_dimensions[row].height = 28


def build_check_sheet(wb):
    ws = wb.active
    ws.title = "SEL Biennial Check"
    ws.sheet_view.showGridLines = False
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 0
    ws.page_setup.fitToWidth = 1
    ws.page_margins.left = 0.5; ws.page_margins.right = 0.5
    ws.page_margins.top = 0.6;  ws.page_margins.bottom = 0.6

    # Column widths  A    B    C    D    E
    for col, w in zip("ABCDE", [30, 32, 14, 14, 12]):
        ws.column_dimensions[col].width = w

    r = 1
    title_row(ws, r, "SEL PROTECTION RELAY — BIENNIAL (2-YEARLY) VISUAL CHECK SHEET")

    # ── 1. RELAY / SITE INFORMATION ─────────────────────────────────────────
    r += 2                                              # 3
    h1(ws, r, "1.  RELAY AND SITE INFORMATION")
    fields = [
        ("Relay Model",         "SEL-300G  /  SEL-311L  /  SEL-787  /  SEL-700G  (circle one)"),
        ("Protected Equipment", ""),
        ("Site / Substation",   ""),
        ("Relay ID (RID / TID)",""),
        ("Panel / Cubicle Tag", ""),
        ("Work Order No.",      ""),
        ("Date of Check",       ""),
        ("Technician Name",     ""),
        ("Next Check Due",      "(2 years from date above)"),
    ]
    for lbl_text, default in fields:
        r += 1
        lbl(ws, r, 1, lbl_text, bold=True)
        inp(ws, r, 2, default, span=4)

    # ── 2. CONNECTION METHOD ─────────────────────────────────────────────────
    r += 2                                              # ~14
    h1(ws, r, "2.  CONNECTION METHOD")
    r += 1
    note(ws, r,
         "OPTION A — Front panel only (no laptop): Press any key to wake display. Navigate with arrow keys.\n"
         "OPTION B — Serial cable: Connect SEL-C662 cable (laptop USB → relay front DB-9). "
         "Open PuTTY/terminal: 19200 baud, 8-N-1, no flow control. Press Enter → prompt appears. "
         "Type ACC <Enter>, enter Level 1 password (default: OTTER).\n"
         "OPTION C — Ethernet (if fitted): Open AcSELerator QuickSet or SSH/Telnet to relay IP.",
         bg=CLB, height=52)
    r += 1
    lbl(ws, r, 1, "Connection method used today", bold=True)
    # Dropdown for connection method
    inp(ws, r, 2, "Front panel / Serial cable / Ethernet  (delete as applicable)", span=4)

    # ── 3. SELF-TEST / DEVICE STATUS ────────────────────────────────────────
    r += 2
    h1(ws, r, "3.  SELF-TEST AND DEVICE STATUS")
    r += 1
    note(ws, r,
         "Front panel: ENABLED LED (green, solid) must be illuminated. "
         "Serial: type STA <Enter> or STATUS <Enter> — all items should show OK. "
         "W = warning (issue STA C to clear), FAIL = fault (requires investigation).",
         bg=CLB, height=36)
    r += 1
    chdr(ws, r, 1, "Check Item")
    chdr(ws, r, 2, "How to Read / Where to Look")
    chdr(ws, r, 3, "Expected")
    chdr(ws, r, 4, "Result")
    chdr(ws, r, 5, "Notes")

    status_checks = [
        ("ENABLED LED (green, solid)",
         "Visual — front panel top-right LED",
         "ON (solid green)"),
        ("No alarm LEDs active",
         "Visual — front panel LEDs",
         "All alarm LEDs off"),
        ("Front panel display",
         "Visual — press any key to wake",
         "No error messages shown"),
        ("Self-test status — all OK",
         "Serial: STA <Enter>\nFront panel: Main > Status",
         "All items = OK"),
        ("HALARM bit (hardware alarm)",
         "STA output: look for HALARM = 0",
         "0 (not asserted)"),
        ("SALARM bit (software alarm)",
         "STA output: look for SALARM = 0",
         "0 (not asserted)"),
        ("Relay shows 'Relay Enabled'",
         "Last line of STA output",
         "'Relay Enabled'"),
    ]
    for chk, how, exp in status_checks:
        r += 1
        lbl(ws, r, 1, chk)
        lbl(ws, r, 2, how, bg=CCR, wrap=True)
        lbl(ws, r, 3, exp, bg=CCR)
        inp(ws, r, 4, center=True)
        inp(ws, r, 5)

    # ── 4. ANALOGUE MEASUREMENTS (PRIMARY VALUES) ────────────────────────────
    r += 2
    h1(ws, r, "4.  ANALOGUE MEASUREMENTS  (Primary Values — relay reports in primary units)")
    r += 1
    note(ws, r,
         "Serial: type METER <Enter> — relay displays primary values (A and V). "
         "Front panel: Main > Meters > Fundamental.  "
         "Record measured values. Compare to expected load (no specific pass/fail for 2-yearly — "
         "flag anything obviously wrong, e.g., one phase zero when others are normal).",
         bg=CLB, height=40)
    r += 1
    chdr(ws, r, 1, "Quantity")
    chdr(ws, r, 2, "Reading Method")
    chdr(ws, r, 3, "Measured Value")
    chdr(ws, r, 4, "Unit")
    chdr(ws, r, 5, "Notes")

    measurements = [
        ("IA — Phase A Current (primary)",  "METER / Meters > Fundamental", "", "A primary"),
        ("IB — Phase B Current (primary)",  "METER / Meters > Fundamental", "", "A primary"),
        ("IC — Phase C Current (primary)",  "METER / Meters > Fundamental", "", "A primary"),
        ("IN / IGX — Residual / Neutral",   "METER / Meters > Fundamental", "", "A primary"),
        ("VA — Phase A Voltage (primary)",  "METER / Meters > Fundamental\n(if VT fitted)", "", "kV primary"),
        ("VB — Phase B Voltage (primary)",  "METER / Meters > Fundamental\n(if VT fitted)", "", "kV primary"),
        ("VC — Phase C Voltage (primary)",  "METER / Meters > Fundamental\n(if VT fitted)", "", "kV primary"),
        ("Frequency",                       "METER / Meters > Fundamental", "", "Hz"),
    ]
    for qty, method, val, unit in measurements:
        r += 1
        lbl(ws, r, 1, qty)
        lbl(ws, r, 2, method, bg=CCR, wrap=True)
        inp(ws, r, 3, val, center=True)
        lbl(ws, r, 4, unit, bg=CCR)
        inp(ws, r, 5)

    # 6 spare rows for additional analogue quantities (e.g. Y-side CTs on 700G/300G,
    # differential quantities, neutral voltages, power, etc.)
    for i in range(6):
        r += 1
        bg = CAL if i % 2 == 0 else CWH
        lbl(ws, r, 1, "", bg=bg)                       # user fills in quantity name
        lbl(ws, r, 2, "METER / Meters > Fundamental", bg=CCR)
        inp(ws, r, 3, center=True, bg=bg)
        inp(ws, r, 4, bg=bg)                           # user fills in unit
        inp(ws, r, 5, bg=bg)

    # ── 5. CLOCK / TIME CHECK ────────────────────────────────────────────────
    r += 2
    h1(ws, r, "5.  CLOCK AND TIME SYNCHRONISATION")
    r += 1
    chdr(ws, r, 1, "Check")
    chdr(ws, r, 2, "How to Read")
    chdr(ws, r, 3, "Relay Value")
    chdr(ws, r, 4, "Reference Value")
    chdr(ws, r, 5, "Difference / Result")

    time_checks = [
        ("Relay Date",
         "Serial: DAT <Enter>\nFront panel: Main > Time or Status",
         "", "Today's date"),
        ("Relay Time",
         "Serial: TIM <Enter>\nFront panel: Main > Time or Status",
         "", "GPS/NTP reference"),
        ("Time source (IRIG-B / NTP / Internal)",
         "Serial: STA <Enter> — look for IRIG-B or NTP line",
         "", "IRIG-B preferred"),
        ("Time difference (accept ≤ 10 s)",
         "Relay time vs phone/GPS reference",
         "", "≤ 10 s"),
    ]
    for chk, how, relay_val, ref in time_checks:
        r += 1
        lbl(ws, r, 1, chk)
        lbl(ws, r, 2, how, bg=CCR, wrap=True)
        inp(ws, r, 3, relay_val, center=True)
        lbl(ws, r, 4, ref, bg=CCR)
        inp(ws, r, 5)

    # ── 6. RECENT EVENT LOG ───────────────────────────────────────────────────
    r += 2
    h1(ws, r, "6.  RECENT EVENTS  (record most recent events — flag any unexplained trips or alarms)")
    r += 1
    note(ws, r,
         "Serial: type HIS <Enter> for event history, or EVE <Enter> for last event report, "
         "or SER <Enter> for Sequential Events Recorder.  "
         "Front panel: Main > Events (or Main > Targets for relay word bits).  "
         "Record the most recent events and their dates below.",
         bg=CLB, height=36)
    r += 1
    chdr(ws, r, 1, "Event #\n(most recent = 1)")
    chdr(ws, r, 2, "Date")
    chdr(ws, r, 3, "Time")
    chdr(ws, r, 4, "Event Description / Type", span=2)
    for i in range(1, 13):
        r += 1
        bg = CAL if i % 2 == 0 else CWH
        c = ws.cell(row=r, column=1, value=i)
        c.font = _font(bold=True); c.border = _thin(); c.fill = _fill(bg)
        c.alignment = _align("center", "center")
        inp(ws, r, 2, bg=bg)
        inp(ws, r, 3, bg=bg)
        inp(ws, r, 4, span=2, bg=bg)

    r += 1
    lbl(ws, r, 1, "Total events since last check (approx.)", bold=True)
    inp(ws, r, 2, span=4)

    # ── 7. DCS / 800XA COMMUNICATION CHECK ──────────────────────────────────
    r += 2
    h1(ws, r, "7.  DCS / 800XA COMMUNICATION CHECK")
    r += 1
    note(ws, r,
         "Open the relay faceplate in ABB 800XA. Verify that live values (currents, status) "
         "are updating and not frozen. Check for any communication alarm objects (yellow/red). "
         "Record the 800XA object path or faceplate name below.",
         bg=CLB, height=36)
    r += 1
    chdr(ws, r, 1, "Check")
    chdr(ws, r, 2, "Description")
    chdr(ws, r, 3, "Result\n(Y / N)")
    chdr(ws, r, 4, "Notes", span=2)

    dcs_checks = [
        ("800XA faceplate opened",
         "Navigate to relay faceplate in 800XA control system"),
        ("Live values updating",
         "Phase currents / status visible and changing with load — not frozen"),
        ("No communication alarm",
         "No yellow or red communication alarm objects on faceplate"),
        ("Record 800XA object path / faceplate name",
         "e.g. \"Site/Substation/Relay Tag/Faceplate\""),
    ]
    for chk, desc in dcs_checks:
        r += 1
        lbl(ws, r, 1, chk, bold=True)
        lbl(ws, r, 2, desc, bg=CCR)
        inp(ws, r, 3, center=True)
        inp(ws, r, 4, span=2)

    # ── 8. OPTIONAL — CLEAR EVENT BUFFERS ───────────────────────────────────
    r += 2
    h1(ws, r, "8.  OPTIONAL — CLEAR EVENT BUFFERS  (only if required by your procedure)")
    r += 1
    note(ws, r,
         "If your site procedure requires clearing event buffers after recording, "
         "issue the following serial commands at Access Level 1 or 2. "
         "Ensure all events have been downloaded/recorded BEFORE clearing.",
         bg=CY, height=32)
    r += 1
    chdr(ws, r, 1, "Serial Command")
    chdr(ws, r, 2, "Action")
    chdr(ws, r, 3, "Performed?\n(Y / N / N/A)")
    chdr(ws, r, 4, "Notes", span=2)
    for cmd, action in [
        ("SUM R <Enter>", "Reset event report and Summary Command buffers"),
        ("SER R <Enter>", "Reset Sequential Events Record buffer"),
        ("HIS C <Enter>", "Clear event history (if required — confirm with supervisor)"),
    ]:
        r += 1
        lbl(ws, r, 1, cmd, bg=CLB, bold=True)
        lbl(ws, r, 2, action, bg=CCR)
        inp(ws, r, 3, center=True)
        inp(ws, r, 4, span=2)

    # ── 9. OBSERVATIONS ──────────────────────────────────────────────────────
    r += 2
    h1(ws, r, "9.  OBSERVATIONS AND CORRECTIVE ACTIONS")
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r+4, end_column=NCOLS)
    c = ws.cell(row=r, column=1)
    c.fill = _fill(CWH); c.border = _thin()
    c.alignment = _align("left", "top", wrap=True)
    ws.row_dimensions[r].height = 18
    for rr in range(r+1, r+5):
        ws.row_dimensions[rr].height = 18
        for col in range(1, NCOLS+1):
            ws.cell(row=rr, column=col).border = _thin()
    r += 5

    # ── 10. SIGN-OFF ─────────────────────────────────────────────────────────
    r += 1
    h1(ws, r, "10.  SIGN-OFF")
    r += 1
    chdr(ws, r, 1, "Role")
    chdr(ws, r, 2, "Name (print)")
    chdr(ws, r, 3, "Signature")
    chdr(ws, r, 4, "Date", span=2)
    for role in ["Technician", "Supervisor / Witness"]:
        r += 1
        ws.row_dimensions[r].height = 24
        lbl(ws, r, 1, role, bold=True)
        inp(ws, r, 2); inp(ws, r, 3); inp(ws, r, 4, span=2)

    return ws


def main():
    wb = openpyxl.Workbook()
    build_check_sheet(wb)

    out = "/home/michael/claudeProject /protection-relay-checksheet/SEL_Biennial_Visual_Check.xlsx"
    wb.save(out)
    print(f"Saved: {out}")


if __name__ == "__main__":
    main()
